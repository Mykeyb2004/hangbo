from __future__ import annotations

import argparse
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DEFAULT_SHEET_NAME = "问卷数据"
PHASE_MARKER_COLUMN_INDEX = 3
PHASE_MARKER_PATTERN = re.compile(r"^\s*(?:第)?(?:[一二三四五六七八九十百千万两0-9]+)期\s*$")

STATUS_UPDATED = "updated"
STATUS_ALREADY_PROCESSED = "already_processed"
STATUS_NO_PHASE_MARKER = "no_phase_marker"
STATUS_INSUFFICIENT_COLUMNS = "insufficient_columns"
STATUS_MISSING_SHEET = "missing_sheet"
STATUS_FILE_MISSING = "file_missing"
STATUS_READ_ERROR = "read_error"
STATUS_SAVE_ERROR = "save_error"


@dataclass(frozen=True)
class PhaseMarkerColumnMatch:
    column_index: int
    column_letter: str
    column_name: str
    matched_markers: tuple[str, ...]


@dataclass(frozen=True)
class PhaseColumnPreprocessResult:
    path: Path
    sheet_name: str
    status: str
    matched_markers: tuple[str, ...] = ()
    detected_phase_columns: tuple[PhaseMarkerColumnMatch, ...] = ()
    error_message: str | None = None


@dataclass(frozen=True)
class PhaseColumnPreprocessSummary:
    results: tuple[PhaseColumnPreprocessResult, ...]

    def _results_by_status(self, *statuses: str) -> tuple[PhaseColumnPreprocessResult, ...]:
        return tuple(item for item in self.results if item.status in set(statuses))

    @property
    def updated_results(self) -> tuple[PhaseColumnPreprocessResult, ...]:
        return self._results_by_status(STATUS_UPDATED)

    @property
    def updated_count(self) -> int:
        return len(self.updated_results)

    @property
    def already_processed_results(self) -> tuple[PhaseColumnPreprocessResult, ...]:
        return self._results_by_status(STATUS_ALREADY_PROCESSED)

    @property
    def already_processed_count(self) -> int:
        return len(self.already_processed_results)

    @property
    def no_phase_marker_results(self) -> tuple[PhaseColumnPreprocessResult, ...]:
        return self._results_by_status(STATUS_NO_PHASE_MARKER)

    @property
    def no_phase_marker_count(self) -> int:
        return len(self.no_phase_marker_results)

    @property
    def insufficient_columns_results(self) -> tuple[PhaseColumnPreprocessResult, ...]:
        return self._results_by_status(STATUS_INSUFFICIENT_COLUMNS)

    @property
    def insufficient_columns_count(self) -> int:
        return len(self.insufficient_columns_results)

    @property
    def no_action_count(self) -> int:
        return (
            self.already_processed_count
            + self.no_phase_marker_count
            + self.insufficient_columns_count
        )

    @property
    def failed_results(self) -> tuple[PhaseColumnPreprocessResult, ...]:
        return self._results_by_status(
            STATUS_FILE_MISSING,
            STATUS_MISSING_SHEET,
            STATUS_READ_ERROR,
            STATUS_SAVE_ERROR,
        )

    @property
    def failed_count(self) -> int:
        return len(self.failed_results)


def is_phase_marker_value(value: object) -> bool:
    if value is None:
        return False
    return PHASE_MARKER_PATTERN.fullmatch(str(value).strip()) is not None


def find_phase_marker_values_in_column(worksheet, column_index: int) -> tuple[str, ...]:
    if worksheet.max_column < column_index:
        return ()

    matched_markers: list[str] = []
    seen_markers: set[str] = set()
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if not is_phase_marker_value(value):
            continue
        normalized_value = str(value).strip()
        if normalized_value in seen_markers:
            continue
        matched_markers.append(normalized_value)
        seen_markers.add(normalized_value)
    return tuple(matched_markers)



def find_phase_marker_values_in_third_column(worksheet) -> tuple[str, ...]:
    return find_phase_marker_values_in_column(worksheet, PHASE_MARKER_COLUMN_INDEX)


def worksheet_has_phase_marker_in_third_column(worksheet) -> bool:
    return bool(find_phase_marker_values_in_third_column(worksheet))


def sheet_has_phase_marker_in_third_column(
    input_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> bool:
    if not input_path.exists():
        return False

    try:
        workbook = load_workbook(input_path, keep_vba=input_path.suffix.lower() == ".xlsm")
    except Exception:
        return False

    try:
        if sheet_name not in workbook.sheetnames:
            return False
        return worksheet_has_phase_marker_in_third_column(workbook[sheet_name])
    finally:
        workbook.close()


def move_worksheet_column_to_end(worksheet, source_column_index: int = PHASE_MARKER_COLUMN_INDEX) -> None:
    max_column = worksheet.max_column
    max_row = worksheet.max_row
    target_column_index = max_column + 1
    source_letter = get_column_letter(source_column_index)
    target_letter = get_column_letter(target_column_index)
    source_dimension = worksheet.column_dimensions[source_letter]
    target_dimension = worksheet.column_dimensions[target_letter]

    if source_dimension.width is not None:
        target_dimension.width = source_dimension.width
    if source_dimension.hidden:
        target_dimension.hidden = source_dimension.hidden
    if source_dimension.bestFit:
        target_dimension.bestFit = source_dimension.bestFit
    if source_dimension.outline_level:
        target_dimension.outline_level = source_dimension.outline_level

    for row_index in range(1, max_row + 1):
        source_cell = worksheet.cell(row=row_index, column=source_column_index)
        target_cell = worksheet.cell(row=row_index, column=target_column_index)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

    worksheet.delete_cols(source_column_index, 1)


def normalize_column_name(value: object, column_letter: str) -> str:
    text = str(value).strip() if value is not None else ""
    if text:
        return text
    return "未命名列"


def collect_phase_marker_columns(
    worksheet,
    exclude_column_indexes: tuple[int, ...] = (),
) -> tuple[PhaseMarkerColumnMatch, ...]:
    excluded = set(exclude_column_indexes)
    matches: list[PhaseMarkerColumnMatch] = []
    for column_index in range(1, worksheet.max_column + 1):
        if column_index in excluded:
            continue
        matched_markers = find_phase_marker_values_in_column(worksheet, column_index)
        if not matched_markers:
            continue
        column_letter = get_column_letter(column_index)
        column_name = normalize_column_name(
            worksheet.cell(row=1, column=column_index).value,
            column_letter,
        )
        matches.append(
            PhaseMarkerColumnMatch(
                column_index=column_index,
                column_letter=column_letter,
                column_name=column_name,
                matched_markers=matched_markers,
            )
        )
    return tuple(matches)


def combine_markers_from_matches(matches: tuple[PhaseMarkerColumnMatch, ...]) -> tuple[str, ...]:
    combined: list[str] = []
    seen: set[str] = set()
    for match in matches:
        for marker in match.matched_markers:
            if marker in seen:
                continue
            combined.append(marker)
            seen.add(marker)
    return tuple(combined)


def build_updated_notice(input_path: Path, sheet_name: str) -> str:
    return (
        "已执行输入文件预处理："
        f"{input_path.name} / {sheet_name} 的第三列检测到“X期”标记，"
        "已移动到最后一列并保存。"
    )


def process_phase_column_workbook(
    input_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> PhaseColumnPreprocessResult:
    if not input_path.exists() or not input_path.is_file():
        return PhaseColumnPreprocessResult(
            path=input_path,
            sheet_name=sheet_name,
            status=STATUS_FILE_MISSING,
        )

    try:
        workbook = load_workbook(input_path, keep_vba=input_path.suffix.lower() == ".xlsm")
    except Exception as exc:
        return PhaseColumnPreprocessResult(
            path=input_path,
            sheet_name=sheet_name,
            status=STATUS_READ_ERROR,
            error_message=str(exc),
        )

    try:
        if sheet_name not in workbook.sheetnames:
            return PhaseColumnPreprocessResult(
                path=input_path,
                sheet_name=sheet_name,
                status=STATUS_MISSING_SHEET,
            )

        worksheet = workbook[sheet_name]
        if worksheet.max_column < PHASE_MARKER_COLUMN_INDEX:
            return PhaseColumnPreprocessResult(
                path=input_path,
                sheet_name=sheet_name,
                status=STATUS_INSUFFICIENT_COLUMNS,
            )

        matched_markers = find_phase_marker_values_in_third_column(worksheet)
        if not matched_markers:
            detected_phase_columns = collect_phase_marker_columns(
                worksheet,
                exclude_column_indexes=(PHASE_MARKER_COLUMN_INDEX,),
            )
            if detected_phase_columns:
                return PhaseColumnPreprocessResult(
                    path=input_path,
                    sheet_name=sheet_name,
                    status=STATUS_ALREADY_PROCESSED,
                    matched_markers=combine_markers_from_matches(detected_phase_columns),
                    detected_phase_columns=detected_phase_columns,
                )
            return PhaseColumnPreprocessResult(
                path=input_path,
                sheet_name=sheet_name,
                status=STATUS_NO_PHASE_MARKER,
            )

        move_worksheet_column_to_end(worksheet, PHASE_MARKER_COLUMN_INDEX)
        try:
            workbook.save(input_path)
        except Exception as exc:
            return PhaseColumnPreprocessResult(
                path=input_path,
                sheet_name=sheet_name,
                status=STATUS_SAVE_ERROR,
                matched_markers=matched_markers,
                error_message=str(exc),
            )

        return PhaseColumnPreprocessResult(
            path=input_path,
            sheet_name=sheet_name,
            status=STATUS_UPDATED,
            matched_markers=matched_markers,
        )
    finally:
        workbook.close()


def preprocess_phase_column_if_needed(
    input_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> str | None:
    result = process_phase_column_workbook(input_path, sheet_name)
    if result.status == STATUS_UPDATED:
        return build_updated_notice(input_path, sheet_name)
    if result.status in {
        STATUS_ALREADY_PROCESSED,
        STATUS_NO_PHASE_MARKER,
        STATUS_INSUFFICIENT_COLUMNS,
        STATUS_MISSING_SHEET,
        STATUS_FILE_MISSING,
    }:
        return None
    if result.status == STATUS_READ_ERROR:
        raise OSError(f"读取 Excel 失败：{input_path}（{result.error_message or '未知错误'}）")
    if result.status == STATUS_SAVE_ERROR:
        raise OSError(
            f"保存预处理后的 Excel 失败：{input_path}（{result.error_message or '未知错误'}）"
        )
    raise ValueError(f"未知预处理状态：{result.status}")


def format_start_message(input_path: Path, sheet_name: str) -> str:
    return f"[INFO] 开始检查文件：{input_path} / {sheet_name}"


def format_markers(markers: tuple[str, ...]) -> str:
    return "、".join(markers)


def format_phase_column_match(match: PhaseMarkerColumnMatch) -> str:
    return (
        f"{match.column_name}"
        f"（第{match.column_letter}列，示例值：{format_markers(match.matched_markers)}）"
    )


def format_phase_column_matches(matches: tuple[PhaseMarkerColumnMatch, ...]) -> str:
    return "；".join(format_phase_column_match(match) for match in matches)


def format_result_location(result: PhaseColumnPreprocessResult) -> str:
    return f"{result.path} / {result.sheet_name}"


def format_failure_reason(result: PhaseColumnPreprocessResult) -> str:
    if result.status == STATUS_MISSING_SHEET:
        return f"未找到 sheet：{result.sheet_name}"
    if result.status == STATUS_FILE_MISSING:
        return "文件不存在"
    if result.status == STATUS_READ_ERROR:
        return f"文件读取失败（{result.error_message or '未知错误'}）"
    if result.status == STATUS_SAVE_ERROR:
        return f"检测到期次标记但保存失败（{result.error_message or '未知错误'}）"
    return f"未知错误（{result.status}）"


def format_result_message(result: PhaseColumnPreprocessResult) -> str:
    if result.status == STATUS_UPDATED:
        return (
            "[OK] 已完成预处理：第三列检测到期次标记"
            f"（{format_markers(result.matched_markers)}），"
            f"已移动到最后一列并保存：{format_result_location(result)}"
        )
    if result.status == STATUS_ALREADY_PROCESSED:
        column_summary = format_phase_column_matches(result.detected_phase_columns)
        return (
            "[INFO] 第三列未检测到期次标记，但发现符合特征的列，"
            f"列名为 {column_summary}，可能已经处理过，"
            f"跳过处理：{format_result_location(result)}"
        )
    if result.status == STATUS_NO_PHASE_MARKER:
        return (
            f"[INFO] 未发现期次特征列，无需处理：{format_result_location(result)}"
        )
    if result.status == STATUS_INSUFFICIENT_COLUMNS:
        return (
            "[WARN] 文件列数不足，未发现第三列，"
            f"跳过处理：{format_result_location(result)}"
        )
    if result.status == STATUS_MISSING_SHEET:
        return f"[ERROR] 未找到 sheet：{result.sheet_name}（文件：{result.path}）"
    if result.status == STATUS_FILE_MISSING:
        return f"[ERROR] 文件不存在：{result.path}"
    if result.status == STATUS_READ_ERROR:
        return f"[ERROR] 文件读取失败：{result.path}（{result.error_message or '未知错误'}）"
    if result.status == STATUS_SAVE_ERROR:
        return (
            "[ERROR] 检测到期次标记，但保存失败："
            f"{result.path} / {result.sheet_name}（{result.error_message or '未知错误'}）"
        )
    return f"[ERROR] 未知处理状态：{result.path} / {result.sheet_name}（{result.status}）"


def format_summary_group(
    title: str,
    results: tuple[PhaseColumnPreprocessResult, ...],
) -> list[str]:
    lines = [f"{title}（{len(results)}）："]
    if not results:
        lines.append("- 无")
        return lines

    for result in results:
        if result.status == STATUS_ALREADY_PROCESSED and result.detected_phase_columns:
            lines.append(
                f"- {format_result_location(result)}："
                f"{format_phase_column_matches(result.detected_phase_columns)}"
            )
            continue
        if result.status in {STATUS_FILE_MISSING, STATUS_MISSING_SHEET, STATUS_READ_ERROR, STATUS_SAVE_ERROR}:
            lines.append(f"- {format_result_location(result)}：{format_failure_reason(result)}")
            continue
        lines.append(f"- {format_result_location(result)}")
    return lines


def format_summary_conclusion(summary: PhaseColumnPreprocessSummary) -> str:
    if hasattr(summary, "results"):
        total_count = len(summary.results)
    else:
        total_count = (
            getattr(summary, "updated_count", 0)
            + getattr(summary, "already_processed_count", 0)
            + getattr(summary, "no_phase_marker_count", 0)
            + getattr(summary, "insufficient_columns_count", 0)
            + getattr(summary, "failed_count", 0)
        )
    return (
        "[INFO] 总结："
        f"共检查 {total_count} 个文件；"
        f"成功处理 {summary.updated_count} 个，"
        f"疑似已处理过 {summary.already_processed_count} 个，"
        f"不含期次特征列 {summary.no_phase_marker_count} 个，"
        f"列数不足 {summary.insufficient_columns_count} 个，"
        f"失败 {summary.failed_count} 个。"
    )


def format_summary_message(summary: PhaseColumnPreprocessSummary) -> str:
    lines = ["[INFO] 处理结束汇总："]
    lines.extend(format_summary_group("成功处理", summary.updated_results))
    lines.extend(format_summary_group("未处理，疑似已处理过", summary.already_processed_results))
    lines.extend(format_summary_group("未处理，不含期次特征列", summary.no_phase_marker_results))
    lines.extend(format_summary_group("未处理，列数不足", summary.insufficient_columns_results))
    lines.extend(format_summary_group("失败", summary.failed_results))
    lines.append(format_summary_conclusion(summary))
    return "\n".join(lines)


def run_phase_column_preprocess(
    input_paths: list[Path] | tuple[Path, ...],
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> PhaseColumnPreprocessSummary:
    results = tuple(process_phase_column_workbook(path, sheet_name=sheet_name) for path in input_paths)
    return PhaseColumnPreprocessSummary(results=results)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="检查 Excel 第三列是否为期次列；若检测到“一期/二期”等标记，则移动到最后一列并原地保存。"
    )
    parser.add_argument("inputs", nargs="+", type=Path, help="要处理的 Excel 文件路径，可传多个")
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"要检查的 sheet 名，默认 {DEFAULT_SHEET_NAME}",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    results: list[PhaseColumnPreprocessResult] = []
    for input_path in args.inputs:
        print(format_start_message(input_path, args.sheet_name))
        result = process_phase_column_workbook(input_path, sheet_name=args.sheet_name)
        print(format_result_message(result))
        results.append(result)

    summary = PhaseColumnPreprocessSummary(results=tuple(results))
    print(format_summary_message(summary))
    return 1 if summary.failed_count else 0


if __name__ == "__main__":
    raise SystemExit(main())

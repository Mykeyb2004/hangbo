from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

DEFAULT_SHEET_NAME = "问卷数据"


@dataclass(frozen=True)
class WorkbookReadError:
    path: Path
    error_message: str


@dataclass(frozen=True)
class DuplicateHeaderError:
    path: Path
    headers: tuple[str, ...]
    duplicate_headers: tuple[str, ...]


@dataclass(frozen=True)
class WorkbookSheetData:
    path: Path
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]


@dataclass(frozen=True)
class MergeResult:
    file_name: str
    source_paths: tuple[Path, ...]
    status: str
    merged_rows: int = 0
    output_path: Path | None = None
    missing_sheet_paths: tuple[Path, ...] = ()
    read_errors: tuple[WorkbookReadError, ...] = ()
    duplicate_header_errors: tuple[DuplicateHeaderError, ...] = ()


@dataclass(frozen=True)
class MergeSummary:
    input_dirs: tuple[Path, ...]
    output_dir: Path
    results: tuple[MergeResult, ...]

    @property
    def merged_count(self) -> int:
        return sum(1 for item in self.results if item.status == "merged")

    @property
    def skipped_count(self) -> int:
        return sum(1 for item in self.results if item.status != "merged")


def make_absolute_path(path: Path) -> Path:
    expanded_path = Path(path).expanduser()
    if expanded_path.is_absolute():
        return expanded_path
    return Path.cwd() / expanded_path


def iter_excel_paths(input_dir: Path, recursive: bool = False) -> list[Path]:
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    paths = sorted(path for path in input_dir.glob(pattern) if path.is_file())
    return [
        path
        for path in paths
        if not path.name.startswith("~$")
        and not path.name.startswith("._")
    ]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def trim_trailing_empty_headers(headers: list[str]) -> tuple[str, ...]:
    last_non_empty_index = -1
    for index, header in enumerate(headers):
        if header:
            last_non_empty_index = index
    if last_non_empty_index == -1:
        return ()
    return tuple(headers[: last_non_empty_index + 1])


def is_empty_row(values: tuple[object, ...]) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return False
    return True


def read_questionnaire_sheet(
    workbook_path: Path,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> tuple[str, WorkbookSheetData | None, str | None]:
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - 防御性分支
        return "read_error", None, str(exc)

    if sheet_name not in workbook.sheetnames:
        return "missing_sheet", None, None

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = trim_trailing_empty_headers([normalize_header(value) for value in header_row])

    rows: list[tuple[object, ...]] = []
    if headers:
        for row in worksheet.iter_rows(
            min_row=2,
            max_col=len(headers),
            values_only=True,
        ):
            normalized_row = tuple(row[: len(headers)])
            if is_empty_row(normalized_row):
                continue
            rows.append(normalized_row)

    return (
        "ok",
        WorkbookSheetData(
            path=make_absolute_path(workbook_path),
            headers=headers,
            rows=tuple(rows),
        ),
        None,
    )


def find_duplicate_headers(headers: tuple[str, ...]) -> tuple[str, ...]:
    counts = Counter(headers)
    duplicates: list[str] = []
    seen: set[str] = set()
    for header in headers:
        if counts[header] <= 1 or header in seen:
            continue
        duplicates.append(header)
        seen.add(header)
    return tuple(duplicates)


def merge_headers(sheet_data_items: tuple[WorkbookSheetData, ...]) -> tuple[str, ...]:
    merged_headers: list[str] = []
    seen_headers: set[str] = set()
    for sheet_data in sheet_data_items:
        for header in sheet_data.headers:
            if header in seen_headers:
                continue
            merged_headers.append(header)
            seen_headers.add(header)
    return tuple(merged_headers)


def align_rows(
    rows: tuple[tuple[object, ...], ...],
    source_headers: tuple[str, ...],
    target_headers: tuple[str, ...],
) -> tuple[tuple[object, ...], ...]:
    if source_headers == target_headers:
        return rows

    index_by_header = {header: index for index, header in enumerate(source_headers)}
    return tuple(
        tuple(row[index_by_header[header]] if header in index_by_header else None for header in target_headers)
        for row in rows
    )


def write_merged_workbook(
    output_path: Path,
    *,
    headers: tuple[str, ...],
    rows: tuple[tuple[object, ...], ...],
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    if headers:
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append(list(row))

    workbook.save(output_path)


def group_workbooks_by_filename(
    input_dirs: list[Path] | tuple[Path, ...],
    *,
    recursive: bool = False,
) -> dict[str, list[Path]]:
    grouped_paths: dict[str, list[Path]] = defaultdict(list)
    for input_dir in input_dirs:
        for path in iter_excel_paths(input_dir, recursive=recursive):
            grouped_paths[path.name].append(make_absolute_path(path))
    return dict(grouped_paths)


def merge_workbooks_by_filename(
    input_dirs: list[Path] | tuple[Path, ...],
    *,
    output_dir: Path,
    recursive: bool = False,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> MergeSummary:
    normalized_input_dirs = tuple(make_absolute_path(path) for path in input_dirs)
    normalized_output_dir = make_absolute_path(output_dir)
    grouped_paths = group_workbooks_by_filename(normalized_input_dirs, recursive=recursive)

    results: list[MergeResult] = []

    for file_name in sorted(grouped_paths):
        source_paths = tuple(grouped_paths[file_name])
        sheet_data_items: list[WorkbookSheetData] = []
        missing_sheet_paths: list[Path] = []
        read_errors: list[WorkbookReadError] = []

        for path in source_paths:
            status, sheet_data, error_message = read_questionnaire_sheet(path, sheet_name=sheet_name)
            if status == "ok" and sheet_data is not None:
                sheet_data_items.append(sheet_data)
            elif status == "missing_sheet":
                missing_sheet_paths.append(path)
            else:
                read_errors.append(
                    WorkbookReadError(
                        path=path,
                        error_message=error_message or "未知错误",
                    )
                )

        if read_errors:
            results.append(
                MergeResult(
                    file_name=file_name,
                    source_paths=source_paths,
                    status="read_error",
                    read_errors=tuple(read_errors),
                )
            )
            continue

        if missing_sheet_paths:
            results.append(
                MergeResult(
                    file_name=file_name,
                    source_paths=source_paths,
                    status="missing_sheet",
                    missing_sheet_paths=tuple(missing_sheet_paths),
                )
            )
            continue

        if not sheet_data_items:
            results.append(
                MergeResult(
                    file_name=file_name,
                    source_paths=source_paths,
                    status="missing_sheet",
                )
            )
            continue

        duplicate_header_errors: list[DuplicateHeaderError] = []
        for sheet_data in sheet_data_items:
            duplicate_headers = find_duplicate_headers(sheet_data.headers)
            if duplicate_headers:
                duplicate_header_errors.append(
                    DuplicateHeaderError(
                        path=sheet_data.path,
                        headers=sheet_data.headers,
                        duplicate_headers=duplicate_headers,
                    )
                )

        if duplicate_header_errors:
            results.append(
                MergeResult(
                    file_name=file_name,
                    source_paths=source_paths,
                    status="duplicate_headers",
                    duplicate_header_errors=tuple(duplicate_header_errors),
                )
            )
            continue

        merged_headers = merge_headers(tuple(sheet_data_items))
        merged_rows: list[tuple[object, ...]] = []
        for sheet_data in sheet_data_items:
            merged_rows.extend(
                align_rows(
                    sheet_data.rows,
                    source_headers=sheet_data.headers,
                    target_headers=merged_headers,
                )
            )

        output_path = normalized_output_dir / file_name
        write_merged_workbook(
            output_path,
            headers=merged_headers,
            rows=tuple(merged_rows),
            sheet_name=sheet_name,
        )
        results.append(
            MergeResult(
                file_name=file_name,
                source_paths=source_paths,
                status="merged",
                merged_rows=len(merged_rows),
                output_path=output_path,
            )
        )

    return MergeSummary(
        input_dirs=normalized_input_dirs,
        output_dir=normalized_output_dir,
        results=tuple(results),
    )


def format_merge_summary(
    summary: MergeSummary,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> str:
    lines = [
        f"输出目录: {summary.output_dir}",
        f"输入目录数: {len(summary.input_dirs)}",
        f"发现文件名分组数: {len(summary.results)}",
        f"合并成功: {summary.merged_count}",
        f"跳过/失败: {summary.skipped_count}",
        "",
        "处理明细：",
    ]

    if not summary.results:
        lines.append("- 未找到可处理的 xlsx 文件")
        return "\n".join(lines)

    for result in summary.results:
        if result.status == "merged":
            lines.append(
                f"- {result.file_name}: 已合并 {len(result.source_paths)} 个文件，共 {result.merged_rows} 行"
            )
            continue

        if result.status == "missing_sheet":
            lines.append(f"- {result.file_name}: 跳过，存在缺少 {sheet_name} sheet 的文件")
            for path in result.missing_sheet_paths:
                lines.append(f"  - 缺少 {sheet_name} sheet: {path}")
            continue

        if result.status == "read_error":
            lines.append(f"- {result.file_name}: 跳过，存在读取失败的文件")
            for error in result.read_errors:
                lines.append(f"  - 读取失败: {error.path}（{error.error_message}）")
            continue

        if result.status == "duplicate_headers":
            lines.append(f"- {result.file_name}: 跳过，存在重复列名，无法按列名安全合并")
            for error in result.duplicate_header_errors:
                lines.append(f"  - 文件: {error.path}")
                lines.append(
                    f"  - 重复列名: {', '.join(error.duplicate_headers) if error.duplicate_headers else '无'}"
                )
                lines.append(
                    "  - 完整列名: "
                    f"{', '.join(error.headers) if error.headers else '空表头'}"
                )
            continue

        lines.append(f"- {result.file_name}: {result.status}")

    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "按文件名合并多个目录中的 Excel 文件，只合并“问卷数据”sheet。"
            "同名列合并到同一列，不同名列追加到结果末尾。"
        )
    )
    parser.add_argument(
        "--input-dir",
        dest="input_dirs",
        action="append",
        type=Path,
        required=True,
        help="要扫描的目录，可重复传入多个",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        required=True,
        help="合并后的 Excel 输出目录",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="是否递归扫描子目录中的 xlsx 文件",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"要合并的 sheet 名，默认 {DEFAULT_SHEET_NAME}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = merge_workbooks_by_filename(
        args.input_dirs,
        output_dir=args.output_dir,
        recursive=args.recursive,
        sheet_name=args.sheet_name,
    )
    print(format_merge_summary(summary, sheet_name=args.sheet_name))


if __name__ == "__main__":
    main()

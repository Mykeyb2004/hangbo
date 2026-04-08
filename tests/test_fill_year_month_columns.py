from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from fill_year_month_columns import (
    DEFAULT_SHEET_NAME,
    apply_year_month_to_directory,
    apply_year_month_to_workbook,
    format_directory_summary,
)


def write_workbook(
    output_path: Path,
    headers: list[str],
    rows: list[list[object]],
    *,
    include_target_sheet: bool = True,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = DEFAULT_SHEET_NAME if include_target_sheet else "其他sheet"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    workbook.save(output_path)


class FillYearMonthColumnsTest(unittest.TestCase):
    def test_apply_year_month_to_workbook_appends_columns_to_questionnaire_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "展览.xlsx"
            write_workbook(
                workbook_path,
                headers=["姓名", "开始填表时间"],
                rows=[
                    ["张三", "2026-02-10 10:00:00"],
                    ["李四", "2026-02-11 11:00:00"],
                ],
            )

            result = apply_year_month_to_workbook(workbook_path, year="2026", month="02")

            self.assertEqual(result.status, "updated")
            self.assertEqual(result.updated_rows, 2)

            worksheet = load_workbook(workbook_path)["问卷数据"]
            self.assertEqual(worksheet.cell(row=1, column=3).value, "年份")
            self.assertEqual(worksheet.cell(row=1, column=4).value, "月份")
            self.assertEqual(worksheet.cell(row=2, column=3).value, "2026")
            self.assertEqual(worksheet.cell(row=2, column=4).value, "02")
            self.assertEqual(worksheet.cell(row=3, column=3).value, "2026")
            self.assertEqual(worksheet.cell(row=3, column=4).value, "02")

    def test_apply_year_month_to_workbook_overwrites_existing_values_as_text(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "会议.xlsx"
            write_workbook(
                workbook_path,
                headers=["姓名", "年份", "月份"],
                rows=[
                    ["张三", 2025, 1],
                    ["李四", "旧年份", "旧月份"],
                ],
            )

            result = apply_year_month_to_workbook(workbook_path, year="2026", month="03")

            self.assertEqual(result.status, "updated")
            self.assertEqual(result.updated_rows, 2)

            worksheet = load_workbook(workbook_path)["问卷数据"]
            self.assertEqual(worksheet.cell(row=2, column=2).value, "2026")
            self.assertEqual(worksheet.cell(row=2, column=3).value, "03")
            self.assertEqual(worksheet.cell(row=2, column=2).data_type, "s")
            self.assertEqual(worksheet.cell(row=2, column=3).data_type, "s")
            self.assertEqual(worksheet.cell(row=3, column=2).value, "2026")
            self.assertEqual(worksheet.cell(row=3, column=3).value, "03")

    def test_apply_year_month_to_directory_skips_files_without_questionnaire_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir)
            write_workbook(
                input_dir / "有sheet.xlsx",
                headers=["姓名"],
                rows=[["张三"]],
            )
            write_workbook(
                input_dir / "缺sheet.xlsx",
                headers=["姓名"],
                rows=[["李四"]],
                include_target_sheet=False,
            )

            summary = apply_year_month_to_directory(input_dir, year="2026", month="04")
            report = format_directory_summary(summary)

            self.assertEqual(summary.updated_count, 1)
            self.assertEqual(summary.skipped_count, 1)

            by_name = {item.path.name: item for item in summary.file_results}
            self.assertEqual(by_name["有sheet.xlsx"].status, "updated")
            self.assertEqual(by_name["缺sheet.xlsx"].status, "missing_sheet")
            self.assertIn("跳过文件：", report)
            self.assertIn("缺sheet.xlsx: 缺少 问卷数据 sheet", report)


if __name__ == "__main__":
    unittest.main()

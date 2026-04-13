from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText

from summary_table import (
    SUMMARY_CHINESE_FONT_NAME,
    DEFAULT_SUMMARY_TITLE,
    SUMMARY_BODY_FILL,
    SUMMARY_BORDER,
    SUMMARY_HEADER_FILL,
    SUMMARY_LATIN_FONT_NAME,
    SUMMARY_NO_DATA_TEXT,
    build_summary_dataframe,
    build_summary_rows,
    generate_summary_report,
    load_report_snapshots,
)
from survey_stats import OVERALL_FILL, SECTION_FILL


def write_role_report(
    output_path: Path,
    role_name: str,
    rows: list[tuple[str, float | None, str]],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = role_name
    worksheet.append(["指标", "满意度", "重要性"])

    for label, satisfaction, row_type in rows:
        worksheet.append([label, satisfaction, None])
        row_index = worksheet.max_row
        if row_type == "overall":
            fill = OVERALL_FILL
        elif row_type == "section":
            fill = SECTION_FILL
        else:
            fill = None

        if fill is not None:
            for column_index in range(1, 4):
                worksheet.cell(row=row_index, column=column_index).fill = fill

    workbook.save(output_path)


class SummaryTableTest(unittest.TestCase):
    def test_build_summary_dataframe_maps_roles_and_combines_hotel_catering(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir)
            write_role_report(
                input_dir / "展览主承办.xlsx",
                "展览主承办",
                [
                    ("展览主承办", 9.14, "overall"),
                    ("会展服务", 9.33, "section"),
                    ("销售经理服务态度", 9.2, "metric"),
                    ("硬件设施", 9.71, "section"),
                    ("配套服务", 9.0, "section"),
                    ("餐饮服务", 8.75, "metric"),
                    ("智慧场馆", 8.5, "section"),
                ],
            )
            write_role_report(
                input_dir / "旅行社工作人员.xlsx",
                "旅行社工作人员",
                [
                    ("旅行社工作人员", 9.2, "overall"),
                    ("旅游服务", 9.4, "section"),
                    ("硬件设施", 8.7, "section"),
                    ("智慧服务", 9.1, "section"),
                ],
            )
            write_role_report(
                input_dir / "散客.xlsx",
                "散客",
                [
                    ("散客", 9.5, "overall"),
                    ("入住服务", 9.6, "section"),
                    ("硬件设施", 9.4, "section"),
                    ("餐饮服务", 9.7, "section"),
                    ("智慧场馆", 8.9, "section"),
                ],
            )
            write_role_report(
                input_dir / "酒店宴会.xlsx",
                "酒店宴会",
                [
                    ("酒店宴会", 9.7, "overall"),
                    ("餐饮服务", 9.8, "section"),
                    ("硬件设施", 9.6, "section"),
                    ("智慧场馆", 8.1, "section"),
                ],
            )
            write_role_report(
                input_dir / "酒店自助餐.xlsx",
                "酒店自助餐",
                [
                    ("酒店自助餐", 9.9, "overall"),
                    ("餐饮服务", 10.0, "section"),
                    ("硬件设施", 9.8, "section"),
                    ("智慧场馆", 8.3, "section"),
                ],
            )

            reports = load_report_snapshots(input_dir)
            rows = build_summary_rows(reports)
            summary_df = build_summary_dataframe(rows)

            organizer_row = summary_df[summary_df["样本类型"] == "展览活动主（承）办"].iloc[0]
            self.assertEqual(organizer_row["客户大类"], "一、会展客户")
            self.assertEqual(organizer_row["总分"], 9.14)
            self.assertEqual(organizer_row["产品服务"], 9.33)
            self.assertEqual(organizer_row["餐饮服务"], 8.75)

            travel_staff_row = summary_df[summary_df["样本类型"] == "旅行社工作人员"].iloc[0]
            self.assertEqual(travel_staff_row["客户大类"], "三、G20峰会体验馆")
            self.assertEqual(travel_staff_row["产品服务"], 9.4)
            self.assertEqual(travel_staff_row["智慧场馆/服务"], 9.1)
            self.assertTrue(travel_staff_row["配套服务"] != travel_staff_row["配套服务"])

            guest_row = summary_df[summary_df["样本类型"] == "散客"].iloc[0]
            self.assertEqual(guest_row["产品服务"], 9.6)
            self.assertEqual(guest_row["餐饮服务"], 9.7)

            hotel_catering_row = summary_df[summary_df["样本类型"] == "酒店餐饮客户"].iloc[0]
            self.assertEqual(hotel_catering_row["客户大类"], "五、酒店客户")
            self.assertEqual(hotel_catering_row["总分"], 9.8)
            self.assertEqual(hotel_catering_row["硬件设施"], 9.7)
            self.assertEqual(hotel_catering_row["智慧场馆/服务"], 8.2)
            self.assertEqual(hotel_catering_row["餐饮服务"], 9.9)

            total_row = summary_df[summary_df["样本类型"] == "总分"].iloc[0]
            self.assertEqual(total_row["总分"], 9.41)
            self.assertEqual(total_row["产品服务"], 9.44)

    def test_hotel_meeting_row_prefers_dining_section_over_support_metric(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir)
            write_role_report(
                input_dir / "酒店会议主承办.xlsx",
                "酒店会议主承办",
                [
                    ("酒店会议主承办", 9.5, "overall"),
                    ("会展服务", 9.7, "section"),
                    ("硬件设施", 9.1, "section"),
                    ("餐饮服务", 9.4, "section"),
                    ("配套服务", 9.2, "section"),
                    ("餐饮服务", 8.2, "metric"),
                    ("智慧场馆", 8.9, "section"),
                ],
            )

            reports = load_report_snapshots(input_dir)
            rows = build_summary_rows(reports)
            summary_df = build_summary_dataframe(rows)

            row = summary_df[summary_df["样本类型"] == "酒店会议活动主（承）办"].iloc[0]
            self.assertEqual(row["总分"], 9.5)
            self.assertEqual(row["产品服务"], 9.7)
            self.assertEqual(row["配套服务"], 9.2)
            self.assertEqual(row["餐饮服务"], 9.4)

    def test_hotel_catering_row_accepts_combined_report_alias(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir)
            write_role_report(
                input_dir / "酒店餐饮客户.xlsx",
                "酒店餐饮客户",
                [
                    ("酒店餐饮客户", 9.8, "overall"),
                    ("餐饮服务", 9.9, "section"),
                    ("硬件设施", 9.7, "section"),
                    ("智慧场馆", 8.2, "section"),
                ],
            )

            reports = load_report_snapshots(input_dir)
            rows = build_summary_rows(reports)
            summary_df = build_summary_dataframe(rows)

            row = summary_df[summary_df["样本类型"] == "酒店餐饮客户"].iloc[0]
            self.assertEqual(row["总分"], 9.8)
            self.assertEqual(row["硬件设施"], 9.7)
            self.assertEqual(row["智慧场馆/服务"], 8.2)
            self.assertEqual(row["餐饮服务"], 9.9)

    def test_food_hall_support_section_is_not_used_in_summary(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir)
            write_role_report(
                input_dir / "特色美食廊.xlsx",
                "特色美食廊",
                [
                    ("特色美食廊", 9.4, "overall"),
                    ("餐饮服务", 9.8, "section"),
                    ("硬件设施", 9.1, "section"),
                    ("配套服务", 8.5, "section"),
                    ("智慧场馆", 8.7, "section"),
                ],
            )

            reports = load_report_snapshots(input_dir)
            rows = build_summary_rows(reports)
            summary_df = build_summary_dataframe(rows)

            row = summary_df[summary_df["样本类型"] == "特色美食廊"].iloc[0]
            self.assertEqual(row["总分"], 9.4)
            self.assertEqual(row["硬件设施"], 9.1)
            self.assertEqual(row["智慧场馆/服务"], 8.7)
            self.assertEqual(row["餐饮服务"], 9.8)
            self.assertTrue(pd.isna(row["配套服务"]))

    def test_generate_summary_report_creates_expected_layout_and_no_data_placeholders(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_dir = temp_path / "inputs"
            output_dir = temp_path / "outputs"
            input_dir.mkdir()
            output_dir.mkdir()

            write_role_report(
                input_dir / "商务简餐.xlsx",
                "商务简餐",
                [
                    ("商务简餐", 9.58, "overall"),
                    ("餐饮服务", 9.98, "section"),
                    ("硬件设施", 9.17, "section"),
                    ("智慧场馆", 8.88, "section"),
                ],
            )

            unrelated_workbook = Workbook()
            unrelated_sheet = unrelated_workbook.active
            unrelated_sheet.title = "Sheet1"
            unrelated_sheet.append(["不是", "统计表"])
            unrelated_workbook.save(input_dir / "原始数据.xlsx")

            output_path = generate_summary_report(input_dir=input_dir, output_dir=output_dir)

            self.assertTrue(output_path.exists())

            workbook = load_workbook(output_path, rich_text=True)
            worksheet = workbook.active

            self.assertEqual(worksheet["A1"].value, DEFAULT_SUMMARY_TITLE)
            self.assertIn("A1:H1", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
            self.assertIn("A3:A8", {str(cell_range) for cell_range in worksheet.merged_cells.ranges})
            self.assertEqual(worksheet["A1"].fill.start_color.rgb, SUMMARY_HEADER_FILL.start_color.rgb)
            self.assertEqual(worksheet["A1"].font.color.rgb, "00FFFFFF")
            self.assertEqual(worksheet["A1"].font.name, SUMMARY_CHINESE_FONT_NAME)
            self.assertEqual(worksheet["A1"].border.left.color.rgb, SUMMARY_BORDER.left.color.rgb)

            business_meal_row = None
            for row_index in range(3, worksheet.max_row + 1):
                if worksheet.cell(row=row_index, column=2).value == "商务简餐":
                    business_meal_row = row_index
                    break

            self.assertIsNotNone(business_meal_row)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=3).value, 9.58)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=2).fill.start_color.rgb, SUMMARY_HEADER_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=2).font.color.rgb, "00FFFFFF")
            self.assertEqual(worksheet.cell(row=business_meal_row, column=2).font.name, SUMMARY_CHINESE_FONT_NAME)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=3).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=3).font.name, SUMMARY_LATIN_FONT_NAME)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=3).number_format, "0.00")
            self.assertEqual(worksheet.cell(row=business_meal_row, column=6).value, SUMMARY_NO_DATA_TEXT)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=6).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=worksheet.max_row, column=3).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=worksheet.max_row, column=3).font.name, SUMMARY_LATIN_FONT_NAME)
            self.assertEqual(worksheet.cell(row=worksheet.max_row, column=3).number_format, "0.00")

            special_research_row = None
            for row_index in range(3, worksheet.max_row + 1):
                if worksheet.cell(row=row_index, column=1).value == "四、专项调研":
                    special_research_row = row_index
                    break

            self.assertIsNotNone(special_research_row)
            self.assertEqual(worksheet.cell(row=special_research_row, column=3).value, SUMMARY_NO_DATA_TEXT)
            self.assertEqual(worksheet.cell(row=special_research_row, column=8).value, SUMMARY_NO_DATA_TEXT)

            g20_label_value = None
            for row_index in range(3, worksheet.max_row + 1):
                candidate = worksheet.cell(row=row_index, column=1).value
                if candidate is not None and str(candidate) == "三、G20峰会体验馆":
                    g20_label_value = candidate
                    break

            self.assertIsInstance(g20_label_value, CellRichText)
            self.assertEqual(len(g20_label_value), 3)
            self.assertEqual(g20_label_value[0].text, "三、")
            self.assertEqual(g20_label_value[0].font.rFont, SUMMARY_CHINESE_FONT_NAME)
            self.assertEqual(g20_label_value[1].text, "G20")
            self.assertEqual(g20_label_value[1].font.rFont, SUMMARY_LATIN_FONT_NAME)
            self.assertEqual(g20_label_value[2].text, "峰会体验馆")
            self.assertEqual(g20_label_value[2].font.rFont, SUMMARY_CHINESE_FONT_NAME)

    def test_food_hall_support_cell_uses_no_data_placeholder(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_dir = temp_path / "inputs"
            output_dir = temp_path / "outputs"
            input_dir.mkdir()
            output_dir.mkdir()

            write_role_report(
                input_dir / "特色美食廊.xlsx",
                "特色美食廊",
                [
                    ("特色美食廊", 9.4, "overall"),
                    ("餐饮服务", 9.8, "section"),
                    ("硬件设施", 9.1, "section"),
                    ("配套服务", 8.5, "section"),
                    ("智慧场馆", 8.7, "section"),
                ],
            )

            output_path = generate_summary_report(input_dir=input_dir, output_dir=output_dir)
            worksheet = load_workbook(output_path).active

            food_hall_row = None
            for row_index in range(3, worksheet.max_row + 1):
                if worksheet.cell(row=row_index, column=2).value == "特色美食廊":
                    food_hall_row = row_index
                    break

            self.assertIsNotNone(food_hall_row)
            self.assertEqual(worksheet.cell(row=food_hall_row, column=6).value, SUMMARY_NO_DATA_TEXT)
            self.assertEqual(worksheet.cell(row=food_hall_row, column=6).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)

    def test_applicable_missing_cells_remain_blank(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_dir = temp_path / "inputs"
            output_dir = temp_path / "outputs"
            input_dir.mkdir()
            output_dir.mkdir()

            write_role_report(
                input_dir / "商务简餐.xlsx",
                "商务简餐",
                [
                    ("商务简餐", 9.4, "overall"),
                ],
            )

            output_path = generate_summary_report(input_dir=input_dir, output_dir=output_dir)
            worksheet = load_workbook(output_path).active

            business_meal_row = None
            for row_index in range(3, worksheet.max_row + 1):
                if worksheet.cell(row=row_index, column=2).value == "商务简餐":
                    business_meal_row = row_index
                    break

            self.assertIsNotNone(business_meal_row)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=4).value, SUMMARY_NO_DATA_TEXT)
            self.assertIsNone(worksheet.cell(row=business_meal_row, column=5).value)
            self.assertIsNone(worksheet.cell(row=business_meal_row, column=7).value)
            self.assertIsNone(worksheet.cell(row=business_meal_row, column=8).value)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=3).value, 9.4)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=5).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=7).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=8).fill.start_color.rgb, SUMMARY_BODY_FILL.start_color.rgb)
            self.assertEqual(worksheet.cell(row=business_meal_row, column=6).value, SUMMARY_NO_DATA_TEXT)


if __name__ == "__main__":
    unittest.main()

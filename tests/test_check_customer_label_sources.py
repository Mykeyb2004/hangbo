from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

import pandas as pd
from openpyxl import Workbook, load_workbook

from tests.check_customer_label_sources import (
    MappingEntry,
    audit_mapping_entries,
    detect_unexpected_combinations,
    find_tag_locator,
    write_augmented_workbook,
)


class CheckCustomerLabelSourcesTest(unittest.TestCase):
    def test_find_tag_locator_supports_split_tag_expression(self) -> None:
        df = pd.DataFrame(
            {
                "Q1-调研类别": ["酒店餐饮", "酒店餐饮", "餐饮"],
                "Q2-您使用了那种餐饮类型": ["酒店自助餐", "酒店宴会", "商务简餐"],
            }
        )

        locator = find_tag_locator(df, "酒店自助餐/酒店宴会")

        self.assertIsNotNone(locator.best_match)
        self.assertEqual(locator.match_mode, "split")
        self.assertEqual(locator.best_match.column_name, "Q2-您使用了那种餐饮类型")
        self.assertEqual(locator.best_match.matched_values, ("酒店宴会", "酒店自助餐"))
        self.assertEqual(locator.best_match.matched_row_count, 2)

    def test_find_tag_locator_supports_four_way_split_tag_expression(self) -> None:
        df = pd.DataFrame(
            {
                "Q1-调研类别": ["酒店餐饮", "酒店餐饮", "酒店餐饮", "酒店餐饮"],
                "Q2-您使用了那种餐饮类型": ["酒店自助餐", "酒店宴会", "商务简餐", "宴会"],
            }
        )

        locator = find_tag_locator(df, "酒店自助餐/酒店宴会/商务简餐/宴会")

        self.assertIsNotNone(locator.best_match)
        self.assertEqual(locator.match_mode, "split")
        self.assertEqual(locator.best_match.column_name, "Q2-您使用了那种餐饮类型")
        self.assertEqual(set(locator.best_match.matched_values), {"酒店自助餐", "酒店宴会", "商务简餐", "宴会"})
        self.assertEqual(locator.best_match.matched_row_count, 4)

    def test_detect_unexpected_combinations_reports_unmapped_pairs(self) -> None:
        df = pd.DataFrame(
            {
                "Q1-调研类别": ["会议", "会议", "酒店会议", "酒店会议"],
                "Q3-您在会议中的身份": ["会议主承办", "参会人员", "参会人员", "酒店参会客户"],
            }
        )
        mapped_pairs = {
            ("会议", "会议主承办"),
            ("会议", "参会人员"),
            ("酒店会议", "酒店参会客户"),
        }

        extras = detect_unexpected_combinations(
            df,
            auxiliary_column_name="Q1-调研类别",
            data_column_name="Q3-您在会议中的身份",
            mapped_pairs=mapped_pairs,
        )

        self.assertEqual(len(extras), 1)
        self.assertEqual(extras[0].auxiliary_value, "酒店会议")
        self.assertEqual(extras[0].data_value, "参会人员")
        self.assertEqual(extras[0].row_count, 1)

    def test_audit_mapping_entries_finds_customer_category_alias_and_extra_pair(self) -> None:
        df = pd.DataFrame(
            {
                "Q1-调研类别": ["会议", "会议", "酒店会议", "酒店会议"],
                "Q2-活动名称": ["A", "B", "C", "D"],
                "Q3-您在会议中的身份": ["会议主承办", "参会人员", "参会人员", "酒店参会客户"],
            }
        )
        entries = (
            MappingEntry(
                sequence_number=1,
                row_number=1,
                customer_group="会展客户",
                customer_category="会议活动主（承）办",
                source_file_name="会议.xlsx",
                data_tag="会议主承办",
                auxiliary_tag="会议",
            ),
            MappingEntry(
                sequence_number=2,
                row_number=2,
                customer_group="酒店客户",
                customer_category="酒店参会客户",
                source_file_name="会议.xlsx",
                data_tag="酒店参会客户/参会人员",
                auxiliary_tag="酒店会议",
            ),
        )

        report = audit_mapping_entries(entries, {"会议.xlsx": df})

        first_rule = report.rule_audits[0]
        self.assertFalse(first_rule.customer_category_literal_found)
        self.assertTrue(first_rule.identifiable)
        self.assertEqual(first_rule.matched_row_count, 1)

        source_audit = report.source_audits[0]
        self.assertEqual(len(source_audit.unexpected_combinations), 1)
        self.assertIn(
            ("会议", "参会人员"),
            {
                (item.auxiliary_value, item.data_value)
                for item in source_audit.unexpected_combinations
            },
        )

    def test_load_current_workbook_and_sources(self) -> None:
        mapping_path = Path("datas/客户类别对照表V1.0.xlsx")
        source_dir = Path("datas/2025")
        if not mapping_path.exists() or not source_dir.exists():
            self.skipTest("当前仓库未提供实际核查样本。")

        from tests.check_customer_label_sources import audit_mapping_workbook

        report = audit_mapping_workbook(mapping_path, source_dir)

        row19 = next(rule for rule in report.rule_audits if rule.row_number == 19)
        self.assertEqual(row19.source_file_name, "餐饮.xlsx")
        self.assertEqual(row19.data_tag_locator.match_mode, "split")
        self.assertTrue(row19.identifiable)
        self.assertEqual(row19.data_tag_locator.best_match.column_name, "Q2-您使用了那种餐饮类型")

    def test_write_augmented_workbook_appends_locator_columns(self) -> None:
        with TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            workbook_path = tmp_path / "mapping.xlsx"
            output_path = tmp_path / "mapping_out.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "客户大类", "客户类别", "数据来源", "数据标签", "辅助标签"])
            sheet.append([1, "会展客户", "会议活动主（承）办", "会议.xlsx", "会议主承办", "会议"])
            sheet.append([2, "酒店客户", "酒店参会客户", "会议.xlsx", "酒店参会客户", "酒店会议"])
            workbook.save(workbook_path)

            df = pd.DataFrame(
                {
                    "Q1-调研类别": ["会议", "酒店会议"],
                    "Q3-您在会议中的身份": ["会议主承办", "酒店参会客户"],
                }
            )
            entries = (
                MappingEntry(1, 2, "会展客户", "会议活动主（承）办", "会议.xlsx", "会议主承办", "会议"),
                MappingEntry(2, 3, "酒店客户", "酒店参会客户", "会议.xlsx", "酒店参会客户", "酒店会议"),
            )
            report = audit_mapping_entries(entries, {"会议.xlsx": df})

            write_augmented_workbook(workbook_path, output_path, report)

            out_wb = load_workbook(output_path, read_only=True, data_only=True)
            out_ws = out_wb.active
            self.assertEqual(out_ws.cell(1, 7).value, "数据标签所在列")
            self.assertEqual(out_ws.cell(1, 8).value, "辅助标签所在列")
            self.assertEqual(out_ws.cell(2, 7).value, "B")
            self.assertEqual(out_ws.cell(2, 8).value, "A")
            self.assertEqual(out_ws.cell(3, 7).value, "B")
            self.assertEqual(out_ws.cell(3, 8).value, "A")


if __name__ == "__main__":
    unittest.main()

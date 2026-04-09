from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from merge_questionnaire_workbooks import (
    DEFAULT_SHEET_NAME,
    merge_workbooks_by_filename,
    format_merge_summary,
)


def write_workbook(
    output_path: Path,
    headers: list[str],
    rows: list[list[object]],
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(output_path)


class MergeQuestionnaireWorkbooksTest(unittest.TestCase):
    def test_merge_workbooks_by_filename_merges_questionnaire_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            january_dir = root / "1月"
            february_dir = root / "2月"
            output_dir = root / "合并结果"
            january_dir.mkdir()
            february_dir.mkdir()

            write_workbook(
                january_dir / "展览.xlsx",
                headers=["姓名", "得分"],
                rows=[["张三", 90], ["李四", 88]],
            )
            write_workbook(
                february_dir / "展览.xlsx",
                headers=["姓名", "得分"],
                rows=[["王五", 95]],
            )

            summary = merge_workbooks_by_filename(
                [january_dir, february_dir],
                output_dir=output_dir,
            )

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "merged")
            self.assertEqual(result.merged_rows, 3)
            self.assertEqual(result.output_path, output_dir / "展览.xlsx")
            self.assertTrue(result.output_path.exists())

            worksheet = load_workbook(result.output_path)[DEFAULT_SHEET_NAME]
            self.assertEqual(worksheet.max_row, 4)
            self.assertEqual(worksheet.cell(row=1, column=1).value, "姓名")
            self.assertEqual(worksheet.cell(row=1, column=2).value, "得分")
            self.assertEqual(worksheet.cell(row=2, column=1).value, "张三")
            self.assertEqual(worksheet.cell(row=3, column=1).value, "李四")
            self.assertEqual(worksheet.cell(row=4, column=1).value, "王五")

    def test_merge_workbooks_by_filename_adds_new_columns_to_the_end(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()

            write_workbook(
                first_dir / "会议.xlsx",
                headers=["姓名", "得分", "城市"],
                rows=[["张三", 90, "上海"]],
            )
            write_workbook(
                second_dir / "会议.xlsx",
                headers=["姓名", "得分", "电话"],
                rows=[["李四", 88, "123456"]],
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir],
                output_dir=output_dir,
            )
            report = format_merge_summary(summary)

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "merged")
            self.assertEqual(result.merged_rows, 2)
            self.assertTrue((output_dir / "会议.xlsx").exists())
            self.assertIn("已合并 2 个文件，共 2 行", report)

            worksheet = load_workbook(result.output_path)[DEFAULT_SHEET_NAME]
            self.assertEqual(
                [worksheet.cell(row=1, column=index).value for index in range(1, 5)],
                ["姓名", "得分", "城市", "电话"],
            )
            self.assertEqual(
                [worksheet.cell(row=2, column=index).value for index in range(1, 5)],
                ["张三", 90, "上海", None],
            )
            self.assertEqual(
                [worksheet.cell(row=3, column=index).value for index in range(1, 5)],
                ["李四", 88, None, "123456"],
            )

    def test_merge_workbooks_by_filename_aligns_rows_by_column_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            third_dir = root / "目录C"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()
            third_dir.mkdir()

            write_workbook(
                first_dir / "论坛.xlsx",
                headers=["姓名", "得分"],
                rows=[["张三", 90]],
            )
            write_workbook(
                second_dir / "论坛.xlsx",
                headers=["电话", "姓名"],
                rows=[["123456", "李四"]],
            )
            write_workbook(
                third_dir / "论坛.xlsx",
                headers=["城市", "得分"],
                rows=[["北京", 95]],
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir, third_dir],
                output_dir=output_dir,
            )

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "merged")
            self.assertEqual(result.merged_rows, 3)

            worksheet = load_workbook(result.output_path)[DEFAULT_SHEET_NAME]
            self.assertEqual(
                [worksheet.cell(row=1, column=index).value for index in range(1, 5)],
                ["姓名", "得分", "电话", "城市"],
            )
            self.assertEqual(
                [worksheet.cell(row=2, column=index).value for index in range(1, 5)],
                ["张三", 90, None, None],
            )
            self.assertEqual(
                [worksheet.cell(row=3, column=index).value for index in range(1, 5)],
                ["李四", None, "123456", None],
            )
            self.assertEqual(
                [worksheet.cell(row=4, column=index).value for index in range(1, 5)],
                [None, 95, None, "北京"],
            )

    def test_merge_workbooks_by_filename_merges_semantic_headers_with_different_question_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()

            write_workbook(
                first_dir / "会议.xlsx",
                headers=["Q1-调研类别", "Q2-活动名称", "扩展字段"],
                rows=[["会议", "人工智能大会", "A"]],
            )
            write_workbook(
                second_dir / "会议.xlsx",
                headers=["Q2-调研类别", "Q3-活动名称", "Q1-期次", "扩展字段"],
                rows=[["会议", "全球峰会", "三期", "B"]],
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir],
                output_dir=output_dir,
            )

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "merged")
            self.assertEqual(result.merged_rows, 2)

            worksheet = load_workbook(result.output_path)[DEFAULT_SHEET_NAME]
            self.assertEqual(
                [worksheet.cell(row=1, column=index).value for index in range(1, 5)],
                ["Q1-调研类别", "Q2-活动名称", "扩展字段", "Q1-期次"],
            )
            self.assertEqual(
                [worksheet.cell(row=2, column=index).value for index in range(1, 5)],
                ["会议", "人工智能大会", "A", None],
            )
            self.assertEqual(
                [worksheet.cell(row=3, column=index).value for index in range(1, 5)],
                ["会议", "全球峰会", "B", "三期"],
            )

    def test_merge_workbooks_by_filename_skips_when_duplicate_headers_exist(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()

            write_workbook(
                first_dir / "回访.xlsx",
                headers=["姓名", "电话", "电话"],
                rows=[["张三", "111", "222"]],
            )
            write_workbook(
                second_dir / "回访.xlsx",
                headers=["姓名", "电话"],
                rows=[["李四", "333"]],
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir],
                output_dir=output_dir,
            )
            report = format_merge_summary(summary)

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "duplicate_headers")
            self.assertFalse((output_dir / "回访.xlsx").exists())
            self.assertIn("存在重复列名", report)
            self.assertIn("完全重复列数: 1", report)
            self.assertIn("语义重复列数（忽略题号前缀后）: 2", report)

    def test_merge_workbooks_by_filename_skips_when_duplicate_semantic_headers_exist(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()

            write_workbook(
                first_dir / "会议.xlsx",
                headers=["Q1-调研类别", "Q2-调研类别", "姓名"],
                rows=[["会议", "会议", "张三"]],
            )
            write_workbook(
                second_dir / "会议.xlsx",
                headers=["Q3-调研类别", "姓名"],
                rows=[["会议", "李四"]],
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir],
                output_dir=output_dir,
            )
            report = format_merge_summary(summary)

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "duplicate_headers")
            self.assertFalse((output_dir / "会议.xlsx").exists())
            self.assertIn("语义重复列名", report)
            self.assertIn("完全重复列数: 0", report)
            self.assertIn("语义重复列数（忽略题号前缀后）: 2", report)

    def test_merge_workbooks_by_filename_skips_when_sheet_missing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_dir = root / "目录A"
            second_dir = root / "目录B"
            output_dir = root / "合并结果"
            first_dir.mkdir()
            second_dir.mkdir()

            write_workbook(
                first_dir / "参展商.xlsx",
                headers=["姓名", "得分"],
                rows=[["张三", 90]],
            )
            write_workbook(
                second_dir / "参展商.xlsx",
                headers=["姓名", "得分"],
                rows=[["李四", 88]],
                sheet_name="其他sheet",
            )

            summary = merge_workbooks_by_filename(
                [first_dir, second_dir],
                output_dir=output_dir,
            )
            report = format_merge_summary(summary)

            self.assertEqual(len(summary.results), 1)
            result = summary.results[0]
            self.assertEqual(result.status, "missing_sheet")
            self.assertFalse((output_dir / "参展商.xlsx").exists())
            self.assertIn("缺少 问卷数据 sheet", report)
            self.assertIn("目录B", report)


if __name__ == "__main__":
    unittest.main()

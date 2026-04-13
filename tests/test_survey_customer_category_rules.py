from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook

from survey_customer_category_rules import (
    CUSTOMER_CATEGORY_RULE_BY_NAME,
    DISPLAY_ORDERED_CUSTOMER_CATEGORY_RULES,
)


def split_mapping_cell(value: object) -> tuple[str, ...]:
    text = str(value).strip() if value is not None else ""
    if not text:
        return ()
    return tuple(part.strip() for part in text.replace("／", "/").split("/") if part.strip())


def load_customer_mapping_rows() -> list[tuple[int, str, str, str, tuple[str, ...], tuple[str, ...], str, str | None]]:
    workbook_path = Path(__file__).resolve().parents[1] / "datas" / "customer_mapping.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[tuple[int, str, str, str, tuple[str, ...], tuple[str, ...], str, str | None]] = []

    for raw_row in worksheet.iter_rows(values_only=True, min_row=2):
        if not raw_row or all(cell is None for cell in raw_row):
            continue
        sequence_number, customer_group, customer_category, source_file_name, data_tag, auxiliary_tag, data_column, auxiliary_column = raw_row
        rows.append(
            (
                int(sequence_number),
                str(customer_group).strip(),
                str(customer_category).strip(),
                str(source_file_name).strip(),
                split_mapping_cell(data_tag),
                split_mapping_cell(auxiliary_tag),
                str(data_column).strip(),
                str(auxiliary_column).strip() if auxiliary_column is not None else None,
            )
        )

    return rows


class SurveyCustomerCategoryRulesTest(unittest.TestCase):
    def test_display_rules_follow_customer_mapping_workbook_order_and_metadata(self) -> None:
        mapping_rows = load_customer_mapping_rows()

        self.assertEqual(
            [rule.sequence_number for rule in DISPLAY_ORDERED_CUSTOMER_CATEGORY_RULES],
            list(range(1, len(mapping_rows) + 1)),
        )
        self.assertEqual(len(DISPLAY_ORDERED_CUSTOMER_CATEGORY_RULES), len(mapping_rows))

        for rule, (
            sequence_number,
            customer_group,
            customer_category,
            source_file_name,
            data_values,
            auxiliary_values,
            data_column,
            auxiliary_column,
        ) in zip(DISPLAY_ORDERED_CUSTOMER_CATEGORY_RULES, mapping_rows, strict=True):
            self.assertEqual(rule.sequence_number, sequence_number)
            self.assertEqual(rule.customer_group, customer_group)
            self.assertEqual(rule.customer_category, customer_category)
            self.assertEqual(rule.source_file_name, source_file_name)
            self.assertEqual(rule.data_values, data_values)
            self.assertEqual(rule.auxiliary_values, auxiliary_values)
            self.assertEqual(rule.data_column, data_column)
            self.assertEqual(rule.auxiliary_column, auxiliary_column)

    def test_internal_component_rules_do_not_consume_display_sequence_numbers(self) -> None:
        for rule_name in ("酒店宴会", "酒店自助餐", "酒店餐饮-商务简餐", "酒店餐饮-宴会"):
            self.assertIsNone(CUSTOMER_CATEGORY_RULE_BY_NAME[rule_name].sequence_number)


if __name__ == "__main__":
    unittest.main()

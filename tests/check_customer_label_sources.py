from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


IGNORED_FILE_PREFIXES = ("._", "~$")


@dataclass(frozen=True)
class MappingEntry:
    sequence_number: int | None
    row_number: int
    customer_group: str
    customer_category: str
    source_file_name: str
    data_tag: str
    auxiliary_tag: str | None


@dataclass(frozen=True)
class ColumnMatch:
    column_name: str
    column_index: int
    column_letter: str
    matched_values: tuple[str, ...]
    matched_row_count: int


@dataclass(frozen=True)
class TagLocator:
    raw_tag: str | None
    target_values: tuple[str, ...]
    match_mode: str
    best_match: ColumnMatch | None
    candidate_matches: tuple[ColumnMatch, ...]

    @property
    def found_values(self) -> tuple[str, ...]:
        if self.best_match is None:
            return ()
        return self.best_match.matched_values

    @property
    def missing_values(self) -> tuple[str, ...]:
        found = set(self.found_values)
        return tuple(value for value in self.target_values if value not in found)

    @property
    def found_all_values(self) -> bool:
        return bool(self.target_values) and not self.missing_values


@dataclass(frozen=True)
class UnexpectedValue:
    data_value: str
    row_count: int


@dataclass(frozen=True)
class UnexpectedCombination:
    auxiliary_value: str
    data_value: str
    row_count: int


@dataclass(frozen=True)
class RuleAudit:
    sequence_number: int | None
    row_number: int
    customer_group: str
    customer_category: str
    source_file_name: str
    data_tag: str
    auxiliary_tag: str | None
    data_tag_locator: TagLocator
    auxiliary_tag_locator: TagLocator
    matched_row_count: int
    identifiable: bool
    customer_category_literal_found: bool
    customer_category_literal_columns: tuple[str, ...]


@dataclass(frozen=True)
class SourceAudit:
    source_file_name: str
    row_count: int
    data_column_name: str | None
    auxiliary_column_name: str | None
    unexpected_values: tuple[UnexpectedValue, ...]
    unexpected_combinations: tuple[UnexpectedCombination, ...]


@dataclass(frozen=True)
class AuditReport:
    rule_audits: tuple[RuleAudit, ...]
    source_audits: tuple[SourceAudit, ...]


def normalize_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def split_tag_values(tag: str | None) -> tuple[str, ...]:
    if tag is None:
        return ()
    text = str(tag).strip()
    if not text:
        return ()

    values: list[str] = []
    seen: set[str] = set()
    for part in text.replace("／", "/").split("/"):
        normalized = part.strip()
        if not normalized or normalized in seen:
            continue
        values.append(normalized)
        seen.add(normalized)
    return tuple(values)


def index_to_excel_column(index: int) -> str:
    value = index + 1
    letters: list[str] = []
    while value:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(ord("A") + remainder))
    return "".join(reversed(letters))


def normalized_series(df: pd.DataFrame, column_name: str) -> pd.Series:
    return df[column_name].map(normalize_cell)


def find_tag_locator(df: pd.DataFrame, tag: str | None) -> TagLocator:
    target_values = split_tag_values(tag)
    if not target_values:
        return TagLocator(
            raw_tag=tag,
            target_values=(),
            match_mode="none",
            best_match=None,
            candidate_matches=(),
        )

    targets = set(target_values)
    candidates: list[ColumnMatch] = []
    for column_index, column_name in enumerate(df.columns):
        series = normalized_series(df, column_name)
        matched_values = tuple(sorted(targets.intersection(set(value for value in series.unique() if value))))
        if not matched_values:
            continue

        matched_row_count = int(series.isin(targets).sum())
        candidates.append(
            ColumnMatch(
                column_name=str(column_name),
                column_index=column_index,
                column_letter=index_to_excel_column(column_index),
                matched_values=matched_values,
                matched_row_count=matched_row_count,
            )
        )

    candidates.sort(
        key=lambda item: (-len(item.matched_values), -item.matched_row_count, item.column_index)
    )
    best_match = candidates[0] if candidates else None
    return TagLocator(
        raw_tag=tag,
        target_values=target_values,
        match_mode="exact" if len(target_values) == 1 else "split",
        best_match=best_match,
        candidate_matches=tuple(candidates),
    )


def detect_unexpected_values(
    df: pd.DataFrame,
    data_column_name: str,
    mapped_values: Iterable[str],
) -> tuple[UnexpectedValue, ...]:
    mapped = {value for value in mapped_values if value}
    counter = Counter(
        value
        for value in normalized_series(df, data_column_name).tolist()
        if value
    )
    extras = [
        UnexpectedValue(data_value=value, row_count=row_count)
        for value, row_count in counter.items()
        if value not in mapped
    ]
    extras.sort(key=lambda item: (-item.row_count, item.data_value))
    return tuple(extras)


def detect_unexpected_combinations(
    df: pd.DataFrame,
    auxiliary_column_name: str,
    data_column_name: str,
    mapped_pairs: set[tuple[str, str]],
) -> tuple[UnexpectedCombination, ...]:
    auxiliary_series = normalized_series(df, auxiliary_column_name)
    data_series = normalized_series(df, data_column_name)
    counter: Counter[tuple[str, str]] = Counter()

    for auxiliary_value, data_value in zip(auxiliary_series.tolist(), data_series.tolist(), strict=False):
        if not auxiliary_value or not data_value:
            continue
        counter[(auxiliary_value, data_value)] += 1

    extras = [
        UnexpectedCombination(
            auxiliary_value=auxiliary_value,
            data_value=data_value,
            row_count=row_count,
        )
        for (auxiliary_value, data_value), row_count in counter.items()
        if (auxiliary_value, data_value) not in mapped_pairs
    ]
    extras.sort(key=lambda item: (-item.row_count, item.auxiliary_value, item.data_value))
    return tuple(extras)


def find_literal_columns(df: pd.DataFrame, target: str) -> tuple[str, ...]:
    normalized_target = normalize_cell(target)
    if not normalized_target:
        return ()

    columns: list[str] = []
    for column_name in df.columns:
        if normalized_series(df, column_name).eq(normalized_target).any():
            columns.append(str(column_name))
    return tuple(columns)


def build_rule_mask(df: pd.DataFrame, data_locator: TagLocator, auxiliary_locator: TagLocator) -> pd.Series:
    if data_locator.best_match is None:
        return pd.Series([False] * len(df), index=df.index)

    data_series = normalized_series(df, data_locator.best_match.column_name)
    mask = data_series.isin(set(data_locator.target_values))

    if auxiliary_locator.best_match is not None:
        auxiliary_series = normalized_series(df, auxiliary_locator.best_match.column_name)
        mask = mask & auxiliary_series.isin(set(auxiliary_locator.target_values))

    return mask


def audit_mapping_entries(
    entries: Iterable[MappingEntry],
    source_frames: dict[str, pd.DataFrame],
) -> AuditReport:
    rule_audits: list[RuleAudit] = []
    rules_by_source: dict[str, list[RuleAudit]] = defaultdict(list)

    for entry in entries:
        df = source_frames[entry.source_file_name]
        data_locator = find_tag_locator(df, entry.data_tag)
        auxiliary_locator = find_tag_locator(df, entry.auxiliary_tag)
        rule_mask = build_rule_mask(df, data_locator, auxiliary_locator)
        customer_category_literal_columns = find_literal_columns(df, entry.customer_category)
        rule_audit = RuleAudit(
            sequence_number=entry.sequence_number,
            row_number=entry.row_number,
            customer_group=entry.customer_group,
            customer_category=entry.customer_category,
            source_file_name=entry.source_file_name,
            data_tag=entry.data_tag,
            auxiliary_tag=entry.auxiliary_tag,
            data_tag_locator=data_locator,
            auxiliary_tag_locator=auxiliary_locator,
            matched_row_count=int(rule_mask.sum()),
            identifiable=bool(rule_mask.any()),
            customer_category_literal_found=bool(customer_category_literal_columns),
            customer_category_literal_columns=customer_category_literal_columns,
        )
        rule_audits.append(rule_audit)
        rules_by_source[entry.source_file_name].append(rule_audit)

    source_audits: list[SourceAudit] = []
    for source_file_name, df in source_frames.items():
        source_rules = rules_by_source.get(source_file_name, [])
        data_column_candidates = [
            rule.data_tag_locator.best_match.column_name
            for rule in source_rules
            if rule.data_tag_locator.best_match is not None
        ]
        auxiliary_column_candidates = [
            rule.auxiliary_tag_locator.best_match.column_name
            for rule in source_rules
            if rule.auxiliary_tag_locator.best_match is not None
        ]
        data_column_name = most_common_or_none(data_column_candidates)
        auxiliary_column_name = most_common_or_none(auxiliary_column_candidates)

        mapped_data_values: set[str] = set()
        mapped_pairs: set[tuple[str, str]] = set()
        for rule in source_rules:
            mapped_data_values.update(rule.data_tag_locator.target_values)
            if rule.auxiliary_tag_locator.target_values:
                for auxiliary_value in rule.auxiliary_tag_locator.target_values:
                    for data_value in rule.data_tag_locator.target_values:
                        mapped_pairs.add((auxiliary_value, data_value))

        unexpected_values: tuple[UnexpectedValue, ...] = ()
        unexpected_combinations: tuple[UnexpectedCombination, ...] = ()
        if data_column_name is not None:
            if auxiliary_column_name is None:
                unexpected_values = detect_unexpected_values(
                    df,
                    data_column_name=data_column_name,
                    mapped_values=mapped_data_values,
                )
            else:
                unexpected_combinations = detect_unexpected_combinations(
                    df,
                    auxiliary_column_name=auxiliary_column_name,
                    data_column_name=data_column_name,
                    mapped_pairs=mapped_pairs,
                )

        source_audits.append(
            SourceAudit(
                source_file_name=source_file_name,
                row_count=len(df),
                data_column_name=data_column_name,
                auxiliary_column_name=auxiliary_column_name,
                unexpected_values=unexpected_values,
                unexpected_combinations=unexpected_combinations,
            )
        )

    source_audits.sort(key=lambda item: item.source_file_name)
    rule_audits.sort(key=lambda item: item.row_number)
    return AuditReport(rule_audits=tuple(rule_audits), source_audits=tuple(source_audits))


def most_common_or_none(values: Iterable[str]) -> str | None:
    counter = Counter(values)
    if not counter:
        return None
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))[0][0]


def load_mapping_entries(workbook_path: Path, sheet_name: str | None = None) -> tuple[MappingEntry, ...]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    entries: list[MappingEntry] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_index == 1:
            continue
        if not row or all(cell is None for cell in row):
            continue

        entries.append(
            MappingEntry(
                sequence_number=int(row[0]) if row[0] is not None else None,
                row_number=row_index,
                customer_group=normalize_cell(row[1]),
                customer_category=normalize_cell(row[2]),
                source_file_name=normalize_cell(row[3]),
                data_tag=normalize_cell(row[4]),
                auxiliary_tag=normalize_cell(row[5]) or None,
            )
        )

    return tuple(entries)


def load_source_frames(source_dir: Path) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for path in sorted(source_dir.glob("*.xlsx")):
        if path.name.startswith(IGNORED_FILE_PREFIXES):
            continue
        frames[path.name] = pd.read_excel(path)
    return frames


def audit_mapping_workbook(workbook_path: Path, source_dir: Path) -> AuditReport:
    entries = load_mapping_entries(workbook_path)
    source_frames = load_source_frames(source_dir)
    missing_sources = sorted({entry.source_file_name for entry in entries if entry.source_file_name not in source_frames})
    if missing_sources:
        raise FileNotFoundError(f"缺少数据来源文件: {', '.join(missing_sources)}")
    return audit_mapping_entries(entries, source_frames)


def render_markdown_report(report: AuditReport) -> str:
    total_rules = len(report.rule_audits)
    identifiable_rules = sum(1 for rule in report.rule_audits if rule.identifiable)
    literal_found_rules = sum(1 for rule in report.rule_audits if rule.customer_category_literal_found)
    partial_rules = [
        rule
        for rule in report.rule_audits
        if rule.data_tag_locator.target_values and not rule.data_tag_locator.found_all_values
    ]
    extra_value_sources = [audit for audit in report.source_audits if audit.unexpected_values]
    extra_combo_sources = [audit for audit in report.source_audits if audit.unexpected_combinations]

    lines = [
        "# 客户类别对照表 V1.0 与数据源核查",
        "",
        "## 结论摘要",
        "",
        f"- 对照表规则总数：`{total_rules}`",
        f"- 可依据标签规则在源数据中识别的规则数：`{identifiable_rules}` / `{total_rules}`",
        f"- 客户类别字面值可直接在源数据中找到的规则数：`{literal_found_rules}` / `{total_rules}`",
        f"- 数据标签部分命中规则数：`{len(partial_rules)}`",
        f"- 存在表外独立标签的数据源数：`{len(extra_value_sources)}`",
        f"- 存在表外辅助标签+数据标签组合的数据源数：`{len(extra_combo_sources)}`",
        "",
        "## 逐条核查",
        "",
        "| 序号 | 客户类别 | 数据来源 | 数据标签定位 | 辅助标签定位 | 匹配行数 | 客户类别字面值 | 备注 |",
        "| --- | --- | --- | --- | --- | ---: | --- | --- |",
    ]

    for rule in report.rule_audits:
        data_desc = describe_locator(rule.data_tag_locator)
        auxiliary_desc = describe_locator(rule.auxiliary_tag_locator)
        literal_desc = "是" if rule.customer_category_literal_found else "否"
        notes: list[str] = []
        if not rule.identifiable:
            notes.append("按当前标签规则未筛到数据")
        if rule.data_tag_locator.target_values and not rule.data_tag_locator.found_all_values:
            notes.append(f"数据标签缺失值: {', '.join(rule.data_tag_locator.missing_values)}")
        if rule.auxiliary_tag_locator.target_values and not rule.auxiliary_tag_locator.found_all_values:
            notes.append(f"辅助标签缺失值: {', '.join(rule.auxiliary_tag_locator.missing_values)}")
        if not rule.customer_category_literal_found:
            notes.append("客户类别需通过标签映射识别")
        row_label = str(rule.sequence_number) if rule.sequence_number is not None else str(rule.row_number)
        lines.append(
            f"| {row_label} | {rule.customer_category} | {rule.source_file_name} | {data_desc} | {auxiliary_desc} | {rule.matched_row_count} | {literal_desc} | {'；'.join(notes) or '正常'} |"
        )

    lines.extend(
        [
            "",
            "## 数据源反查",
            "",
        ]
    )

    for source_audit in report.source_audits:
        lines.append(f"### {source_audit.source_file_name}")
        lines.append("")
        lines.append(f"- 数据行数：`{source_audit.row_count}`")
        lines.append(f"- 数据标签列：`{source_audit.data_column_name or '未定位'}`")
        lines.append(f"- 辅助标签列：`{source_audit.auxiliary_column_name or '无'}`")
        if source_audit.unexpected_values:
            lines.append("- 表外独立标签：")
            for item in source_audit.unexpected_values:
                lines.append(f"  - `{item.data_value}`：{item.row_count} 行")
        if source_audit.unexpected_combinations:
            lines.append("- 表外标签组合：")
            for item in source_audit.unexpected_combinations:
                lines.append(
                    f"  - `{item.auxiliary_value}` + `{item.data_value}`：{item.row_count} 行"
                )
        if not source_audit.unexpected_values and not source_audit.unexpected_combinations:
            lines.append("- 未发现表外标签或表外组合。")
        lines.append("")

    return "\n".join(lines).strip() + "\n"


def describe_locator(locator: TagLocator) -> str:
    if not locator.target_values:
        return "无"
    if locator.best_match is None:
        return f"未找到 `{locator.raw_tag}`"
    values = " / ".join(locator.best_match.matched_values)
    return (
        f"`{locator.best_match.column_name}` ({locator.best_match.column_letter}列)"
        f" -> {values}"
    )


def build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="核查客户类别对照表与数据源标签的一致性。")
    parser.add_argument(
        "--mapping-workbook",
        type=Path,
        default=Path("datas/客户类别对照表V1.0.xlsx"),
        help="客户类别对照表路径",
    )
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=Path("datas/2025"),
        help="数据源目录",
    )
    parser.add_argument(
        "--markdown-output",
        type=Path,
        help="可选：把 Markdown 报告写入指定文件",
    )
    return parser


def main() -> int:
    parser = build_cli_parser()
    args = parser.parse_args()
    report = audit_mapping_workbook(args.mapping_workbook, args.source_dir)
    markdown = render_markdown_report(report)
    print(markdown)
    if args.markdown_output is not None:
        args.markdown_output.write_text(markdown, encoding="utf-8")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from check_unmapped_customer_records import (
    format_directory_audit_report,
    run_directory_audit,
    write_audit_log,
)
from pipeline_models import PipelineIssue, PipelinePaths, PrecheckResult
from pipeline_paths import STANDARD_SOURCE_FILE_NAMES


def workbook_has_year_month_headers(workbook_path: Path, sheet_name: str) -> bool:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"缺少 sheet：{sheet_name}")

        worksheet = workbook[sheet_name]
        first_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            (),
        )
        headers = {str(value).strip() for value in first_row if value is not None}
        return "年份" in headers and "月份" in headers
    finally:
        workbook.close()


def run_unmapped_audit(
    input_dir: Path,
    *,
    sheet_name: str,
    log_path: Path,
) -> tuple[int, Path]:
    report = run_directory_audit(input_dir, sheet_name=sheet_name)
    report_text = format_directory_audit_report(report, log_path=log_path)
    write_audit_log(report_text, log_path)
    return report.total_unmapped_records, log_path


def run_precheck(
    paths: PipelinePaths,
    *,
    sheet_name: str,
    single_month: int | None,
) -> PrecheckResult:
    blocking_issues: list[PipelineIssue] = []
    warning_issues: list[PipelineIssue] = []
    should_autofill_year_month = False

    if not paths.raw_dir.exists() or not paths.raw_dir.is_dir():
        blocking_issues.append(
            PipelineIssue(
                severity="blocking",
                code="missing_raw_dir",
                message=f"原始批次目录不存在：{paths.raw_dir}",
                path=paths.raw_dir,
            )
        )
        return PrecheckResult(
            blocking_issues=tuple(blocking_issues),
            warning_issues=tuple(warning_issues),
            should_autofill_year_month=should_autofill_year_month,
        )

    present_source_paths = [
        source_path
        for source_path in paths.standard_source_paths
        if source_path.exists() and source_path.is_file()
    ]
    if not present_source_paths:
        blocking_issues.append(
            PipelineIssue(
                severity="blocking",
                code="missing_standard_sources",
                message=f"原始批次目录缺少标准来源文件：{paths.raw_dir}",
                path=paths.raw_dir,
            )
        )

    if present_source_paths:
        missing_year_month_paths: list[Path] = []
        for source_path in present_source_paths:
            try:
                has_year_month = workbook_has_year_month_headers(source_path, sheet_name)
            except ValueError as exc:
                blocking_issues.append(
                    PipelineIssue(
                        severity="blocking",
                        code="missing_sheet",
                        message=str(exc),
                        path=source_path,
                    )
                )
                break
            except Exception as exc:
                blocking_issues.append(
                    PipelineIssue(
                        severity="blocking",
                        code="precheck_error",
                        message=f"预查错过程失败：{source_path.name}: {exc}",
                        path=source_path,
                    )
                )
                break

            if not has_year_month:
                missing_year_month_paths.append(source_path)

        if missing_year_month_paths:
            if single_month is not None:
                should_autofill_year_month = True
                warning_issues.append(
                    PipelineIssue(
                        severity="warning",
                        code="autofill_year_month",
                        message="单月批次来源文件缺少“年份”/“月份”列，将自动补齐。",
                        path=missing_year_month_paths[0],
                    )
                )
            else:
                blocking_issues.append(
                    PipelineIssue(
                        severity="blocking",
                        code="missing_year_month_columns",
                        message="合并批次来源文件缺少“年份”/“月份”列，无法自动补齐。",
                        path=missing_year_month_paths[0],
                    )
                )

    if not blocking_issues:
        try:
            unmapped_count, audit_log_path = run_unmapped_audit(
                paths.raw_dir,
                sheet_name=sheet_name,
                log_path=paths.unmapped_log_path,
            )
        except Exception as exc:
            blocking_issues.append(
                PipelineIssue(
                    severity="blocking",
                    code="precheck_error",
                    message=f"预查错过程失败：{exc}",
                    path=paths.raw_dir,
                )
            )
        else:
            if unmapped_count > 0:
                blocking_issues.append(
                    PipelineIssue(
                        severity="blocking",
                        code="unmapped_customer_records",
                        message=f"发现 {unmapped_count} 条未映射标签记录，请查看：{audit_log_path}",
                        path=audit_log_path,
                    )
                )

    return PrecheckResult(
        blocking_issues=tuple(blocking_issues),
        warning_issues=tuple(warning_issues),
        should_autofill_year_month=should_autofill_year_month,
    )

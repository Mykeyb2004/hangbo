from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_SHEET_NAME = "问卷数据"
YEAR_HEADER = "年份"
MONTH_HEADER = "月份"


@dataclass(frozen=True)
class FileUpdateResult:
    path: Path
    status: str
    updated_rows: int
    error_message: str | None = None


@dataclass(frozen=True)
class DirectoryUpdateSummary:
    input_dir: Path
    file_results: tuple[FileUpdateResult, ...]

    @property
    def updated_count(self) -> int:
        return sum(1 for item in self.file_results if item.status == "updated")

    @property
    def skipped_count(self) -> int:
        return sum(1 for item in self.file_results if item.status != "updated")


def iter_excel_paths(input_dir: Path, recursive: bool = False) -> list[Path]:
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    paths = sorted(path for path in input_dir.glob(pattern) if path.is_file())
    return [
        path
        for path in paths
        if not path.name.startswith("~$")
        and not path.name.startswith("._")
    ]


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_header_index(worksheet) -> dict[str, int]:
    header_index: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header_value = normalize_header(worksheet.cell(row=1, column=column_index).value)
        if header_value:
            header_index[header_value] = column_index
    return header_index


def ensure_header_column(worksheet, header_name: str, column_index: int) -> int:
    worksheet.cell(row=1, column=column_index, value=header_name)
    return column_index


def resolve_target_columns(worksheet) -> tuple[int, int]:
    header_index = build_header_index(worksheet)
    last_header_column = max(header_index.values(), default=0)
    next_column_index = last_header_column + 1 if last_header_column else 1

    year_column = header_index.get(YEAR_HEADER)
    if year_column is None:
        year_column = ensure_header_column(worksheet, YEAR_HEADER, next_column_index)
        next_column_index += 1

    month_column = header_index.get(MONTH_HEADER)
    if month_column is None:
        month_column = ensure_header_column(worksheet, MONTH_HEADER, next_column_index)

    return year_column, month_column


def apply_year_month_to_workbook(
    workbook_path: Path,
    *,
    year: str,
    month: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> FileUpdateResult:
    try:
        workbook = load_workbook(workbook_path)
    except Exception as exc:  # pragma: no cover - 防御性分支
        return FileUpdateResult(
            path=workbook_path,
            status="read_error",
            updated_rows=0,
            error_message=str(exc),
        )

    if sheet_name not in workbook.sheetnames:
        return FileUpdateResult(
            path=workbook_path,
            status="missing_sheet",
            updated_rows=0,
        )

    worksheet = workbook[sheet_name]
    year_column, month_column = resolve_target_columns(worksheet)

    updated_rows = max(worksheet.max_row - 1, 0)
    for row_index in range(2, worksheet.max_row + 1):
        year_cell = worksheet.cell(row=row_index, column=year_column)
        month_cell = worksheet.cell(row=row_index, column=month_column)
        year_cell.value = str(year)
        month_cell.value = str(month)
        year_cell.number_format = "@"
        month_cell.number_format = "@"

    workbook.save(workbook_path)
    return FileUpdateResult(
        path=workbook_path,
        status="updated",
        updated_rows=updated_rows,
    )


def apply_year_month_to_directory(
    input_dir: Path,
    *,
    year: str,
    month: str,
    recursive: bool = False,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> DirectoryUpdateSummary:
    file_results = tuple(
        apply_year_month_to_workbook(path, year=year, month=month, sheet_name=sheet_name)
        for path in iter_excel_paths(input_dir, recursive=recursive)
    )
    return DirectoryUpdateSummary(
        input_dir=input_dir.resolve(),
        file_results=file_results,
    )


def describe_file_result(result: FileUpdateResult, *, sheet_name: str = DEFAULT_SHEET_NAME) -> str:
    if result.status == "updated":
        return f"已更新 {result.updated_rows} 行"
    if result.status == "missing_sheet":
        return f"缺少 {sheet_name} sheet"
    if result.status == "read_error":
        return f"读取失败（{result.error_message or '未知错误'}）"
    return result.status


def format_directory_summary(
    summary: DirectoryUpdateSummary,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> str:
    skipped_results = [item for item in summary.file_results if item.status != "updated"]
    lines = [
        f"目录: {summary.input_dir}",
        f"扫描文件数: {len(summary.file_results)}",
        f"更新成功: {summary.updated_count}",
        f"跳过/失败: {summary.skipped_count}",
        "",
        "文件明细：",
    ]

    if not summary.file_results:
        lines.append("- 未找到可处理的 xlsx 文件")
        return "\n".join(lines)

    for item in summary.file_results:
        lines.append(f"- {item.path.name}: {describe_file_result(item, sheet_name=sheet_name)}")

    if skipped_results:
        lines.append("")
        lines.append("跳过文件：")
        for item in skipped_results:
            lines.append(f"- {item.path.name}: {describe_file_result(item, sheet_name=sheet_name)}")
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="为指定目录下 xlsx 文件的问卷数据 sheet 写入年份、月份两列。若列已存在，则覆盖原值。"
    )
    parser.add_argument("--input-dir", type=Path, required=True, help="要处理的目录")
    parser.add_argument("--year", required=True, help="年份文本值，例如 2026")
    parser.add_argument("--month", required=True, help="月份文本值，例如 02 或 2月")
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="是否递归处理子目录中的 xlsx 文件",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"要写入的 sheet 名，默认 {DEFAULT_SHEET_NAME}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    summary = apply_year_month_to_directory(
        args.input_dir,
        year=str(args.year),
        month=str(args.month),
        recursive=args.recursive,
        sheet_name=args.sheet_name,
    )
    print(format_directory_summary(summary, sheet_name=args.sheet_name))


if __name__ == "__main__":
    main()

from __future__ import annotations

import argparse
import tomllib
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from phase_column_preprocess import preprocess_phase_column_if_needed
from survey_customer_category_rules import (
    CUSTOMER_CATEGORY_RULE_BY_NAME,
    CUSTOMER_CATEGORY_RULES,
    CustomerCategoryRule,
)

DEFAULT_SHEET_NAME = "问卷数据"
DEFAULT_OUTPUT_FORMAT = "xlsx"
DEFAULT_ROLE_COLUMN = "E"
VALID_OUTPUT_FORMATS = ("xlsx", "csv", "md")
DEFAULT_CALCULATION_MODE = "template"
VALID_CALCULATION_MODES = ("template", "summary")

ORGANIZER_ROLE_NAME = "展览主承办"
EXHIBITOR_ROLE_NAME = "参展商"
VISITOR_ROLE_NAME = "专业观众"
SERVICE_PROVIDER_ROLE_NAME = "会展服务商"
MEETING_ORGANIZER_ROLE_NAME = "会议主承办"
HOTEL_MEETING_ORGANIZER_ROLE_NAME = "酒店会议主承办"
HOTEL_MEETING_ATTENDEE_ROLE_NAME = "酒店参会客户"
MEETING_ATTENDEE_ROLE_NAME = "参会人员"
MEETING_CATEGORY_COLUMN = "C"
MEETING_CATEGORY_NAME = "会议"
HOTEL_MEETING_CATEGORY_NAME = "酒店会议"
TRAVEL_STAFF_ROLE_NAME = "旅行社工作人员"
TOURIST_ROLE_NAME = "游客"
HOTEL_INDIVIDUAL_GUEST_ROLE_NAME = "散客"
HOTEL_GROUP_GUEST_ROLE_NAME = "住宿团队"
CATERING_FOOD_HALL_ROLE_NAME = "特色美食廊"
CATERING_BUSINESS_MEAL_ROLE_NAME = "商务简餐"
CATERING_TOUR_MEAL_ROLE_NAME = "旅游团餐"
CATERING_BANQUET_ROLE_NAME = "宴会"
CATERING_WEDDING_BANQUET_ROLE_NAME = "婚宴"
CATERING_BUFFET_ROLE_NAME = "自助餐"
CATERING_HOTEL_BANQUET_ROLE_NAME = "酒店宴会"
CATERING_HOTEL_BUFFET_ROLE_NAME = "酒店自助餐"
HOTEL_CATERING_BUSINESS_MEAL_COMPONENT_ROLE_NAME = "酒店餐饮-商务简餐"
HOTEL_CATERING_BANQUET_COMPONENT_ROLE_NAME = "酒店餐饮-宴会"
HOTEL_ROLE_COLUMN = "C"
CATERING_ROLE_COLUMN = "D"
TOURISM_ROLE_COLUMN = "C"

OVERALL_FILL = PatternFill(fill_type="solid", start_color="F4B183", end_color="F4B183")
SECTION_FILL = PatternFill(fill_type="solid", start_color="C6EFD1", end_color="C6EFD1")
EMPHASIS_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


@dataclass(frozen=True)
class MetricDefinition:
    name: str
    satisfaction_column: str
    importance_column: str
    satisfaction_gt_zero_column: str | None = None
    satisfaction_lt_eleven_column: str | None = None
    importance_gt_zero_column: str | None = None
    importance_lt_eleven_column: str | None = None


@dataclass(frozen=True)
class SectionDefinition:
    name: str
    metrics: tuple[MetricDefinition, ...]


@dataclass(frozen=True)
class MatchCondition:
    column: str
    expected_value: str | tuple[str, ...]


@dataclass(frozen=True)
class RoleDefinition:
    role_name: str
    role_column: str
    sections: tuple[SectionDefinition, ...]
    role_match_value: str | None = None
    role_match_values: tuple[str, ...] | None = None
    row_conditions: tuple[MatchCondition, ...] = ()


@dataclass(frozen=True)
class JobConfig:
    name: str
    path: Path
    sheet_name: str
    template_name: str
    role_name: str
    output_name: str
    output_format: str | None = None
    category_rule_name: str | None = None


@dataclass(frozen=True)
class SourceFileOverride:
    standard_file_name: str
    actual_file_name: str


@dataclass(frozen=True)
class BatchConfig:
    config_path: Path
    output_dir: Path
    output_format: str
    calculation_mode: str
    sheet_name: str
    jobs: tuple[JobConfig, ...] = ()
    input_dir: Path | None = None
    source_file_overrides: tuple[SourceFileOverride, ...] = ()


@dataclass(frozen=True)
class MetricResult:
    name: str
    satisfaction: float | None
    importance: float | None


@dataclass(frozen=True)
class SectionResult:
    name: str
    satisfaction: float | None
    importance: float | None
    metrics: tuple[MetricResult, ...]


@dataclass(frozen=True)
class SurveyStatistics:
    role_name: str
    satisfaction: float | None
    importance: float | None
    sections: tuple[SectionResult, ...]
    matched_row_count: int


@dataclass(frozen=True)
class GeneratedReport:
    result_df: pd.DataFrame
    output_path: Path
    stats: SurveyStatistics
    preprocess_notice: str | None = None


@dataclass(frozen=True)
class MissingGroupNotice:
    job_name: str
    input_path: Path
    sheet_name: str


DIRECTORY_NOTICE_REASON_MISSING_SOURCE_FILE = "missing_source_file"
DIRECTORY_NOTICE_REASON_MISSING_ROLE_DATA = "missing_role_data"


@dataclass(frozen=True)
class MissingCustomerTypeNotice:
    customer_type_name: str
    source_reference: str
    sheet_name: str
    reason: str


@dataclass(frozen=True)
class PreprocessNoticeRecord:
    input_path: Path
    notice: str


@dataclass(frozen=True)
class UnmappedCustomerCategoryNotice:
    source_file_name: str
    auxiliary_value: str | None
    data_value: str
    row_count: int


@dataclass(frozen=True)
class DirectoryDiscoveryResult:
    jobs: tuple[JobConfig, ...]
    missing_customer_type_notices: tuple[MissingCustomerTypeNotice, ...]
    preprocess_notices: tuple[PreprocessNoticeRecord, ...]
    unmapped_customer_category_notices: tuple[UnmappedCustomerCategoryNotice, ...] = ()


ORGANIZER_TEMPLATE = RoleDefinition(
    role_name=ORGANIZER_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "会展服务",
            (
                MetricDefinition("销售经理服务态度", "AE", "AF"),
                MetricDefinition("销售经理业务能力", "AG", "AH"),
                MetricDefinition("现场对接协调沟通", "AI", "AJ"),
                MetricDefinition("现场对接服务效率", "AK", "AL"),
                MetricDefinition("工作人员仪容仪表", "AM", "AN"),
                MetricDefinition("工作人员服务态度", "AO", "AP"),
                MetricDefinition("工作人员业务技能", "AQ", "AR"),
                MetricDefinition("报馆流程及服务", "AW", "AX"),
                MetricDefinition("现场布置与搭建", "BA", "BB"),
                MetricDefinition("在线服务渠道便捷", "AY", "AZ"),
                MetricDefinition("撤展服务", "BC", "BD"),
                MetricDefinition("展后回访", "BE", "BF"),
                MetricDefinition("整体协调和配合", "AU", "AV"),
                MetricDefinition("服务响应及时性", "AS", "AT"),
                MetricDefinition("设备租赁", "AA", "AB"),
            ),
        ),
        SectionDefinition(
            "硬件设施",
            (
                MetricDefinition("园区停车方便", "O", "P"),
                MetricDefinition("交通便利，容易到达", "M", "N"),
                MetricDefinition("交通流线", "K", "L"),
                MetricDefinition("货运通道", "U", "V"),
                MetricDefinition("标识标牌清晰", "S", "T"),
                MetricDefinition("设施设备齐全", "W", "X"),
                MetricDefinition("展厅使用情况", "Y", "Z"),
            ),
        ),
        SectionDefinition(
            "配套服务",
            (
                MetricDefinition("餐饮服务", "BG", "BH"),
                MetricDefinition("客房服务", "BI", "BJ"),
                MetricDefinition("安保服务", "BK", "BL"),
                MetricDefinition("保洁服务", "BM", "BN"),
            ),
        ),
        SectionDefinition(
            "智慧场馆",
            (
                MetricDefinition("杭州国博APP", "BR", "BS"),
                MetricDefinition("室内导航系统", "BU", "BV"),
                MetricDefinition("寻车系统", "BX", "BY"),
                MetricDefinition("云上看馆", "CA", "CB"),
            ),
        ),
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "展览"),),
)

EXHIBITOR_TEMPLATE = RoleDefinition(
    role_name=EXHIBITOR_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "会场服务",
            (
                MetricDefinition("工作人员仪容仪表", "AM", "AN"),
                MetricDefinition("工作人员服务态度", "AO", "AP"),
                MetricDefinition("工作人员业务技能", "AQ", "AR"),
                MetricDefinition("现场布置与搭建", "BA", "BB"),
                MetricDefinition("接待引导服务", "BO", "BP"),
            ),
        ),
        SectionDefinition(
            "硬件设施",
            (
                MetricDefinition("园区停车方便", "O", "P"),
                MetricDefinition("交通便利，容易到达", "M", "N"),
                MetricDefinition("标识标牌清晰", "S", "T"),
                MetricDefinition("设施设备齐全", "W", "X"),
                MetricDefinition("展厅使用情况", "Y", "Z"),
                MetricDefinition("参展环境", "AC", "AD"),
            ),
        ),
        SectionDefinition(
            "配套服务",
            (
                MetricDefinition("餐饮服务", "BG", "BH"),
                MetricDefinition("客房服务", "BI", "BJ"),
                MetricDefinition("安保服务", "BK", "BL"),
                MetricDefinition("保洁服务", "BM", "BN"),
            ),
        ),
        SectionDefinition(
            "智慧场馆",
            (
                MetricDefinition("杭州国博APP", "BR", "BS"),
                MetricDefinition("室内导航系统", "BU", "BV"),
                MetricDefinition("寻车系统", "BX", "BY"),
                MetricDefinition("云上看馆", "CA", "CB"),
            ),
        ),
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "展览"),),
)

VISITOR_TEMPLATE = RoleDefinition(
    role_name=VISITOR_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "会展服务",
            (
                MetricDefinition("工作人员仪容仪表", "AM", "AN"),
                MetricDefinition("工作人员服务态度", "AO", "AP"),
                MetricDefinition("工作人员业务技能", "AQ", "AR"),
                MetricDefinition("接待引导服务", "BO", "BP"),
            ),
        ),
        SectionDefinition(
            "硬件设施",
            (
                MetricDefinition("展会路线安排", "Q", "R"),
                MetricDefinition("园区停车方便", "O", "P"),
                MetricDefinition("交通便利，容易到达", "M", "N"),
                MetricDefinition("标识标牌清晰", "S", "T"),
                MetricDefinition("设施设备齐全", "W", "X"),
                MetricDefinition("展厅使用情况", "Y", "Z"),
                MetricDefinition("参展环境", "AC", "AD"),
            ),
        ),
        SectionDefinition(
            "配套服务",
            (
                MetricDefinition("餐饮服务", "BG", "BH"),
                MetricDefinition("客房服务", "BI", "BJ"),
                MetricDefinition("安保服务", "BK", "BL"),
                MetricDefinition("保洁服务", "BM", "BN"),
            ),
        ),
        SectionDefinition(
            "智慧场馆",
            (
                MetricDefinition("杭州国博APP", "BR", "BS"),
                MetricDefinition("室内导航系统", "BU", "BV"),
                MetricDefinition("寻车系统", "BX", "BY"),
                MetricDefinition("云上看馆", "CA", "CB"),
            ),
        ),
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "展览"),),
)

SERVICE_PROVIDER_TEMPLATE = RoleDefinition(
    role_name=SERVICE_PROVIDER_ROLE_NAME,
    role_column="D",
    sections=(
        SectionDefinition(
            "会展服务",
            (
                MetricDefinition("现场对接协调沟通", "AN", "AO"),
                MetricDefinition("现场对接服务效率", "AL", "AM"),
                MetricDefinition("工作人员仪容仪表", "AF", "AG"),
                MetricDefinition(
                    "工作人员服务态度",
                    "AH",
                    "AI",
                    satisfaction_gt_zero_column="AH",
                    satisfaction_lt_eleven_column="AL",
                ),
                MetricDefinition("工作人员业务技能", "AJ", "AK"),
                MetricDefinition("现场布置与搭建", "Z", "AA"),
                MetricDefinition("报馆流程及服务", "AB", "AC"),
                MetricDefinition("在线服务渠道便捷", "AD", "AE"),
                MetricDefinition("工程设备服务", "X", "Y"),
                MetricDefinition("整体协调和配合", "AR", "AS"),
                MetricDefinition("服务响应及时性", "AP", "AQ"),
                MetricDefinition("撤展服务", "AT", "AU"),
                MetricDefinition("展后回访", "AV", "AX"),
            ),
        ),
        SectionDefinition(
            "硬件设施",
            (
                MetricDefinition("交通便利，容易到达", "H", "I"),
                MetricDefinition("标识标牌清晰", "R", "S"),
                MetricDefinition("交通流线", "L", "M"),
                MetricDefinition("园区停车便利", "J", "K"),
                MetricDefinition("货运通道", "N", "O"),
                MetricDefinition("快递物流", "P", "Q"),
                MetricDefinition("休息空间", "V", "W"),
                MetricDefinition("设施设备齐全", "T", "U"),
            ),
        ),
        SectionDefinition(
            "配套服务",
            (
                MetricDefinition("餐饮服务", "AX", "AY"),
                MetricDefinition("客房服务", "AZ", "BA"),
                MetricDefinition("安保服务", "BB", "BC"),
                MetricDefinition("保洁服务", "BD", "BE"),
            ),
        ),
        SectionDefinition(
            "智慧场馆",
            (
                MetricDefinition("室内导航系统", "BJ", "BK"),
                MetricDefinition("寻车系统", "BM", "BN"),
                MetricDefinition("杭州国博APP", "BG", "BH"),
                MetricDefinition("云上看馆", "BP", "BQ"),
            ),
        ),
    ),
)

MEETING_ORGANIZER_SERVICE_SECTION = SectionDefinition(
    "会展服务",
    (
        MetricDefinition("销售经理服务态度", "AS", "AT"),
        MetricDefinition("销售经理业务能力", "AU", "AV"),
        MetricDefinition("现场对接协调沟通", "AW", "AX"),
        MetricDefinition("现场对接服务效率", "AY", "AZ"),
        MetricDefinition("工作人员仪容仪表", "AE", "AF"),
        MetricDefinition("工作人员服务态度", "AI", "AJ"),
        MetricDefinition("工作人员业务技能", "AG", "AH"),
        MetricDefinition("现场布置与搭建", "BC", "BD"),
        MetricDefinition("茶歇服务与品质", "BG", "BH"),
        MetricDefinition("在线服务渠道便捷", "BI", "BJ"),
        MetricDefinition("报馆流程及服务", "BA", "BB"),
        MetricDefinition("工程设施设备", "BE", "BF"),
        MetricDefinition("撤会服务", "BO", "BP"),
        MetricDefinition("会后回访", "BQ", "BR"),
        MetricDefinition("整体协调和配合", "BM", "BN"),
        MetricDefinition("服务响应及时性", "BK", "BL"),
    ),
)

MEETING_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("交通便利，容易到达", "O", "P"),
        MetricDefinition("标识标牌清晰", "S", "T"),
        MetricDefinition("会议厅室匹配性，方便性", "W", "X"),
        MetricDefinition("设施设备齐全", "U", "V"),
        MetricDefinition("休息空间", "AA", "AB"),
        MetricDefinition("园区停车方便", "Q", "R"),
        MetricDefinition("AV效果", "Y", "Z"),
        MetricDefinition("交通流线", "M", "N"),
    ),
)

MEETING_SUPPORT_SECTION = SectionDefinition(
    "配套服务",
    (
        MetricDefinition("餐饮服务", "AK", "AL"),
        MetricDefinition("客房服务", "AM", "AN"),
        MetricDefinition("安保服务", "AQ", "AR"),
        MetricDefinition("保洁服务", "AO", "AP"),
    ),
)

MEETING_SMART_SECTION = SectionDefinition(
    "智慧场馆",
    (
        MetricDefinition("杭州国博APP", "BV", "BW"),
        MetricDefinition("室内导航系统", "BY", "BZ"),
        MetricDefinition("寻车系统", "CB", "CC"),
        MetricDefinition("云上看馆", "CE", "CF"),
    ),
)

MEETING_ATTENDEE_SERVICE_SECTION = SectionDefinition(
    "会展服务",
    (
        MetricDefinition("工作人员仪容仪表", "AE", "AF"),
        MetricDefinition("工作人员业务技能", "AG", "AH"),
        MetricDefinition("工作人员服务态度", "AI", "AJ"),
        MetricDefinition("接待引导服务", "BS", "BT"),
        MetricDefinition("茶歇服务品质", "BG", "BH"),
    ),
)

MEETING_ATTENDEE_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("交通便利，容易到达", "O", "P"),
        MetricDefinition("标识标牌清晰", "S", "T"),
        MetricDefinition("会议厅室匹配性，方便性", "W", "X"),
        MetricDefinition("设施设备齐全", "U", "V"),
        MetricDefinition("休息空间", "AA", "AB"),
        MetricDefinition("园区停车方便", "Q", "R"),
        MetricDefinition("AV效果", "Y", "Z"),
        MetricDefinition("参会环境", "AC", "AD"),
    ),
)

MEETING_ORGANIZER_TEMPLATE = RoleDefinition(
    role_name=MEETING_ORGANIZER_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        MEETING_ORGANIZER_SERVICE_SECTION,
        MEETING_HARDWARE_SECTION,
        MEETING_SUPPORT_SECTION,
        MEETING_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, MEETING_CATEGORY_NAME),),
)

HOTEL_MEETING_ORGANIZER_TEMPLATE = RoleDefinition(
    role_name=HOTEL_MEETING_ORGANIZER_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        MEETING_ORGANIZER_SERVICE_SECTION,
        MEETING_HARDWARE_SECTION,
        MEETING_SUPPORT_SECTION,
        MEETING_SMART_SECTION,
    ),
    role_match_value=HOTEL_MEETING_ORGANIZER_ROLE_NAME,
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, HOTEL_MEETING_CATEGORY_NAME),),
)

HOTEL_MEETING_ATTENDEE_TEMPLATE = RoleDefinition(
    role_name=HOTEL_MEETING_ATTENDEE_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        MEETING_ATTENDEE_SERVICE_SECTION,
        MEETING_ATTENDEE_HARDWARE_SECTION,
        MEETING_SUPPORT_SECTION,
        MEETING_SMART_SECTION,
    ),
    role_match_value=HOTEL_MEETING_ATTENDEE_ROLE_NAME,
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, HOTEL_MEETING_CATEGORY_NAME),),
)

MEETING_ATTENDEE_TEMPLATE = RoleDefinition(
    role_name=MEETING_ATTENDEE_ROLE_NAME,
    role_column=DEFAULT_ROLE_COLUMN,
    sections=(
        MEETING_ATTENDEE_SERVICE_SECTION,
        MEETING_ATTENDEE_HARDWARE_SECTION,
        MEETING_SUPPORT_SECTION,
        MEETING_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, MEETING_CATEGORY_NAME),),
)

TOURISM_SERVICE_SECTION = SectionDefinition(
    "旅游服务",
    (
        MetricDefinition("售票/销售服务", "T", "U"),
        MetricDefinition("景点设施", "V", "W"),
        MetricDefinition("环境卫生", "X", "Y"),
        MetricDefinition("安全服务", "Z", "AA"),
        MetricDefinition("工作人员仪容仪表", "AB", "AC"),
        MetricDefinition("工作人员服务态度", "AD", "AE"),
        MetricDefinition("讲解人员专业性", "AF", "AG"),
    ),
)

TOURISM_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("交通便利，容易到达", "N", "O"),
        MetricDefinition("园区停车方便", "P", "Q"),
        MetricDefinition("标识标牌清晰", "R", "S"),
    ),
)

TOURISM_SMART_SECTION = SectionDefinition(
    "智慧服务",
    (
        MetricDefinition("语音导览", "AI", "AJ"),
        MetricDefinition("AR导览", "AL", "AM"),
        MetricDefinition("自助售卖", "AO", "AP"),
        MetricDefinition("线上商城", "AR", "AS"),
    ),
)

TRAVEL_STAFF_TEMPLATE = RoleDefinition(
    role_name=TRAVEL_STAFF_ROLE_NAME,
    role_column=TOURISM_ROLE_COLUMN,
    sections=(
        TOURISM_SERVICE_SECTION,
        TOURISM_HARDWARE_SECTION,
        TOURISM_SMART_SECTION,
    ),
)

TOURIST_TEMPLATE = RoleDefinition(
    role_name=TOURIST_ROLE_NAME,
    role_column=TOURISM_ROLE_COLUMN,
    sections=(
        TOURISM_SERVICE_SECTION,
        TOURISM_HARDWARE_SECTION,
        TOURISM_SMART_SECTION,
    ),
)

HOTEL_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("交通便利，容易到达", "G", "H"),
        MetricDefinition("标识标牌清晰", "K", "L"),
        MetricDefinition("适合商务旅行", "M", "N"),
        MetricDefinition("园区停车方便", "I", "J"),
        MetricDefinition("客房整体环境", "O", "P"),
        MetricDefinition("客房设施设备", "Q", "R"),
        MetricDefinition("住宿用品质量", "S", "T"),
        MetricDefinition("客房私密性", "U", "V"),
    ),
)

HOTEL_CHECKIN_SERVICE_SECTION = SectionDefinition(
    "入住服务",
    (
        MetricDefinition("礼宾服务", "W", "X"),
        MetricDefinition("入住登记", "Y", "Z"),
        MetricDefinition("客房服务", "AA", "AB"),
        MetricDefinition("离店退房", "AC", "AD"),
        MetricDefinition("工作人员仪容仪表", "AI", "AJ"),
        MetricDefinition("工作人员服务态度", "AG", "AH"),
        MetricDefinition("工作人员业务技能", "AE", "AF"),
    ),
)

HOTEL_DINING_SECTION = SectionDefinition(
    "餐饮服务",
    (
        MetricDefinition("自助早餐", "AK", "AL"),
        MetricDefinition("自助晚餐", "AM", "AN"),
        MetricDefinition("送餐服务", "AO", "AP"),
        MetricDefinition("零点服务", "AQ", "AR"),
    ),
)

HOTEL_SMART_SECTION = SectionDefinition(
    "智慧场馆",
    (
        MetricDefinition("杭州国博APP", "AT", "AU"),
        MetricDefinition("室内导航系统", "AW", "AX"),
        MetricDefinition("寻车系统", "AZ", "BA"),
        MetricDefinition("云上看馆", "BC", "BD"),
    ),
)

HOTEL_INDIVIDUAL_GUEST_TEMPLATE = RoleDefinition(
    role_name=HOTEL_INDIVIDUAL_GUEST_ROLE_NAME,
    role_column=HOTEL_ROLE_COLUMN,
    sections=(
        HOTEL_HARDWARE_SECTION,
        HOTEL_CHECKIN_SERVICE_SECTION,
        HOTEL_DINING_SECTION,
        HOTEL_SMART_SECTION,
    ),
)

HOTEL_GROUP_GUEST_TEMPLATE = RoleDefinition(
    role_name=HOTEL_GROUP_GUEST_ROLE_NAME,
    role_column=HOTEL_ROLE_COLUMN,
    sections=(
        HOTEL_HARDWARE_SECTION,
        HOTEL_CHECKIN_SERVICE_SECTION,
        HOTEL_DINING_SECTION,
        HOTEL_SMART_SECTION,
    ),
)

CATERING_BASIC_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("园区停车方便", "I", "J"),
        MetricDefinition("交通便利，容易到达", "G", "H"),
        MetricDefinition("标识标牌清晰", "K", "L"),
    ),
)

CATERING_STANDARD_SMART_SECTION = SectionDefinition(
    "智慧场馆",
    (
        MetricDefinition("杭州国博APP", "AV", "AW"),
        MetricDefinition("寻车系统", "BB", "BC"),
        MetricDefinition("室内导航系统", "AY", "AZ"),
        MetricDefinition("云上看馆", "BE", "BF"),
    ),
)

CATERING_BUFFET_SMART_SECTION = SectionDefinition(
    "智慧场馆",
    (
        MetricDefinition("杭州国博APP", "AV", "AW"),
        MetricDefinition("室内导航系统", "AY", "AZ"),
        MetricDefinition("寻车系统", "BB", "BC"),
        MetricDefinition("云上看馆", "BE", "BF"),
    ),
)

CATERING_BANQUET_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("标识标牌清晰", "K", "L"),
        MetricDefinition("宴会视听设备", "Q", "R"),
        MetricDefinition("园区停车方便", "I", "J"),
        MetricDefinition("交通便利，容易到达", "G", "H"),
    ),
)

CATERING_WEDDING_HARDWARE_SECTION = SectionDefinition(
    "硬件设施",
    (
        MetricDefinition("园区停车方便", "I", "J"),
        MetricDefinition("交通便利，容易到达", "G", "H"),
        MetricDefinition("标识标牌清晰", "K", "L"),
        MetricDefinition("宴会视听设备", "Q", "R"),
    ),
)

CATERING_FOOD_HALL_TEMPLATE = RoleDefinition(
    role_name=CATERING_FOOD_HALL_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜肴品质", "Y", "Z"),
                MetricDefinition("菜品口味口相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("菜肴份量", "AI", "AJ"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
            ),
        ),
        CATERING_BASIC_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_BUSINESS_MEAL_TEMPLATE = RoleDefinition(
    role_name=CATERING_BUSINESS_MEAL_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜品温度", "AK", "AL"),
                MetricDefinition("菜品分量", "AI", "AJ"),
                MetricDefinition("供应速度", "AC", "AD"),
                MetricDefinition("菜品卫生", "AE", "AF"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
            ),
        ),
        CATERING_BASIC_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_TOUR_MEAL_TEMPLATE = RoleDefinition(
    role_name=CATERING_TOUR_MEAL_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜品温度", "AK", "AL"),
                MetricDefinition("菜品分量", "AI", "AJ"),
                MetricDefinition("供应速度", "AC", "AD"),
                MetricDefinition("菜品卫生", "AE", "AF"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
            ),
        ),
        CATERING_BASIC_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_BANQUET_TEMPLATE = RoleDefinition(
    role_name=CATERING_BANQUET_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜品温度", "AK", "AL"),
                MetricDefinition("菜肴品种", "AG", "AH"),
                MetricDefinition("菜品供应速度", "AC", "AD"),
                MetricDefinition("菜品口味品相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("接待指引", "S", "T"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
            ),
        ),
        CATERING_BANQUET_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_WEDDING_BANQUET_TEMPLATE = RoleDefinition(
    role_name=CATERING_WEDDING_BANQUET_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("接待指引", "S", "T"),
                MetricDefinition("菜品口味品相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("菜品供应速度", "AC", "AD"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
                MetricDefinition("工作人员仪容仪表", "AQ", "AR"),
                MetricDefinition("工作人员服务态度", "AO", "AP"),
                MetricDefinition("工作人员业务技能", "AS", "AT"),
                MetricDefinition("婚宴茶歇", "W", "X"),
                MetricDefinition("菜品温度", "AK", "AL"),
            ),
        ),
        CATERING_WEDDING_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_BUFFET_TEMPLATE = RoleDefinition(
    role_name=CATERING_BUFFET_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜肴品种", "AG", "AH"),
                MetricDefinition("菜品供应速度", "AC", "AD"),
                MetricDefinition("菜品口味口相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("接待指引", "S", "T"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
            ),
        ),
        CATERING_BASIC_HARDWARE_SECTION,
        CATERING_BUFFET_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "餐饮"),),
)

CATERING_HOTEL_BANQUET_TEMPLATE = RoleDefinition(
    role_name=CATERING_HOTEL_BANQUET_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜品温度", "AK", "AL"),
                MetricDefinition("菜肴品种", "AG", "AH"),
                MetricDefinition("菜品供应速度", "AC", "AD"),
                MetricDefinition("菜品口味品相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("接待指引", "S", "T"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
            ),
        ),
        CATERING_BANQUET_HARDWARE_SECTION,
        CATERING_STANDARD_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "酒店餐饮"),),
)

CATERING_HOTEL_BUFFET_TEMPLATE = RoleDefinition(
    role_name=CATERING_HOTEL_BUFFET_ROLE_NAME,
    role_column=CATERING_ROLE_COLUMN,
    sections=(
        SectionDefinition(
            "餐饮服务",
            (
                MetricDefinition("菜肴品种", "AG", "AH"),
                MetricDefinition("菜品供应速度", "AC", "AD"),
                MetricDefinition("菜品口味口相", "AA", "AB"),
                MetricDefinition("菜品地方特色", "U", "V"),
                MetricDefinition("接待指引", "S", "T"),
                MetricDefinition("工作人员服务品质", "AM", "AN"),
                MetricDefinition("就餐区域温度", "M", "N"),
                MetricDefinition("就餐区域卫生", "O", "P"),
            ),
        ),
        CATERING_BASIC_HARDWARE_SECTION,
        CATERING_BUFFET_SMART_SECTION,
    ),
    row_conditions=(MatchCondition(MEETING_CATEGORY_COLUMN, "酒店餐饮"),),
)

TEMPLATE_DEFINITIONS: dict[str, RoleDefinition] = {
    "organizer": ORGANIZER_TEMPLATE,
    "exhibitor": EXHIBITOR_TEMPLATE,
    "visitor": VISITOR_TEMPLATE,
    "service_provider": SERVICE_PROVIDER_TEMPLATE,
    "meeting_organizer": MEETING_ORGANIZER_TEMPLATE,
    "hotel_meeting_organizer": HOTEL_MEETING_ORGANIZER_TEMPLATE,
    "hotel_meeting_attendee": HOTEL_MEETING_ATTENDEE_TEMPLATE,
    "meeting_attendee": MEETING_ATTENDEE_TEMPLATE,
    "travel_staff": TRAVEL_STAFF_TEMPLATE,
    "tourist": TOURIST_TEMPLATE,
    "hotel_individual_guest": HOTEL_INDIVIDUAL_GUEST_TEMPLATE,
    "hotel_group_guest": HOTEL_GROUP_GUEST_TEMPLATE,
    "catering_food_hall": CATERING_FOOD_HALL_TEMPLATE,
    "catering_business_meal": CATERING_BUSINESS_MEAL_TEMPLATE,
    "catering_tour_meal": CATERING_TOUR_MEAL_TEMPLATE,
    "catering_banquet": CATERING_BANQUET_TEMPLATE,
    "catering_wedding_banquet": CATERING_WEDDING_BANQUET_TEMPLATE,
    "catering_buffet": CATERING_BUFFET_TEMPLATE,
    "catering_hotel_banquet": CATERING_HOTEL_BANQUET_TEMPLATE,
    "catering_hotel_buffet": CATERING_HOTEL_BUFFET_TEMPLATE,
}

SUMMARY_EVENT_ROLE_NAMES = frozenset(
    {
        ORGANIZER_ROLE_NAME,
        EXHIBITOR_ROLE_NAME,
        VISITOR_ROLE_NAME,
        SERVICE_PROVIDER_ROLE_NAME,
        MEETING_ORGANIZER_ROLE_NAME,
        HOTEL_MEETING_ORGANIZER_ROLE_NAME,
        HOTEL_MEETING_ATTENDEE_ROLE_NAME,
        MEETING_ATTENDEE_ROLE_NAME,
    }
)
SUMMARY_HOTEL_ROLE_NAMES = frozenset(
    {
        HOTEL_INDIVIDUAL_GUEST_ROLE_NAME,
        HOTEL_GROUP_GUEST_ROLE_NAME,
    }
)
SUMMARY_CATERING_ROLE_NAMES = frozenset(
    {
        CATERING_FOOD_HALL_ROLE_NAME,
        CATERING_BUSINESS_MEAL_ROLE_NAME,
        CATERING_TOUR_MEAL_ROLE_NAME,
        CATERING_BANQUET_ROLE_NAME,
        CATERING_WEDDING_BANQUET_ROLE_NAME,
        CATERING_BUFFET_ROLE_NAME,
        CATERING_HOTEL_BANQUET_ROLE_NAME,
        CATERING_HOTEL_BUFFET_ROLE_NAME,
        HOTEL_CATERING_BUSINESS_MEAL_COMPONENT_ROLE_NAME,
        HOTEL_CATERING_BANQUET_COMPONENT_ROLE_NAME,
    }
)


def excel_column_to_index(column_name: str) -> int:
    index = 0
    for char in column_name.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"非法 Excel 列名: {column_name}")
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def excel_round(value: float | int | Decimal | None, digits: int = 2) -> float | None:
    if value is None or pd.isna(value):
        return None

    quantizer = Decimal("1").scaleb(-digits)
    rounded = Decimal(str(value)).quantize(quantizer, rounding=ROUND_HALF_UP)
    return float(rounded)


def decimal_mean_ignore_empty(values: list[object]) -> Decimal | None:
    defined_values = [
        Decimal(str(value))
        for value in values
        if value is not None and not pd.isna(value)
    ]
    if not defined_values:
        return None
    return sum(defined_values) / Decimal(len(defined_values))


def mean_ignore_empty(values: list[float | None]) -> float | None:
    return excel_round(decimal_mean_ignore_empty(list(values)))


def format_value(value: float | None) -> str:
    if value is None or pd.isna(value):
        return ""
    text = f"{value:.2f}"
    return text.rstrip("0").rstrip(".")


def resolve_role_definition(template_name: str, role_name: str | None = None) -> RoleDefinition:
    template = TEMPLATE_DEFINITIONS.get(template_name)
    if template is None:
        supported = ", ".join(sorted(TEMPLATE_DEFINITIONS))
        raise ValueError(f"未知模板: {template_name}，支持的模板有: {supported}")
    resolved_role_name = role_name or template.role_name
    role_match_values = template.role_match_values
    role_match_value = template.role_match_value
    if role_name is not None and role_match_value is None and role_match_values is None:
        role_match_value = resolved_role_name
    return RoleDefinition(
        role_name=resolved_role_name,
        role_column=template.role_column,
        sections=template.sections,
        role_match_value=role_match_value,
        role_match_values=role_match_values,
        row_conditions=template.row_conditions,
    )


def build_role_definition_from_customer_category_rule(
    rule: CustomerCategoryRule,
) -> RoleDefinition:
    if rule.template_name is None:
        raise ValueError(f"{rule.name} 是聚合客户类别，不能直接构建单一模板定义。")

    template = TEMPLATE_DEFINITIONS.get(rule.template_name)
    if template is None:
        supported = ", ".join(sorted(TEMPLATE_DEFINITIONS))
        raise ValueError(f"未知模板: {rule.template_name}，支持的模板有: {supported}")

    row_conditions: tuple[MatchCondition, ...] = ()
    if rule.auxiliary_column and rule.auxiliary_values:
        expected_value: str | tuple[str, ...]
        if len(rule.auxiliary_values) == 1:
            expected_value = rule.auxiliary_values[0]
        else:
            expected_value = rule.auxiliary_values
        row_conditions = (MatchCondition(rule.auxiliary_column, expected_value),)

    return RoleDefinition(
        role_name=rule.name,
        role_column=rule.data_column or template.role_column,
        sections=template.sections,
        role_match_value=rule.data_values[0] if len(rule.data_values) == 1 else None,
        role_match_values=rule.data_values if len(rule.data_values) > 1 else None,
        row_conditions=row_conditions,
    )


def build_customer_category_rule_mask(
    df: pd.DataFrame,
    rule: CustomerCategoryRule,
) -> pd.Series:
    if rule.data_column is None or not rule.data_values:
        return pd.Series([False] * len(df), index=df.index)

    role_definition = RoleDefinition(
        role_name=rule.name,
        role_column=rule.data_column,
        sections=(),
        role_match_value=rule.data_values[0] if len(rule.data_values) == 1 else None,
        role_match_values=rule.data_values if len(rule.data_values) > 1 else None,
        row_conditions=(
            (
                MatchCondition(
                    rule.auxiliary_column,
                    rule.auxiliary_values[0]
                    if len(rule.auxiliary_values) == 1
                    else rule.auxiliary_values,
                ),
            )
            if rule.auxiliary_column and rule.auxiliary_values
            else ()
        ),
    )
    return build_role_mask(df, role_definition)


def unique_preserve_order(values: list[str]) -> tuple[str, ...]:
    ordered: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value in seen:
            continue
        ordered.append(value)
        seen.add(value)
    return tuple(ordered)


def merge_survey_statistics(
    role_name: str,
    stats_list: list[SurveyStatistics],
) -> SurveyStatistics:
    section_names = unique_preserve_order(
        [section.name for stats in stats_list for section in stats.sections]
    )
    section_results: list[SectionResult] = []

    for section_name in section_names:
        matched_sections = [
            section
            for stats in stats_list
            for section in stats.sections
            if section.name == section_name
        ]
        metric_names = unique_preserve_order(
            [metric.name for section in matched_sections for metric in section.metrics]
        )
        metric_results: list[MetricResult] = []
        for metric_name in metric_names:
            matched_metrics = [
                metric
                for section in matched_sections
                for metric in section.metrics
                if metric.name == metric_name
            ]
            metric_results.append(
                MetricResult(
                    name=metric_name,
                    satisfaction=mean_ignore_empty(
                        [metric.satisfaction for metric in matched_metrics]
                    ),
                    importance=mean_ignore_empty(
                        [metric.importance for metric in matched_metrics]
                    ),
                )
            )

        section_results.append(
            SectionResult(
                name=section_name,
                satisfaction=mean_ignore_empty(
                    [section.satisfaction for section in matched_sections]
                ),
                importance=mean_ignore_empty(
                    [section.importance for section in matched_sections]
                ),
                metrics=tuple(metric_results),
            )
        )

    return SurveyStatistics(
        role_name=role_name,
        satisfaction=mean_ignore_empty([stats.satisfaction for stats in stats_list]),
        importance=mean_ignore_empty([stats.importance for stats in stats_list]),
        sections=tuple(section_results),
        matched_row_count=sum(stats.matched_row_count for stats in stats_list),
    )


def normalize_calculation_mode(calculation_mode: str | None) -> str:
    normalized = str(calculation_mode or DEFAULT_CALCULATION_MODE).strip().lower()
    if normalized not in VALID_CALCULATION_MODES:
        raise ValueError(
            f"calculation_mode 仅支持: {', '.join(VALID_CALCULATION_MODES)}"
        )
    return normalized


def clone_section(
    section: SectionDefinition,
    *,
    name: str | None = None,
    metrics: tuple[MetricDefinition, ...] | None = None,
) -> SectionDefinition:
    return SectionDefinition(
        name=section.name if name is None else name,
        metrics=section.metrics if metrics is None else metrics,
    )


def find_section(role_definition: RoleDefinition, *section_names: str) -> SectionDefinition:
    for section_name in section_names:
        for section in role_definition.sections:
            if section.name == section_name:
                return section
    expected = " / ".join(section_names)
    raise ValueError(f"{role_definition.role_name} 缺少汇总口径所需的二级指标: {expected}")


def find_metric(role_definition: RoleDefinition, metric_name: str) -> MetricDefinition:
    for section in role_definition.sections:
        for metric in section.metrics:
            if metric.name == metric_name:
                return metric
    raise ValueError(f"{role_definition.role_name} 缺少汇总口径所需的三级指标: {metric_name}")


def exclude_metrics(section: SectionDefinition, *metric_names: str) -> tuple[MetricDefinition, ...]:
    excluded = set(metric_names)
    return tuple(metric for metric in section.metrics if metric.name not in excluded)


def build_summary_role_definition(role_definition: RoleDefinition) -> RoleDefinition:
    role_name = role_definition.role_name

    if role_name in SUMMARY_EVENT_ROLE_NAMES:
        product_section = clone_section(
            find_section(role_definition, "会展服务", "会场服务"),
            name="产品服务",
        )
        support_section = clone_section(
            find_section(role_definition, "配套服务"),
            metrics=exclude_metrics(find_section(role_definition, "配套服务"), "餐饮服务"),
        )
        smart_section = clone_section(
            find_section(role_definition, "智慧场馆", "智慧服务"),
            name="智慧场馆/服务",
        )
        dining_section = SectionDefinition(
            name="餐饮服务",
            metrics=(find_metric(role_definition, "餐饮服务"),),
        )
        return RoleDefinition(
            role_name=role_name,
            role_column=role_definition.role_column,
            sections=(
                product_section,
                clone_section(find_section(role_definition, "硬件设施")),
                support_section,
                smart_section,
                dining_section,
            ),
            role_match_value=role_definition.role_match_value,
            role_match_values=role_definition.role_match_values,
            row_conditions=role_definition.row_conditions,
        )

    if role_name in SUMMARY_HOTEL_ROLE_NAMES:
        return RoleDefinition(
            role_name=role_name,
            role_column=role_definition.role_column,
            sections=(
                clone_section(find_section(role_definition, "入住服务"), name="产品服务"),
                clone_section(find_section(role_definition, "硬件设施")),
                clone_section(find_section(role_definition, "智慧场馆", "智慧服务"), name="智慧场馆/服务"),
                clone_section(find_section(role_definition, "餐饮服务")),
            ),
            role_match_value=role_definition.role_match_value,
            role_match_values=role_definition.role_match_values,
            row_conditions=role_definition.row_conditions,
        )

    if role_name in SUMMARY_CATERING_ROLE_NAMES:
        return RoleDefinition(
            role_name=role_name,
            role_column=role_definition.role_column,
            sections=(
                clone_section(find_section(role_definition, "硬件设施")),
                clone_section(find_section(role_definition, "智慧场馆", "智慧服务"), name="智慧场馆/服务"),
                clone_section(find_section(role_definition, "餐饮服务")),
            ),
            role_match_value=role_definition.role_match_value,
            role_match_values=role_definition.role_match_values,
            row_conditions=role_definition.row_conditions,
        )

    return role_definition


def get_effective_role_definition(
    role_definition: RoleDefinition,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
) -> RoleDefinition:
    normalized_mode = normalize_calculation_mode(calculation_mode)
    if normalized_mode == DEFAULT_CALCULATION_MODE:
        return role_definition
    return build_summary_role_definition(role_definition)


def required_columns(role_definition: RoleDefinition) -> set[str]:
    columns = {role_definition.role_column}
    for row_condition in role_definition.row_conditions:
        columns.add(row_condition.column)
    for section in role_definition.sections:
        for metric in section.metrics:
            columns.add(metric.satisfaction_column)
            columns.add(metric.importance_column)
            columns.add(metric.satisfaction_gt_zero_column or metric.satisfaction_column)
            columns.add(metric.satisfaction_lt_eleven_column or metric.satisfaction_column)
            columns.add(metric.importance_gt_zero_column or metric.importance_column)
            columns.add(metric.importance_lt_eleven_column or metric.importance_column)
    return columns


def validate_dataframe(df: pd.DataFrame, role_definition: RoleDefinition) -> None:
    max_required_index = max(
        excel_column_to_index(column_name) for column_name in required_columns(role_definition)
    )
    if len(df.columns) <= max_required_index:
        required_max = max(required_columns(role_definition), key=excel_column_to_index)
        raise ValueError(
            f"{role_definition.role_name} 所需的来源 sheet 列数不足，"
            f" 需要至少覆盖到 Excel 列 {required_max}。"
        )


def load_survey_dataframe(
    input_path: Path,
    role_definition: RoleDefinition,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> tuple[pd.DataFrame, str | None]:
    preprocess_notice = preprocess_phase_column_if_needed(input_path, sheet_name)
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    validate_dataframe(df, role_definition)
    return df, preprocess_notice


def compute_metric_average(
    df: pd.DataFrame,
    role_mask: pd.Series,
    column_name: str,
    gt_zero_column_name: str | None = None,
    lt_eleven_column_name: str | None = None,
) -> float | None:
    series = pd.to_numeric(df.iloc[:, excel_column_to_index(column_name)], errors="coerce")
    gt_zero_series = pd.to_numeric(
        df.iloc[:, excel_column_to_index(gt_zero_column_name or column_name)],
        errors="coerce",
    )
    lt_eleven_series = pd.to_numeric(
        df.iloc[:, excel_column_to_index(lt_eleven_column_name or column_name)],
        errors="coerce",
    )
    valid_series = series[role_mask & gt_zero_series.gt(0) & lt_eleven_series.lt(11)]
    if valid_series.empty:
        return None
    return excel_round(decimal_mean_ignore_empty(valid_series.tolist()))


def compute_role_stats(
    df: pd.DataFrame,
    role_definition: RoleDefinition,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
) -> SurveyStatistics:
    effective_role_definition = get_effective_role_definition(role_definition, calculation_mode)
    role_mask = build_role_mask(df, effective_role_definition)
    matched_row_count = int(role_mask.sum())

    section_results: list[SectionResult] = []
    for section in effective_role_definition.sections:
        metric_results: list[MetricResult] = []
        for metric in section.metrics:
            satisfaction = compute_metric_average(
                df,
                role_mask,
                metric.satisfaction_column,
                gt_zero_column_name=metric.satisfaction_gt_zero_column,
                lt_eleven_column_name=metric.satisfaction_lt_eleven_column,
            )
            importance = compute_metric_average(
                df,
                role_mask,
                metric.importance_column,
                gt_zero_column_name=metric.importance_gt_zero_column,
                lt_eleven_column_name=metric.importance_lt_eleven_column,
            )
            metric_results.append(
                MetricResult(
                    name=metric.name,
                    satisfaction=satisfaction,
                    importance=importance,
                )
            )

        section_results.append(
            SectionResult(
                name=section.name,
                satisfaction=mean_ignore_empty([metric.satisfaction for metric in metric_results]),
                importance=mean_ignore_empty([metric.importance for metric in metric_results]),
                metrics=tuple(metric_results),
            )
        )

    overall_satisfaction = mean_ignore_empty([section.satisfaction for section in section_results])
    overall_importance = mean_ignore_empty([section.importance for section in section_results])
    return SurveyStatistics(
        role_name=effective_role_definition.role_name,
        satisfaction=overall_satisfaction,
        importance=overall_importance,
        sections=tuple(section_results),
        matched_row_count=matched_row_count,
    )


def build_result_dataframe(stats: SurveyStatistics) -> pd.DataFrame:
    rows: list[dict[str, float | str | None]] = [
        {
            "指标": stats.role_name,
            "满意度": stats.satisfaction,
            "重要性": stats.importance,
        }
    ]

    for section in stats.sections:
        rows.append(
            {
                "指标": section.name,
                "满意度": section.satisfaction,
                "重要性": section.importance,
            }
        )
        skip_single_metric_detail = (
            len(section.metrics) == 1 and section.metrics[0].name == section.name
        )
        if skip_single_metric_detail:
            continue
        for metric in section.metrics:
            rows.append(
                {
                    "指标": metric.name,
                    "满意度": metric.satisfaction,
                    "重要性": metric.importance,
                }
            )

    return pd.DataFrame(rows, columns=["指标", "满意度", "重要性"])


def render_markdown_table(df: pd.DataFrame) -> str:
    headers = df.columns.tolist()
    rows = [
        [row["指标"], format_value(row["满意度"]), format_value(row["重要性"])]
        for _, row in df.iterrows()
    ]

    widths = []
    for index, header in enumerate(headers):
        width = len(str(header))
        for row in rows:
            width = max(width, len(str(row[index])))
        widths.append(width)

    def build_row(values: list[str]) -> str:
        cells = [f" {str(value).ljust(widths[index])} " for index, value in enumerate(values)]
        return "|" + "|".join(cells) + "|"

    separator = "|" + "|".join("-" * (width + 2) for width in widths) + "|"

    lines = [build_row(headers), separator]
    lines.extend(build_row(row) for row in rows)
    return "\n".join(lines)


def style_worksheet(worksheet, role_definition: RoleDefinition) -> None:
    section_names = {section.name for section in role_definition.sections}

    for cell in worksheet[1]:
        cell.alignment = CENTER_ALIGNMENT
        cell.font = EMPHASIS_FONT

    for row_index in range(2, worksheet.max_row + 1):
        indicator_name = worksheet.cell(row=row_index, column=1).value
        row_fill = None
        if indicator_name == role_definition.role_name:
            row_fill = OVERALL_FILL
        elif indicator_name in section_names:
            row_fill = SECTION_FILL

        for column_index in range(1, 4):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = CENTER_ALIGNMENT
            if row_fill is not None:
                cell.fill = row_fill
                cell.font = EMPHASIS_FONT

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 12
    worksheet.column_dimensions["C"].width = 12


def normalize_output_dir(output_dir: Path) -> Path:
    if output_dir.exists() and output_dir.is_dir():
        return output_dir

    if output_dir.exists() and output_dir.is_file():
        return output_dir.parent / f"{output_dir.stem}_outputs"

    if output_dir.suffix:
        return output_dir.with_suffix("")

    return output_dir


def build_output_path(
    output_dir: Path,
    output_name: str,
    output_format: str = DEFAULT_OUTPUT_FORMAT,
) -> Path:
    return normalize_output_dir(output_dir) / f"{output_name}.{output_format}"


def save_results(
    df: pd.DataFrame,
    output_path: Path,
    role_definition: RoleDefinition,
    sheet_title: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()

    if suffix == ".xlsx":
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_title, index=False)
            style_worksheet(writer.sheets[sheet_title], role_definition)
    elif suffix == ".csv":
        df.to_csv(output_path, index=False, encoding="utf-8-sig")
    elif suffix in {".md", ".markdown"}:
        output_path.write_text(render_markdown_table(df), encoding="utf-8")
    else:
        raise ValueError("仅支持输出为 .xlsx、.csv、.md 或 .markdown 文件。")


def default_config_output_dir(config_path: Path) -> Path:
    return config_path.parent / f"{config_path.stem}_outputs"


def load_source_file_overrides(raw_overrides: object) -> tuple[SourceFileOverride, ...]:
    if raw_overrides is None:
        return ()

    if not isinstance(raw_overrides, dict):
        raise ValueError("source_file_overrides 必须是表结构。")

    supported_source_files = {rule.source_file_name for rule in CUSTOMER_CATEGORY_RULES}
    overrides: list[SourceFileOverride] = []
    for standard_file_name, actual_file_name_value in raw_overrides.items():
        normalized_standard_file_name = str(standard_file_name).strip()
        normalized_actual_file_name = str(actual_file_name_value).strip()
        if not normalized_standard_file_name or not normalized_actual_file_name:
            raise ValueError("source_file_overrides 的键和值都不能为空。")
        if normalized_standard_file_name not in supported_source_files:
            supported = ", ".join(sorted(supported_source_files))
            raise ValueError(
                f"source_file_overrides 包含未知来源文件: {normalized_standard_file_name}；"
                f"支持的来源文件有: {supported}"
            )
        overrides.append(
            SourceFileOverride(
                standard_file_name=normalized_standard_file_name,
                actual_file_name=normalized_actual_file_name,
            )
        )
    return tuple(overrides)


def load_batch_config(config_path: Path, default_sheet_name: str = DEFAULT_SHEET_NAME) -> BatchConfig:
    with config_path.open("rb") as file:
        raw_config = tomllib.load(file)

    raw_jobs = raw_config.get("jobs")
    raw_input_dir = raw_config.get("input_dir")
    if raw_jobs is not None and raw_input_dir is not None:
        raise ValueError("配置文件不能同时包含 input_dir 和 [[jobs]]。")
    if raw_jobs is None and raw_input_dir is None:
        raise ValueError("配置文件必须包含 input_dir 或至少一个 [[jobs]] 定义。")

    output_dir_value = raw_config.get("output_dir")
    if output_dir_value is None:
        output_dir = default_config_output_dir(config_path)
    else:
        output_dir = (config_path.parent / str(output_dir_value)).resolve()

    output_format = str(raw_config.get("output_format", DEFAULT_OUTPUT_FORMAT)).lower()
    if output_format not in VALID_OUTPUT_FORMATS:
        raise ValueError(f"配置文件 output_format 仅支持: {', '.join(VALID_OUTPUT_FORMATS)}")
    calculation_mode = normalize_calculation_mode(
        raw_config.get("calculation_mode", DEFAULT_CALCULATION_MODE)
    )
    sheet_name = str(raw_config.get("sheet_name", default_sheet_name)).strip() or default_sheet_name
    source_file_overrides = load_source_file_overrides(raw_config.get("source_file_overrides"))

    if raw_input_dir is not None:
        if not str(raw_input_dir).strip():
            raise ValueError("input_dir 不能为空。")
        return BatchConfig(
            config_path=config_path.resolve(),
            output_dir=output_dir,
            output_format=output_format,
            calculation_mode=calculation_mode,
            sheet_name=sheet_name,
            input_dir=(config_path.parent / str(raw_input_dir)).resolve(),
            source_file_overrides=source_file_overrides,
        )

    if source_file_overrides:
        raise ValueError("source_file_overrides 仅能与 input_dir 一起使用。")
    if not isinstance(raw_jobs, list) or not raw_jobs:
        raise ValueError("配置文件必须包含至少一个 [[jobs]] 定义。")

    jobs: list[JobConfig] = []
    for index, job_data in enumerate(raw_jobs, start=1):
        if not isinstance(job_data, dict):
            raise ValueError(f"jobs 第 {index} 项必须是表结构。")
        name = str(job_data.get("name", "")).strip()
        path_value = job_data.get("path")
        template_name = str(job_data.get("template", "")).strip()
        if not name or not path_value or not template_name:
            raise ValueError(f"jobs 第 {index} 项必须包含 name/path/template。")
        default_role_name = resolve_role_definition(template_name).role_name
        role_name = str(job_data.get("role_name", default_role_name)).strip() or default_role_name
        output_name = str(job_data.get("output_name", name)).strip() or name
        job_sheet_name = str(job_data.get("sheet", sheet_name)).strip() or sheet_name
        job_path = (config_path.parent / str(path_value)).resolve()
        job_output_format_value = job_data.get("output_format")
        job_output_format = None
        if job_output_format_value is not None:
            job_output_format = str(job_output_format_value).lower()
            if job_output_format not in VALID_OUTPUT_FORMATS:
                raise ValueError(
                    f"jobs 第 {index} 项的 output_format 仅支持: {', '.join(VALID_OUTPUT_FORMATS)}"
                )
        jobs.append(
            JobConfig(
                name=name,
                path=job_path,
                sheet_name=job_sheet_name,
                template_name=template_name,
                role_name=role_name,
                output_name=output_name,
                output_format=job_output_format,
            )
        )

    return BatchConfig(
        config_path=config_path.resolve(),
        output_dir=output_dir,
        output_format=output_format,
        calculation_mode=calculation_mode,
        sheet_name=sheet_name,
        jobs=tuple(jobs),
    )


def select_jobs(
    jobs: tuple[JobConfig, ...],
    job_filters: list[str],
) -> tuple[JobConfig, ...]:
    selected_jobs = jobs
    if job_filters:
        selected_job_names = set(job_filters)
        selected_jobs = tuple(job for job in selected_jobs if job.name in selected_job_names)
    return selected_jobs


def select_missing_customer_type_notices(
    notices: tuple[MissingCustomerTypeNotice, ...],
    job_filters: list[str],
) -> tuple[MissingCustomerTypeNotice, ...]:
    selected_notices = notices
    if job_filters:
        selected_job_names = set(job_filters)
        selected_notices = tuple(
            notice for notice in selected_notices if notice.customer_type_name in selected_job_names
        )
    return selected_notices


def build_source_file_override_lookup(
    source_file_overrides: tuple[SourceFileOverride, ...],
) -> dict[str, str]:
    return {
        override.standard_file_name: override.actual_file_name
        for override in source_file_overrides
    }


def collect_role_values(df: pd.DataFrame, role_column: str) -> set[str]:
    role_index = excel_column_to_index(role_column)
    if role_index >= len(df.columns):
        raise ValueError(
            f"来源数据缺少身份列 {role_column}，当前仅有 {len(df.columns)} 列。"
        )

    return {
        text
        for text in (str(value).strip() for value in df.iloc[:, role_index].dropna().tolist())
        if text
    }


def load_text_column(
    df: pd.DataFrame,
    column_name: str,
    *,
    column_label: str,
) -> pd.Series:
    column_index = excel_column_to_index(column_name)
    if column_index >= len(df.columns):
        raise ValueError(
            f"来源数据缺少{column_label}列 {column_name}，当前仅有 {len(df.columns)} 列。"
        )

    return (
        df.iloc[:, column_index]
        .astype("string")
        .fillna("")
        .str.strip()
    )


def normalize_expected_values(expected_value: str | tuple[str, ...]) -> tuple[str, ...]:
    if isinstance(expected_value, tuple):
        return tuple(str(value).strip() for value in expected_value if str(value).strip())
    text = str(expected_value).strip()
    return (text,) if text else ()


def build_role_mask(df: pd.DataFrame, role_definition: RoleDefinition) -> pd.Series:
    role_series = load_text_column(df, role_definition.role_column, column_label="身份")
    role_match_values = role_definition.role_match_values or normalize_expected_values(
        role_definition.role_match_value or role_definition.role_name
    )
    role_mask = role_series.isin(role_match_values)

    for row_condition in role_definition.row_conditions:
        condition_series = load_text_column(df, row_condition.column, column_label="筛选")
        role_mask = role_mask & condition_series.isin(
            normalize_expected_values(row_condition.expected_value)
        )

    return role_mask


def build_unmapped_customer_category_notices_for_source(
    source_file_name: str,
    df: pd.DataFrame,
) -> tuple[UnmappedCustomerCategoryNotice, ...]:
    rules = [rule for rule in CUSTOMER_CATEGORY_RULES if rule.source_file_name == source_file_name]
    if not rules:
        return ()

    data_columns = {rule.data_column for rule in rules if rule.data_column}
    auxiliary_columns = {rule.auxiliary_column for rule in rules if rule.auxiliary_column}
    if len(data_columns) != 1:
        return ()

    data_column = next(iter(data_columns))
    data_series = load_text_column(df, data_column, column_label="数据标签")

    if auxiliary_columns:
        if len(auxiliary_columns) != 1:
            return ()
        auxiliary_column = next(iter(auxiliary_columns))
        auxiliary_series = load_text_column(df, auxiliary_column, column_label="辅助标签")
        mapped_pairs: set[tuple[str, str]] = set()
        for rule in rules:
            for auxiliary_value in (rule.auxiliary_values or ("",)):
                for data_value in rule.data_values:
                    mapped_pairs.add((auxiliary_value, data_value))

        pair_counts = (
            pd.DataFrame({"auxiliary": auxiliary_series, "data": data_series})
            .query("auxiliary != '' and data != ''")
            .value_counts(sort=False)
        )
        notices = [
            UnmappedCustomerCategoryNotice(
                source_file_name=source_file_name,
                auxiliary_value=auxiliary_value,
                data_value=data_value,
                row_count=int(row_count),
            )
            for (auxiliary_value, data_value), row_count in pair_counts.items()
            if (auxiliary_value, data_value) not in mapped_pairs
        ]
        notices.sort(key=lambda item: (-item.row_count, item.auxiliary_value or "", item.data_value))
        return tuple(notices)

    mapped_values = {
        data_value
        for rule in rules
        for data_value in rule.data_values
    }
    value_counts = data_series[data_series != ""].value_counts(sort=False)
    notices = [
        UnmappedCustomerCategoryNotice(
            source_file_name=source_file_name,
            auxiliary_value=None,
            data_value=data_value,
            row_count=int(row_count),
        )
        for data_value, row_count in value_counts.items()
        if data_value not in mapped_values
    ]
    notices.sort(key=lambda item: (-item.row_count, item.data_value))
    return tuple(notices)


def discover_directory_jobs(config: BatchConfig) -> DirectoryDiscoveryResult:
    if config.input_dir is None:
        raise ValueError("仅目录模式配置支持自动发现 jobs。")

    source_file_override_lookup = build_source_file_override_lookup(config.source_file_overrides)
    jobs: list[JobConfig] = []
    missing_customer_type_notices: list[MissingCustomerTypeNotice] = []
    preprocess_notices: list[PreprocessNoticeRecord] = []
    unmapped_customer_category_notices: list[UnmappedCustomerCategoryNotice] = []
    dataframe_cache: dict[Path, pd.DataFrame] = {}

    for rule in CUSTOMER_CATEGORY_RULES:
        source_reference = source_file_override_lookup.get(
            rule.source_file_name,
            rule.source_file_name,
        )
        input_path = (config.input_dir / source_reference).resolve()
        if not input_path.exists() or not input_path.is_file():
            missing_customer_type_notices.append(
                MissingCustomerTypeNotice(
                    customer_type_name=rule.name,
                    source_reference=source_reference,
                    sheet_name=config.sheet_name,
                    reason=DIRECTORY_NOTICE_REASON_MISSING_SOURCE_FILE,
                )
            )
            continue

        if input_path not in dataframe_cache:
            preprocess_notice = preprocess_phase_column_if_needed(input_path, config.sheet_name)
            if preprocess_notice is not None:
                preprocess_notices.append(
                    PreprocessNoticeRecord(input_path=input_path, notice=preprocess_notice)
                )
            dataframe_cache[input_path] = pd.read_excel(input_path, sheet_name=config.sheet_name)
            unmapped_customer_category_notices.extend(
                build_unmapped_customer_category_notices_for_source(
                    rule.source_file_name,
                    dataframe_cache[input_path],
                )
            )

        if not build_customer_category_rule_mask(dataframe_cache[input_path], rule).any():
            missing_customer_type_notices.append(
                MissingCustomerTypeNotice(
                    customer_type_name=rule.name,
                    source_reference=source_reference,
                    sheet_name=config.sheet_name,
                    reason=DIRECTORY_NOTICE_REASON_MISSING_ROLE_DATA,
                )
            )
            continue

        jobs.append(
            JobConfig(
                name=rule.name,
                path=input_path,
                sheet_name=config.sheet_name,
                template_name=rule.template_name or "",
                role_name=rule.name,
                output_name=rule.name,
                category_rule_name=rule.name,
            )
        )

    return DirectoryDiscoveryResult(
        jobs=tuple(jobs),
        missing_customer_type_notices=tuple(missing_customer_type_notices),
        preprocess_notices=tuple(preprocess_notices),
        unmapped_customer_category_notices=tuple(unmapped_customer_category_notices),
    )


def generate_role_report_bundle(
    input_path: Path,
    role_definition: RoleDefinition,
    output_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    sheet_title: str | None = None,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
    dry_run: bool = False,
    save_empty_report: bool = True,
) -> GeneratedReport:
    effective_role_definition = get_effective_role_definition(role_definition, calculation_mode)
    survey_df, preprocess_notice = load_survey_dataframe(
        input_path,
        effective_role_definition,
        sheet_name=sheet_name,
    )
    stats = compute_role_stats(survey_df, effective_role_definition)
    result_df = build_result_dataframe(stats)
    final_sheet_title = sheet_title or role_definition.role_name
    if not dry_run and (save_empty_report or stats.matched_row_count > 0):
        save_results(result_df, output_path, effective_role_definition, final_sheet_title)
    return GeneratedReport(
        result_df=result_df,
        output_path=output_path,
        stats=stats,
        preprocess_notice=preprocess_notice,
    )


def generate_customer_category_report_bundle(
    input_path: Path,
    category_rule: CustomerCategoryRule,
    output_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
    dry_run: bool = False,
    save_empty_report: bool = True,
) -> GeneratedReport:
    if not category_rule.is_aggregate:
        role_definition = build_role_definition_from_customer_category_rule(category_rule)
        return generate_role_report_bundle(
            input_path=input_path,
            role_definition=role_definition,
            output_path=output_path,
            sheet_name=sheet_name,
            sheet_title=category_rule.name,
            calculation_mode=calculation_mode,
            dry_run=dry_run,
            save_empty_report=save_empty_report,
        )

    component_rules = [
        CUSTOMER_CATEGORY_RULE_BY_NAME[rule_name]
        for rule_name in category_rule.aggregate_rule_names
    ]
    if not component_rules:
        raise ValueError(f"{category_rule.name} 未配置聚合组件规则。")

    first_role_definition = build_role_definition_from_customer_category_rule(component_rules[0])
    preprocess_notice = preprocess_phase_column_if_needed(input_path, sheet_name)
    df = pd.read_excel(input_path, sheet_name=sheet_name)
    validate_dataframe(df, first_role_definition)

    component_stats: list[SurveyStatistics] = []
    for component_rule in component_rules:
        role_definition = build_role_definition_from_customer_category_rule(component_rule)
        validate_dataframe(df, role_definition)
        component_stats.append(
            compute_role_stats(
                df,
                role_definition,
                calculation_mode=calculation_mode,
            )
        )

    stats = merge_survey_statistics(category_rule.name, component_stats)
    result_df = build_result_dataframe(stats)
    if not dry_run and (save_empty_report or stats.matched_row_count > 0):
        save_results(
            result_df,
            output_path,
            RoleDefinition(role_name=category_rule.name, role_column="A", sections=stats.sections),
            category_rule.name,
        )

    return GeneratedReport(
        result_df=result_df,
        output_path=output_path,
        stats=stats,
        preprocess_notice=preprocess_notice,
    )


def generate_role_report(
    input_path: Path,
    role_definition: RoleDefinition,
    output_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    sheet_title: str | None = None,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
    dry_run: bool = False,
) -> tuple[pd.DataFrame, Path]:
    report = generate_role_report_bundle(
        input_path=input_path,
        role_definition=role_definition,
        output_path=output_path,
        sheet_name=sheet_name,
        sheet_title=sheet_title,
        calculation_mode=calculation_mode,
        dry_run=dry_run,
    )
    return report.result_df, report.output_path


def build_missing_group_summary(notices: list[MissingGroupNotice]) -> str | None:
    if not notices:
        return None

    lines = ["以下指定的客户分组在来源数据中未找到任何匹配记录，已输出空白统计结果："]
    for notice in notices:
        lines.append(f"- {notice.job_name} [{notice.input_path.name} / {notice.sheet_name}]")
    return "\n".join(lines)


def print_missing_group_summary(notices: list[MissingGroupNotice]) -> None:
    summary = build_missing_group_summary(notices)
    if summary is not None:
        print(f"\n{summary}")


def build_missing_customer_type_summary(notices: list[MissingCustomerTypeNotice]) -> str | None:
    if not notices:
        return None

    missing_source_file_notices = [
        notice
        for notice in notices
        if notice.reason == DIRECTORY_NOTICE_REASON_MISSING_SOURCE_FILE
    ]
    missing_role_data_notices = [
        notice
        for notice in notices
        if notice.reason == DIRECTORY_NOTICE_REASON_MISSING_ROLE_DATA
    ]

    lines = ["以下客户类型因缺少来源数据被跳过，未生成统计结果："]
    if missing_source_file_notices:
        lines.append("[缺少来源文件]")
        for notice in missing_source_file_notices:
            lines.append(
                f"- {notice.customer_type_name} [{notice.source_reference} / {notice.sheet_name}]"
            )
    if missing_role_data_notices:
        lines.append("[来源文件存在但未找到匹配身份值]")
        for notice in missing_role_data_notices:
            lines.append(
                f"- {notice.customer_type_name} [{notice.source_reference} / {notice.sheet_name}]"
            )
    return "\n".join(lines)


def print_missing_customer_type_summary(notices: list[MissingCustomerTypeNotice]) -> None:
    summary = build_missing_customer_type_summary(notices)
    if summary is not None:
        print(f"\n{summary}")


def build_unmapped_customer_category_summary(
    notices: list[UnmappedCustomerCategoryNotice],
) -> str | None:
    if not notices:
        return None

    lines = ["以下来源数据中存在未纳入 V1.0 客户类别口径的标签组合，本次统计已排除："]
    for notice in notices:
        if notice.auxiliary_value:
            label = f"{notice.auxiliary_value} + {notice.data_value}"
        else:
            label = notice.data_value
        lines.append(f"- {notice.source_file_name}: {label}（{notice.row_count} 行）")
    return "\n".join(lines)


def print_unmapped_customer_category_summary(
    notices: list[UnmappedCustomerCategoryNotice],
) -> None:
    summary = build_unmapped_customer_category_summary(notices)
    if summary is not None:
        print(f"\n{summary}")


def build_progress_prefix(current: int, total: int) -> str:
    return f"[{current}/{total}]"


def print_file_progress_start(current: int, total: int, input_path: Path, job_name: str) -> None:
    print(f"{build_progress_prefix(current, total)} 正在处理文件：{input_path.name}（{job_name}）")


def print_preprocess_notice(current: int, total: int, notice: str | None) -> None:
    if notice is not None:
        print(f"{build_progress_prefix(current, total)} {notice}")


def print_file_progress_result(
    current: int,
    total: int,
    input_path: Path,
    job_name: str,
    output_path: Path,
    dry_run: bool = False,
) -> None:
    status = "已完成校验" if dry_run else "结果已保存"
    print(
        f"{build_progress_prefix(current, total)} "
        f"{status}：{input_path.name}（{job_name}） -> {output_path}"
    )


def run_single_mode(args: argparse.Namespace) -> None:
    if not all([args.input, args.template, args.role_name, args.output]):
        raise ValueError("--input、--template、--role-name、--output 必须同时提供。")

    role_definition = resolve_role_definition(args.template, args.role_name)
    print_file_progress_start(1, 1, args.input, args.role_name)
    report = generate_role_report_bundle(
        input_path=args.input,
        role_definition=role_definition,
        output_path=args.output,
        sheet_name=args.sheet_name,
        sheet_title=args.output.stem,
        calculation_mode=args.calculation_mode,
        dry_run=args.dry_run,
    )
    print_preprocess_notice(1, 1, report.preprocess_notice)
    print_file_progress_result(1, 1, args.input, args.role_name, report.output_path, dry_run=args.dry_run)
    if report.stats.matched_row_count == 0:
        print_missing_group_summary(
            [MissingGroupNotice(args.role_name, args.input, args.sheet_name)]
        )


def run_legacy_batch_mode(args: argparse.Namespace) -> None:
    if not all([args.organizer_input, args.exhibitor_input, args.visitor_input, args.output_dir]):
        raise ValueError(
            "--organizer-input、--exhibitor-input、--visitor-input、--output-dir 必须同时提供。"
        )

    jobs = (
        (ORGANIZER_TEMPLATE, args.organizer_input),
        (EXHIBITOR_TEMPLATE, args.exhibitor_input),
        (VISITOR_TEMPLATE, args.visitor_input),
    )
    output_dir = normalize_output_dir(args.output_dir)
    output_format = args.output_format or DEFAULT_OUTPUT_FORMAT
    missing_group_notices: list[MissingGroupNotice] = []
    total_jobs = len(jobs)

    for index, (role_definition, input_path) in enumerate(jobs, start=1):
        output_path = build_output_path(output_dir, role_definition.role_name, output_format)
        print_file_progress_start(index, total_jobs, input_path, role_definition.role_name)
        report = generate_role_report_bundle(
            input_path=input_path,
            role_definition=role_definition,
            output_path=output_path,
            sheet_name=args.sheet_name,
            sheet_title=role_definition.role_name,
            calculation_mode=args.calculation_mode,
            dry_run=args.dry_run,
        )
        print_preprocess_notice(index, total_jobs, report.preprocess_notice)
        print_file_progress_result(
            index,
            total_jobs,
            input_path,
            role_definition.role_name,
            report.output_path,
            dry_run=args.dry_run,
        )
        if report.stats.matched_row_count == 0:
            missing_group_notices.append(
                MissingGroupNotice(role_definition.role_name, input_path, args.sheet_name)
            )

    print_missing_group_summary(missing_group_notices)


def run_batch_config(
    config: BatchConfig,
    *,
    output_dir_override: Path | None = None,
    output_format_override: str | None = None,
    calculation_mode_override: str | None = None,
    selected_job_names: list[str] | tuple[str, ...] = (),
    dry_run: bool = False,
) -> None:
    output_dir = normalize_output_dir(output_dir_override or config.output_dir)
    global_output_format = output_format_override or config.output_format
    calculation_mode = normalize_calculation_mode(
        calculation_mode_override or config.calculation_mode
    )
    job_filters = list(selected_job_names)
    selected_jobs: tuple[JobConfig, ...]
    missing_group_notices: list[MissingGroupNotice] = []
    missing_customer_type_notices: list[MissingCustomerTypeNotice] = []
    unmapped_customer_category_notices: list[UnmappedCustomerCategoryNotice] = []
    preprocess_notice_lookup: dict[Path, str] = {}

    if config.input_dir is None:
        selected_jobs = select_jobs(config.jobs, job_filters)
        if not selected_jobs:
            raise ValueError("筛选后没有可运行的 jobs。")
    else:
        discovery_result = discover_directory_jobs(config)
        selected_jobs = select_jobs(discovery_result.jobs, job_filters)
        missing_customer_type_notices = list(
            select_missing_customer_type_notices(
                discovery_result.missing_customer_type_notices,
                job_filters,
            )
        )
        preprocess_notice_lookup = {
            record.input_path: record.notice for record in discovery_result.preprocess_notices
        }
        unmapped_customer_category_notices = list(
            discovery_result.unmapped_customer_category_notices
        )
        if not selected_jobs and not missing_customer_type_notices:
            raise ValueError("筛选后没有可运行的 jobs。")

    total_jobs = len(selected_jobs)

    for index, job in enumerate(selected_jobs, start=1):
        output_format = job.output_format or global_output_format
        output_path = build_output_path(output_dir, job.output_name, output_format)
        print_file_progress_start(index, total_jobs, job.path, job.name)
        if job.category_rule_name is None:
            role_definition = resolve_role_definition(job.template_name, job.role_name)
            report = generate_role_report_bundle(
                input_path=job.path,
                role_definition=role_definition,
                output_path=output_path,
                sheet_name=job.sheet_name,
                sheet_title=job.name,
                calculation_mode=calculation_mode,
                dry_run=dry_run,
                save_empty_report=config.input_dir is None,
            )
        else:
            category_rule = CUSTOMER_CATEGORY_RULE_BY_NAME[job.category_rule_name]
            report = generate_customer_category_report_bundle(
                input_path=job.path,
                category_rule=category_rule,
                output_path=output_path,
                sheet_name=job.sheet_name,
                calculation_mode=calculation_mode,
                dry_run=dry_run,
                save_empty_report=config.input_dir is None,
            )
        preprocess_notice = report.preprocess_notice or preprocess_notice_lookup.pop(job.path, None)
        print_preprocess_notice(index, total_jobs, preprocess_notice)
        if config.input_dir is not None and report.stats.matched_row_count == 0:
            missing_customer_type_notices.append(
                MissingCustomerTypeNotice(
                    customer_type_name=job.name,
                    source_reference=job.path.name,
                    sheet_name=job.sheet_name,
                    reason=DIRECTORY_NOTICE_REASON_MISSING_ROLE_DATA,
                )
            )
            continue
        print_file_progress_result(
            index,
            total_jobs,
            job.path,
            job.name,
            report.output_path,
            dry_run=dry_run,
        )
        if report.stats.matched_row_count == 0:
            missing_group_notices.append(
                MissingGroupNotice(job.name, job.path, job.sheet_name)
            )

    if config.input_dir is None:
        print_missing_group_summary(missing_group_notices)
    else:
        print_missing_customer_type_summary(missing_customer_type_notices)
        print_unmapped_customer_category_summary(unmapped_customer_category_notices)


def run_directory_batch(
    *,
    input_dir: Path,
    output_dir: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    output_format: str = DEFAULT_OUTPUT_FORMAT,
    calculation_mode: str = DEFAULT_CALCULATION_MODE,
    job_filters: list[str] | tuple[str, ...] = (),
    dry_run: bool = False,
) -> None:
    config = BatchConfig(
        config_path=Path("<programmatic-directory-mode>"),
        output_dir=output_dir,
        output_format=output_format,
        calculation_mode=normalize_calculation_mode(calculation_mode),
        sheet_name=sheet_name,
        input_dir=input_dir,
    )
    try:
        run_batch_config(
            config,
            selected_job_names=job_filters,
            dry_run=dry_run,
        )
    except ValueError as exc:
        if str(exc) != "筛选后没有可运行的 jobs。":
            raise


def run_config_mode(args: argparse.Namespace) -> None:
    config = load_batch_config(args.config, default_sheet_name=args.sheet_name)
    run_batch_config(
        config,
        output_dir_override=args.output_dir,
        output_format_override=args.output_format,
        calculation_mode_override=args.calculation_mode,
        selected_job_names=args.job,
        dry_run=args.dry_run,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="按会展问卷模板批量或单次计算统计结果，并导出 Excel/CSV/Markdown 文件。"
    )
    parser.add_argument("--config", type=Path, help="批量模式：TOML 配置文件路径")
    parser.add_argument("--job", action="append", default=[], help="批量模式：只运行某个 job 名称，可重复传入")
    parser.add_argument("--dry-run", action="store_true", help="只校验并显示处理进度，不实际写文件")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help=f"默认 sheet 名，默认 {DEFAULT_SHEET_NAME}")
    parser.add_argument(
        "--output-format",
        choices=VALID_OUTPUT_FORMATS,
        help=f"覆盖输出格式，支持 {', '.join(VALID_OUTPUT_FORMATS)}",
    )
    parser.add_argument(
        "--calculation-mode",
        choices=VALID_CALCULATION_MODES,
        help=f"计算口径，支持 {', '.join(VALID_CALCULATION_MODES)}；默认 template",
    )
    parser.add_argument("--output-dir", type=Path, help="批量模式输出目录，或覆盖配置里的 output_dir")

    parser.add_argument("--input", type=Path, help="单任务模式：来源 Excel 文件")
    parser.add_argument(
        "--template",
        choices=sorted(TEMPLATE_DEFINITIONS),
        help="单任务模式：模板类型",
    )
    parser.add_argument("--role-name", help="单任务模式：按该身份分组")
    parser.add_argument("--output", type=Path, help="单任务模式：输出文件路径")

    parser.add_argument("--organizer-input", type=Path, help="兼容模式：展览主承办来源 Excel")
    parser.add_argument("--exhibitor-input", type=Path, help="兼容模式：参展商来源 Excel")
    parser.add_argument("--visitor-input", type=Path, help="兼容模式：专业观众来源 Excel")

    return parser.parse_args()


def main() -> None:
    args = parse_args()

    if args.config:
        run_config_mode(args)
        return

    if any([args.input, args.template, args.role_name, args.output]):
        run_single_mode(args)
        return

    if any([args.organizer_input, args.exhibitor_input, args.visitor_input]):
        run_legacy_batch_mode(args)
        return

    raise SystemExit(
        "请使用 --config 批量模式，或提供单任务参数 --input/--template/--role-name/--output。"
    )


if __name__ == "__main__":
    main()

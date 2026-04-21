from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_FILES = {
    "1-2月": ROOT / "汇总结果" / "1-2月" / "1-2月客户类型满意度汇总表.xlsx",
    "3月": ROOT / "汇总结果" / "3月" / "3月客户类型满意度汇总表.xlsx",
}
SAMPLE_FILE = ROOT / "汇总结果" / "Q1" / "Q1客户类型样本统计表.xlsx"
OUTPUT_HTML = ROOT / "output" / "满意度趋势演示仪表板.html"
OUTPUT_JSON = ROOT / "output" / "满意度趋势演示数据.json"
GROUPS = ["一、会展客户", "二、餐饮客户", "三、G20峰会体验馆", "五、酒店客户"]
DIMENSION_LABELS = {
    "score": "总分",
    "product": "产品服务",
    "hardware": "硬件设施",
    "support": "配套服务",
    "smart": "智慧场馆/服务",
    "dining": "餐饮服务",
}
DIMENSION_KEYS = ["score", "product", "hardware", "support", "smart", "dining"]


def norm(value: object) -> float | None:
    if value in (None, "", "-", "--", "—"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def normalize_name(name: object) -> str:
    text = str(name or "").strip()
    mapping = {
        "会议活动主（承）办": "会议活动主（承）办",
        "参会客户": "参会客户",
        "展览活动主（承）办": "展览活动主（承）办",
        "酒店住宿团队": "酒店住宿团队",
        "酒店参会客户": "酒店参会客户",
        "酒店餐饮客户": "酒店餐饮客户",
        "酒店散客": "酒店散客",
    }
    return mapping.get(text, text)


def read_summary(path: Path) -> tuple[pd.DataFrame, dict[str, float | None]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    records: list[dict[str, Any]] = []
    current_group = None
    total_row: dict[str, float | None] = {}
    for row in rows[2:]:
        first, second, score, product, hardware, support, smart, dining = row[:8]
        if first == "总分":
            total_row = {
                "score": norm(score),
                "product": norm(product),
                "hardware": norm(hardware),
                "support": norm(support),
                "smart": norm(smart),
                "dining": norm(dining),
            }
            break
        if first:
            current_group = str(first).strip()
        if not second:
            continue
        records.append(
            {
                "group": current_group,
                "type": normalize_name(second),
                "score": norm(score),
                "product": norm(product),
                "hardware": norm(hardware),
                "support": norm(support),
                "smart": norm(smart),
                "dining": norm(dining),
            }
        )
    return pd.DataFrame(records), total_row


def read_samples(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    idx_group = header.index("客户大类")
    idx_type = header.index("样本类型")
    idx_12 = header.index("1-2月")
    idx_3 = header.index("3月")
    current_group = None
    records: list[dict[str, Any]] = []
    skip_names = {"小计", "四、专项调研", "六、酒店暗访（次）", "会展流失主办客户"}
    for row in rows[2:]:
        group = row[idx_group]
        sample_type = row[idx_type]
        if group == "合计":
            break
        if group and group != "小计":
            current_group = str(group).strip()
        if not sample_type or sample_type in skip_names:
            continue
        records.append(
            {
                "group": current_group,
                "type": normalize_name(sample_type),
                "1-2月_samples": int(row[idx_12] or 0),
                "3月_samples": int(row[idx_3] or 0),
            }
        )
    return pd.DataFrame(records)


def weighted_mean(df: pd.DataFrame, value_col: str, weight_col: str) -> float | None:
    temp = pd.DataFrame(
        {
            "v": pd.to_numeric(df[value_col], errors="coerce"),
            "w": pd.to_numeric(df[weight_col], errors="coerce"),
        }
    ).dropna()
    temp = temp[temp["w"] > 0]
    if temp.empty:
        return None
    return float((temp["v"] * temp["w"]).sum() / temp["w"].sum())


def round_or_none(value: float | None, digits: int = 3) -> float | None:
    if value is None or pd.isna(value):
        return None
    return round(float(value), digits)


def build_data() -> dict[str, Any]:
    summary_frames: list[pd.DataFrame] = []
    official_rows: dict[str, dict[str, float | None]] = {}
    for period, path in SUMMARY_FILES.items():
        df, total_row = read_summary(path)
        df["period"] = period
        summary_frames.append(df)
        official_rows[period] = total_row

    summary_df = pd.concat(summary_frames, ignore_index=True)
    sample_df = read_samples(SAMPLE_FILE)
    merged = summary_df.merge(sample_df, on=["group", "type"], how="left")
    merged["samples"] = merged.apply(lambda row: row[f"{row['period']}_samples"], axis=1)

    cards = []
    group_dimension_trend = []
    detail_by_group: dict[str, list[dict[str, Any]]] = {}
    for group in GROUPS:
        g12 = merged[(merged["group"].eq(group)) & (merged["period"].eq("1-2月"))].copy()
        g3 = merged[(merged["group"].eq(group)) & (merged["period"].eq("3月"))].copy()
        score12 = weighted_mean(g12, "score", "samples")
        score3 = weighted_mean(g3, "score", "samples")
        sample12 = int(pd.to_numeric(g12.loc[pd.to_numeric(g12["score"], errors="coerce").notna(), "samples"], errors="coerce").fillna(0).sum())
        sample3 = int(pd.to_numeric(g3.loc[pd.to_numeric(g3["score"], errors="coerce").notna(), "samples"], errors="coerce").fillna(0).sum())
        cards.append(
            {
                "group": group,
                "score12": round_or_none(score12),
                "score3": round_or_none(score3),
                "delta": round_or_none(score3 - score12 if score12 is not None and score3 is not None else None),
                "sample12": sample12,
                "sample3": sample3,
            }
        )

        dims = []
        for key in DIMENSION_KEYS[1:]:
            val12 = weighted_mean(g12, key, "samples")
            val3 = weighted_mean(g3, key, "samples")
            dims.append(
                {
                    "dimension": DIMENSION_LABELS[key],
                    "value12": round_or_none(val12),
                    "value3": round_or_none(val3),
                    "delta": round_or_none(val3 - val12 if val12 is not None and val3 is not None else None),
                }
            )
            group_dimension_trend.append(
                {
                    "group": group,
                    "dimension": DIMENSION_LABELS[key],
                    "value12": round_or_none(val12),
                    "value3": round_or_none(val3),
                    "delta": round_or_none(val3 - val12 if val12 is not None and val3 is not None else None),
                }
            )

        detail_rows = []
        type_names = sorted(set(g12["type"].dropna().tolist()) | set(g3["type"].dropna().tolist()))
        for type_name in type_names:
            row12 = g12[g12["type"].eq(type_name)]
            row3 = g3[g3["type"].eq(type_name)]
            base: dict[str, Any] = {
                "type": type_name,
                "sample12": int(row12["samples"].iloc[0]) if not row12.empty and pd.notna(row12["samples"].iloc[0]) else 0,
                "sample3": int(row3["samples"].iloc[0]) if not row3.empty and pd.notna(row3["samples"].iloc[0]) else 0,
                "dimensions": {},
            }
            for key in DIMENSION_KEYS:
                v12 = None if row12.empty else row12[key].iloc[0]
                v3 = None if row3.empty else row3[key].iloc[0]
                base["dimensions"][key] = {
                    "label": DIMENSION_LABELS[key],
                    "value12": round_or_none(v12),
                    "value3": round_or_none(v3),
                    "delta": round_or_none(v3 - v12 if pd.notna(v12) and pd.notna(v3) else None),
                }
            detail_rows.append(base)
        detail_by_group[group] = detail_rows

    official_dimension_trend = []
    for key in DIMENSION_KEYS[1:]:
        v12 = official_rows["1-2月"].get(key)
        v3 = official_rows["3月"].get(key)
        official_dimension_trend.append(
            {
                "dimension": DIMENSION_LABELS[key],
                "value12": round_or_none(v12),
                "value3": round_or_none(v3),
                "delta": round_or_none(v3 - v12 if v12 is not None and v3 is not None else None),
            }
        )

    return {
        "cards": cards,
        "officialDimensionTrend": official_dimension_trend,
        "groupDimensionTrend": group_dimension_trend,
        "detailByGroup": detail_by_group,
        "meta": {
            "title": "满意度趋势演示仪表板",
            "periods": ["1-2月", "3月"],
            "groups": GROUPS,
        },
    }


HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>满意度趋势演示仪表板</title>
  <script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
  <style>
    :root {
      --bg: #081225;
      --panel: rgba(14, 29, 56, 0.88);
      --panel-2: rgba(20, 40, 75, 0.92);
      --text: #e9f1ff;
      --muted: #99abd1;
      --accent: #5aa9ff;
      --accent-2: #55d8c1;
      --rise: #31c48d;
      --fall: #ff6b7d;
      --warning: #ffb84d;
      --grid: rgba(255,255,255,0.08);
      --shadow: 0 20px 40px rgba(0, 0, 0, 0.35);
      --radius: 20px;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      background: radial-gradient(circle at top right, #173867, #081225 50%, #050b16 100%);
      color: var(--text);
      min-height: 100vh;
    }
    .page {
      width: min(1500px, calc(100vw - 40px));
      margin: 0 auto;
      padding: 28px 0 32px;
    }
    .hero {
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 18px;
    }
    .hero h1 {
      margin: 0;
      font-size: 34px;
      font-weight: 800;
      letter-spacing: 0.5px;
    }
    .hero p {
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 15px;
      line-height: 1.6;
      max-width: 920px;
    }
    .stamp {
      padding: 12px 18px;
      background: linear-gradient(135deg, rgba(90,169,255,0.18), rgba(85,216,193,0.14));
      border: 1px solid rgba(255,255,255,0.12);
      border-radius: 999px;
      color: #d7e8ff;
      font-size: 13px;
      white-space: nowrap;
      box-shadow: var(--shadow);
    }
    .cards {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 14px;
      margin-bottom: 18px;
    }
    .card {
      background: linear-gradient(160deg, rgba(20, 40, 75, 0.96), rgba(10, 22, 44, 0.96));
      border: 1px solid rgba(255,255,255,0.1);
      border-radius: var(--radius);
      padding: 18px 18px 16px;
      box-shadow: var(--shadow);
      position: relative;
      overflow: hidden;
      transform: translateY(16px);
      opacity: 0;
      animation: fadeUp 0.8s ease forwards;
    }
    .card::after {
      content: "";
      position: absolute;
      inset: auto -30px -60px auto;
      width: 140px;
      height: 140px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(90,169,255,0.18), transparent 70%);
      pointer-events: none;
    }
    .card:nth-child(2) { animation-delay: 0.08s; }
    .card:nth-child(3) { animation-delay: 0.16s; }
    .card:nth-child(4) { animation-delay: 0.24s; }
    .card .title {
      font-size: 14px;
      color: var(--muted);
      margin-bottom: 10px;
    }
    .card .value-row {
      display: flex;
      justify-content: space-between;
      align-items: baseline;
      gap: 10px;
      margin-bottom: 8px;
    }
    .card .value {
      font-size: 38px;
      font-weight: 800;
      letter-spacing: -1px;
    }
    .delta {
      padding: 4px 10px;
      border-radius: 999px;
      font-size: 13px;
      font-weight: 700;
    }
    .delta.fall { background: rgba(255,107,125,0.16); color: #ffb3bd; }
    .delta.rise { background: rgba(49,196,141,0.18); color: #8ef1cb; }
    .delta.flat { background: rgba(255,255,255,0.12); color: #d6e3ff; }
    .subline {
      display: flex;
      justify-content: space-between;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.7;
    }
    .grid-2 {
      display: grid;
      grid-template-columns: 1.2fr 1fr;
      gap: 16px;
      margin-bottom: 18px;
    }
    .panel {
      background: linear-gradient(180deg, rgba(16, 32, 60, 0.96), rgba(9, 20, 40, 0.96));
      border: 1px solid rgba(255,255,255,0.08);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      overflow: hidden;
    }
    .panel-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 16px 18px 0;
    }
    .panel h2 {
      margin: 0;
      font-size: 18px;
    }
    .panel .desc {
      color: var(--muted);
      font-size: 13px;
      margin: 6px 0 0;
    }
    .chart {
      height: 370px;
      width: 100%;
    }
    .bottom-panel .chart {
      height: 490px;
    }
    .toolbar {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      padding: 14px 18px 0;
    }
    .btn {
      border: 1px solid rgba(255,255,255,0.12);
      background: rgba(255,255,255,0.05);
      color: var(--muted);
      padding: 8px 14px;
      border-radius: 999px;
      cursor: pointer;
      font-size: 13px;
      transition: all .2s ease;
    }
    .btn:hover { color: var(--text); border-color: rgba(255,255,255,0.24); }
    .btn.active {
      color: white;
      background: linear-gradient(135deg, rgba(90,169,255,0.34), rgba(85,216,193,0.28));
      border-color: rgba(125, 188, 255, 0.55);
      box-shadow: 0 10px 20px rgba(13, 38, 76, 0.35);
    }
    .insight-row {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
      padding: 0 18px 18px;
    }
    .insight {
      background: rgba(255,255,255,0.04);
      border: 1px solid rgba(255,255,255,0.08);
      border-radius: 14px;
      padding: 12px 14px;
      min-height: 88px;
    }
    .insight .kicker {
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 6px;
    }
    .insight .headline {
      font-size: 15px;
      line-height: 1.45;
      font-weight: 700;
    }
    .insight .detail {
      margin-top: 6px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.55;
    }
    .legend-note {
      padding: 0 18px 12px;
      color: var(--muted);
      font-size: 12px;
    }
    @keyframes fadeUp {
      from { opacity: 0; transform: translateY(16px); }
      to { opacity: 1; transform: translateY(0); }
    }
    @media (max-width: 1200px) {
      .cards, .grid-2, .insight-row { grid-template-columns: 1fr 1fr; }
    }
    @media (max-width: 900px) {
      .cards, .grid-2, .insight-row { grid-template-columns: 1fr; }
      .hero { flex-direction: column; align-items: flex-start; }
      .page { width: min(100vw - 24px, 1500px); }
    }
  </style>
</head>
<body>
  <div class="page">
    <div class="hero">
      <div>
        <h1>满意度趋势演示仪表板</h1>
        <p>基于 1-2月 与 3月 的满意度汇总表、样本统计表整理。顶部看四大类总体变化，中部看总体维度趋势和大类差异，底部按客户大类切换客户分组动画图表，突出 1-2月 → 3月 的变化。</p>
      </div>
      <div class="stamp">单页 HTML 动画演示版</div>
    </div>

    <div id="cards" class="cards"></div>

    <div class="grid-2">
      <section class="panel">
        <div class="panel-header">
          <div>
            <h2>总体五维度趋势</h2>
            <div class="desc">官方汇总表底部总分口径，比较 1-2月 与 3月 五个维度的整体变化。</div>
          </div>
        </div>
        <div id="officialChart" class="chart"></div>
      </section>

      <section class="panel">
        <div class="panel-header">
          <div>
            <h2>四大类维度变化热力图</h2>
            <div class="desc">样本加权口径。红色越深表示下降越明显，绿色表示改善。</div>
          </div>
        </div>
        <div id="heatmapChart" class="chart"></div>
      </section>
    </div>

    <section class="panel bottom-panel">
      <div class="panel-header">
        <div>
          <h2>客户分组趋势演示</h2>
          <div class="desc">点击客户大类和维度按钮，查看客户分组在 1-2月 与 3月 的动态变化。图表会自动动画过渡。</div>
        </div>
      </div>
      <div id="groupButtons" class="toolbar"></div>
      <div id="dimensionButtons" class="toolbar" style="padding-top:10px;"></div>
      <div class="legend-note">图中标签格式为：客户分组（样本 1-2月→3月）。绿色表示提升，红色表示下降，灰色表示无可比值。</div>
      <div id="detailChart" class="chart"></div>
      <div id="insights" class="insight-row"></div>
    </section>
  </div>

  <script>
    const dashboardData = __DATA__;
    const colors = {
      blue: '#5aa9ff',
      cyan: '#55d8c1',
      rise: '#31c48d',
      fall: '#ff6b7d',
      text: '#e9f1ff',
      muted: '#99abd1',
      grid: 'rgba(255,255,255,0.08)',
    };

    const cardsEl = document.getElementById('cards');
    const groupButtonsEl = document.getElementById('groupButtons');
    const dimensionButtonsEl = document.getElementById('dimensionButtons');
    const insightsEl = document.getElementById('insights');
    const officialChart = echarts.init(document.getElementById('officialChart'));
    const heatmapChart = echarts.init(document.getElementById('heatmapChart'));
    const detailChart = echarts.init(document.getElementById('detailChart'));

    let currentGroup = dashboardData.meta.groups[0];
    let currentDimension = 'score';

    function fmt(value, digits = 3) {
      if (value === null || value === undefined || Number.isNaN(value)) return '-';
      return Number(value).toFixed(digits);
    }

    function deltaText(delta) {
      if (delta === null || delta === undefined || Number.isNaN(delta)) return '无可比值';
      const sign = delta > 0 ? '+' : '';
      return `${sign}${fmt(delta)}`;
    }

    function deltaClass(delta) {
      if (delta === null || delta === undefined || Number.isNaN(delta)) return 'flat';
      if (delta > 0.001) return 'rise';
      if (delta < -0.001) return 'fall';
      return 'flat';
    }

    function renderCards() {
      cardsEl.innerHTML = dashboardData.cards.map(card => `
        <div class="card">
          <div class="title">${card.group}</div>
          <div class="value-row">
            <div class="value">${fmt(card.score3)}</div>
            <div class="delta ${deltaClass(card.delta)}">${deltaText(card.delta)}</div>
          </div>
          <div class="subline"><span>1-2月加权均分</span><strong>${fmt(card.score12)}</strong></div>
          <div class="subline"><span>3月加权均分</span><strong>${fmt(card.score3)}</strong></div>
          <div class="subline"><span>样本变化</span><strong>${card.sample12} → ${card.sample3}</strong></div>
        </div>
      `).join('');
    }

    function renderOfficialChart() {
      const dims = dashboardData.officialDimensionTrend.map(d => d.dimension);
      const values12 = dashboardData.officialDimensionTrend.map(d => d.value12);
      const values3 = dashboardData.officialDimensionTrend.map(d => d.value3);
      officialChart.setOption({
        backgroundColor: 'transparent',
        animationDuration: 900,
        tooltip: { trigger: 'axis' },
        legend: { top: 6, textStyle: { color: colors.muted } },
        grid: { left: 40, right: 24, top: 58, bottom: 32 },
        xAxis: {
          type: 'category',
          data: dims,
          axisLabel: { color: colors.muted, interval: 0 },
          axisLine: { lineStyle: { color: colors.grid } },
          axisTick: { show: false },
        },
        yAxis: {
          type: 'value',
          min: 8,
          max: 10.1,
          axisLabel: { color: colors.muted },
          splitLine: { lineStyle: { color: colors.grid } },
        },
        series: [
          {
            name: '1-2月',
            type: 'bar',
            data: values12,
            barWidth: 18,
            itemStyle: { color: colors.blue, borderRadius: [8, 8, 0, 0] },
            emphasis: { focus: 'series' },
            animationDelay: idx => idx * 80,
          },
          {
            name: '3月',
            type: 'bar',
            data: values3,
            barWidth: 18,
            itemStyle: { color: colors.cyan, borderRadius: [8, 8, 0, 0] },
            emphasis: { focus: 'series' },
            animationDelay: idx => 250 + idx * 80,
          },
        ],
      });
    }

    function renderHeatmap() {
      const groups = dashboardData.meta.groups;
      const dimensions = [...new Set(dashboardData.groupDimensionTrend.map(d => d.dimension))];
      const heatmapData = [];
      dashboardData.groupDimensionTrend.forEach(item => {
        const x = dimensions.indexOf(item.dimension);
        const y = groups.indexOf(item.group);
        heatmapData.push([x, y, item.delta]);
      });
      heatmapChart.setOption({
        backgroundColor: 'transparent',
        animationDuration: 900,
        tooltip: {
          formatter: params => {
            const value = params.value[2];
            return `${groups[params.value[1]]}<br/>${dimensions[params.value[0]]}<br/>变化：${deltaText(value)}`;
          }
        },
        grid: { top: 52, left: 88, right: 18, bottom: 30 },
        xAxis: {
          type: 'category',
          data: dimensions,
          axisLabel: { color: colors.muted, interval: 0 },
          axisLine: { lineStyle: { color: colors.grid } },
        },
        yAxis: {
          type: 'category',
          data: groups,
          axisLabel: { color: colors.muted },
          axisLine: { lineStyle: { color: colors.grid } },
        },
        visualMap: {
          min: -1.0,
          max: 0.4,
          calculable: false,
          orient: 'horizontal',
          left: 'center',
          top: 8,
          textStyle: { color: colors.muted },
          inRange: {
            color: ['#ff6b7d', '#f7b2b8', '#2e3e5f', '#9fe6cf', '#31c48d']
          }
        },
        series: [{
          type: 'heatmap',
          data: heatmapData,
          label: {
            show: true,
            color: '#ffffff',
            formatter: params => params.value[2] == null ? '-' : Number(params.value[2]).toFixed(3)
          },
          itemStyle: {
            borderColor: 'rgba(255,255,255,0.08)',
            borderWidth: 1,
            borderRadius: 10,
          }
        }],
      });
    }

    function renderGroupButtons() {
      groupButtonsEl.innerHTML = dashboardData.meta.groups.map(group => `
        <button class="btn ${group === currentGroup ? 'active' : ''}" data-group="${group}">${group}</button>
      `).join('');
      [...groupButtonsEl.querySelectorAll('.btn')].forEach(btn => {
        btn.onclick = () => {
          currentGroup = btn.dataset.group;
          renderGroupButtons();
          renderDetailChart();
        };
      });
    }

    function renderDimensionButtons() {
      const dims = Object.entries({
        score: '总分',
        product: '产品服务',
        hardware: '硬件设施',
        support: '配套服务',
        smart: '智慧场馆/服务',
        dining: '餐饮服务',
      });
      dimensionButtonsEl.innerHTML = dims.map(([key, label]) => `
        <button class="btn ${key === currentDimension ? 'active' : ''}" data-dimension="${key}">${label}</button>
      `).join('');
      [...dimensionButtonsEl.querySelectorAll('.btn')].forEach(btn => {
        btn.onclick = () => {
          currentDimension = btn.dataset.dimension;
          renderDimensionButtons();
          renderDetailChart();
        };
      });
    }

    function buildInsights(rows) {
      const comparable = rows.filter(row => row.dimensions[currentDimension] && row.dimensions[currentDimension].delta !== null);
      const declines = [...comparable].filter(r => r.dimensions[currentDimension].delta < 0).sort((a, b) => a.dimensions[currentDimension].delta - b.dimensions[currentDimension].delta);
      const rises = [...comparable].filter(r => r.dimensions[currentDimension].delta > 0).sort((a, b) => b.dimensions[currentDimension].delta - a.dimensions[currentDimension].delta);
      const sampleShift = [...rows].sort((a, b) => (b.sample3 - b.sample12) - (a.sample3 - a.sample12));

      const cards = [
        {
          kicker: '最大下降',
          headline: declines[0] ? `${declines[0].type} ${deltaText(declines[0].dimensions[currentDimension].delta)}` : '暂无可比下降项',
          detail: declines[0] ? `${declines[0].dimensions[currentDimension].label}：${fmt(declines[0].dimensions[currentDimension].value12)} → ${fmt(declines[0].dimensions[currentDimension].value3)}` : '当前维度没有 1-2月 与 3月 的可比数据。'
        },
        {
          kicker: '最大提升',
          headline: rises[0] ? `${rises[0].type} ${deltaText(rises[0].dimensions[currentDimension].delta)}` : '暂无可比提升项',
          detail: rises[0] ? `${rises[0].dimensions[currentDimension].label}：${fmt(rises[0].dimensions[currentDimension].value12)} → ${fmt(rises[0].dimensions[currentDimension].value3)}` : '当前维度没有显著提升项。'
        },
        {
          kicker: '样本变化最大',
          headline: sampleShift[0] ? `${sampleShift[0].type} ${sampleShift[0].sample12} → ${sampleShift[0].sample3}` : '暂无样本信息',
          detail: sampleShift[0] ? `说明该客户分组在 3月 的影响权重更大，演示时建议优先关注。` : '无样本信息。'
        }
      ];

      insightsEl.innerHTML = cards.map(card => `
        <div class="insight">
          <div class="kicker">${card.kicker}</div>
          <div class="headline">${card.headline}</div>
          <div class="detail">${card.detail}</div>
        </div>
      `).join('');
    }

    function renderDetailChart() {
      const rows = dashboardData.detailByGroup[currentGroup] || [];
      const normalized = rows.map(row => {
        const dim = row.dimensions[currentDimension];
        return {
          type: row.type,
          sample12: row.sample12,
          sample3: row.sample3,
          value12: dim ? dim.value12 : null,
          value3: dim ? dim.value3 : null,
          delta: dim ? dim.delta : null,
        };
      }).sort((a, b) => {
        const av = a.value3 ?? -999;
        const bv = b.value3 ?? -999;
        return av - bv;
      });

      const categories = normalized.map(item => `${item.type}（${item.sample12}→${item.sample3}）`);
      const values12 = normalized.map(item => item.value12);
      const values3 = normalized.map(item => item.value3);
      const deltas = normalized.map(item => item.delta);

      detailChart.setOption({
        backgroundColor: 'transparent',
        animationDuration: 900,
        animationEasing: 'quarticOut',
        tooltip: {
          trigger: 'axis',
          axisPointer: { type: 'shadow' },
          formatter: params => {
            const idx = params[0].dataIndex;
            const row = normalized[idx];
            return [
              `<strong>${row.type}</strong>`,
              `样本：${row.sample12} → ${row.sample3}`,
              `1-2月：${fmt(row.value12)}`,
              `3月：${fmt(row.value3)}`,
              `变化：${deltaText(row.delta)}`
            ].join('<br/>');
          }
        },
        legend: {
          top: 10,
          textStyle: { color: colors.muted },
        },
        grid: { left: 180, right: 36, top: 58, bottom: 46 },
        xAxis: {
          type: 'value',
          min: 7.5,
          max: 10.1,
          axisLabel: { color: colors.muted },
          splitLine: { lineStyle: { color: colors.grid } },
        },
        yAxis: {
          type: 'category',
          data: categories,
          axisLabel: { color: colors.text, width: 170, overflow: 'truncate' },
          axisLine: { show: false },
          axisTick: { show: false },
        },
        series: [
          {
            name: '1-2月',
            type: 'bar',
            data: values12,
            barWidth: 16,
            itemStyle: { color: 'rgba(90,169,255,0.85)', borderRadius: [0, 8, 8, 0] },
            emphasis: { focus: 'series' },
          },
          {
            name: '3月',
            type: 'bar',
            data: values3,
            barWidth: 16,
            itemStyle: {
              color: params => {
                const delta = deltas[params.dataIndex];
                if (delta == null) return 'rgba(85,216,193,0.88)';
                return delta < 0 ? colors.fall : (delta > 0 ? colors.rise : colors.cyan);
              },
              borderRadius: [0, 8, 8, 0],
            },
            label: {
              show: true,
              position: 'right',
              color: '#ffffff',
              formatter: params => deltaText(deltas[params.dataIndex]),
            },
            emphasis: { focus: 'series' },
          }
        ]
      }, true);

      buildInsights(rows);
    }

    function bootstrap() {
      renderCards();
      renderOfficialChart();
      renderHeatmap();
      renderGroupButtons();
      renderDimensionButtons();
      renderDetailChart();
      window.addEventListener('resize', () => {
        officialChart.resize();
        heatmapChart.resize();
        detailChart.resize();
      });
    }

    bootstrap();
  </script>
</body>
</html>
'''


def build_html(data: dict[str, Any]) -> str:
    return HTML_TEMPLATE.replace("__DATA__", json.dumps(data, ensure_ascii=False))


def main() -> None:
    data = build_data()
    OUTPUT_HTML.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_HTML.write_text(build_html(data), encoding="utf-8")
    print(f"JSON: {OUTPUT_JSON}")
    print(f"HTML: {OUTPUT_HTML}")
    for card in data["cards"]:
        print(card)


if __name__ == "__main__":
    main()

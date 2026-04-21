from __future__ import annotations

from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_FILES = {
    '1-2月': ROOT / '汇总结果' / '1-2月' / '1-2月客户类型满意度汇总表.xlsx',
    '3月': ROOT / '汇总结果' / '3月' / '3月客户类型满意度汇总表.xlsx',
}
SAMPLE_FILE = ROOT / '汇总结果' / 'Q1' / 'Q1客户类型样本统计表.xlsx'
TARGET_GROUPS = ['一、会展客户', '二、餐饮客户', '三、G20峰会体验馆', '五、酒店客户']
DIMENSIONS = [
    ('product', '产品服务'),
    ('hardware', '硬件设施'),
    ('support', '配套服务'),
    ('smart', '智慧场馆/服务'),
    ('dining', '餐饮服务'),
]


def norm(value):
    if value in ('--', '—', '-', '', None):
        return None
    return value


def normalize_name(name: str) -> str:
    text = str(name or '').strip()
    mapping = {
        '会议活动主（承）办': '会议活动主（承）办',
        '参会客户': '参会客户',
        '展览活动主（承）办': '展览活动主（承）办',
        '酒店住宿团队': '酒店住宿团队',
        '酒店参会客户': '酒店参会客户',
        '酒店餐饮客户': '酒店餐饮客户',
        '酒店散客': '酒店散客',
    }
    return mapping.get(text, text)


def read_summary(path: Path) -> tuple[pd.DataFrame, dict[str, float | None]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    current_group = None
    records: list[dict[str, object]] = []
    total_row: dict[str, float | None] = {}
    for row in rows[2:]:
        first, second, total, product, hardware, support, smart, dining = row[:8]
        if first == '总分':
            total_row = {
                'score': norm(total),
                'product': norm(product),
                'hardware': norm(hardware),
                'support': norm(support),
                'smart': norm(smart),
                'dining': norm(dining),
            }
            break
        if first:
            current_group = str(first).strip()
        if not second:
            continue
        records.append({
            'group': current_group,
            'type': normalize_name(second),
            'score': norm(total),
            'product': norm(product),
            'hardware': norm(hardware),
            'support': norm(support),
            'smart': norm(smart),
            'dining': norm(dining),
        })
    return pd.DataFrame(records), total_row


def read_samples(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    idx_group = header.index('客户大类')
    idx_type = header.index('样本类型')
    idx_12 = header.index('1-2月')
    idx_3 = header.index('3月')
    records = []
    current_group = None
    for row in rows[2:]:
        group = row[idx_group]
        sample_type = row[idx_type]
        if group == '合计':
            break
        if group and group != '小计':
            current_group = str(group).strip()
        if not sample_type or sample_type in ('小计', '四、专项调研', '六、酒店暗访（次）', '会展流失主办客户'):
            continue
        records.append({
            'group': current_group,
            'type': normalize_name(sample_type),
            '1-2月_samples': int(row[idx_12] or 0),
            '3月_samples': int(row[idx_3] or 0),
        })
    return pd.DataFrame(records)


def weighted_mean(df: pd.DataFrame, value_col: str, weight_col: str):
    values = pd.to_numeric(df[value_col], errors='coerce')
    weights = pd.to_numeric(df[weight_col], errors='coerce')
    temp = pd.DataFrame({'v': values, 'w': weights}).dropna()
    temp = temp[temp['w'] > 0]
    if temp.empty:
        return None
    return float((temp['v'] * temp['w']).sum() / temp['w'].sum())


def fmt(v):
    if v is None or pd.isna(v):
        return '-'
    return f'{float(v):.3f}'


def main() -> None:
    summary_frames = []
    total_rows = []
    for period, path in SUMMARY_FILES.items():
        df, total_row = read_summary(path)
        df['period'] = period
        summary_frames.append(df)
        total_rows.append({
            'period': period,
            **total_row,
        })
    summary_df = pd.concat(summary_frames, ignore_index=True)
    total_df = pd.DataFrame(total_rows)
    sample_df = read_samples(SAMPLE_FILE)
    merged = summary_df.merge(sample_df, on=['group', 'type'], how='left')
    merged['samples'] = merged.apply(lambda r: r[f"{r['period']}_samples"], axis=1)

    print('\n=== 官方汇总表底部总分维度趋势（客户类型等权）===')
    total_rows_out = []
    for _, row in total_df.iterrows():
        total_rows_out.append({
            'period': row['period'],
            '产品服务': fmt(row['product']),
            '硬件设施': fmt(row['hardware']),
            '配套服务': fmt(row['support']),
            '智慧场馆/服务': fmt(row['smart']),
            '餐饮服务': fmt(row['dining']),
        })
    print(pd.DataFrame(total_rows_out).to_string(index=False))

    print('\n=== 四个客户大类维度趋势对比（样本加权）===')
    trend_rows = []
    for group in TARGET_GROUPS:
        for dim_col, dim_label in DIMENSIONS:
            g12 = merged[(merged['group'].eq(group)) & (merged['period'].eq('1-2月'))]
            g3 = merged[(merged['group'].eq(group)) & (merged['period'].eq('3月'))]
            v12 = weighted_mean(g12, dim_col, 'samples')
            v3 = weighted_mean(g3, dim_col, 'samples')
            trend_rows.append({
                '客户大类': group,
                '维度': dim_label,
                '1-2月': fmt(v12),
                '3月': fmt(v3),
                '变化': fmt(v3 - v12) if v12 is not None and v3 is not None else '-',
            })
    trend_df = pd.DataFrame(trend_rows)
    print(trend_df.to_string(index=False))


if __name__ == '__main__':
    main()

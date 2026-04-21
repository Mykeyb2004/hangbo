from __future__ import annotations

from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_FILES = {
    '1-2月': ROOT / '汇总结果' / '1-2月' / '1-2月客户类型满意度汇总表.xlsx',
    '3月': ROOT / '汇总结果' / '3月' / '3月客户类型满意度汇总表.xlsx',
}
SAMPLE_FILE = ROOT / '汇总结果' / 'Q1' / 'Q1客户类型样本统计表.xlsx'
TARGET_GROUPS = ['一、会展客户', '二、餐饮客户', '三、G20峰会体验馆', '五、酒店客户']
DIMENSIONS = ['score', 'product', 'hardware', 'support', 'smart', 'dining']
DIMENSION_LABELS = {
    'score': '总分',
    'product': '产品服务',
    'hardware': '硬件设施',
    'support': '配套服务',
    'smart': '智慧场馆/服务',
    'dining': '餐饮服务',
}
GROUP_ORDER = {name: idx for idx, name in enumerate(TARGET_GROUPS, start=1)}


def norm(value):
    if value in ('--', '—', '-', '', None):
        return None
    return value


def normalize_name(name: str) -> str:
    text = str(name or '').strip()
    mapping = {
        '会议活动主（承）办': '会议活动主（承）办',
        '参会客户': '参会客户',
        '展览活动主（承）办': '展览活动主（承）办',
        '酒店住宿团队': '酒店住宿团队',
        '酒店参会客户': '酒店参会客户',
        '酒店餐饮客户': '酒店餐饮客户',
        '酒店散客': '酒店散客',
    }
    return mapping.get(text, text)


def read_summary(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    current_group = None
    records: list[dict[str, object]] = []
    for row in rows[2:]:
        first, second, total, product, hardware, support, smart, dining = row[:8]
        if first == '总分':
            break
        if first:
            current_group = str(first).strip()
        if not second:
            continue
        records.append({
            '客户大类': current_group,
            '客户分组': normalize_name(second),
            '总分': norm(total),
            '产品服务': norm(product),
            '硬件设施': norm(hardware),
            '配套服务': norm(support),
            '智慧场馆/服务': norm(smart),
            '餐饮服务': norm(dining),
        })
    return pd.DataFrame(records)


def read_samples(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    idx_group = header.index('客户大类')
    idx_type = header.index('样本类型')
    idx_12 = header.index('1-2月')
    idx_3 = header.index('3月')
    records: list[dict[str, object]] = []
    current_group = None
    for row in rows[2:]:
        group = row[idx_group]
        sample_type = row[idx_type]
        if group == '合计':
            break
        if group and group != '小计':
            current_group = str(group).strip()
        if not sample_type or sample_type in ('小计', '四、专项调研', '六、酒店暗访（次）', '会展流失主办客户'):
            continue
        records.append({
            '客户大类': current_group,
            '客户分组': normalize_name(sample_type),
            '1-2月样本': int(row[idx_12] or 0),
            '3月样本': int(row[idx_3] or 0),
        })
    return pd.DataFrame(records)


def fmt(v):
    if v is None or pd.isna(v):
        return '-'
    return f'{float(v):.3f}'


def fmt_int(v):
    if v is None or pd.isna(v):
        return 0
    return int(v)


def main() -> None:
    frames = []
    for period, path in SUMMARY_FILES.items():
        df = read_summary(path)
        renamed = df.rename(columns={
            '总分': f'{period}-总分',
            '产品服务': f'{period}-产品服务',
            '硬件设施': f'{period}-硬件设施',
            '配套服务': f'{period}-配套服务',
            '智慧场馆/服务': f'{period}-智慧场馆/服务',
            '餐饮服务': f'{period}-餐饮服务',
        })
        frames.append(renamed)

    merged = frames[0].merge(frames[1], on=['客户大类', '客户分组'], how='outer')
    merged = merged.merge(read_samples(SAMPLE_FILE), on=['客户大类', '客户分组'], how='left')
    merged['group_order'] = merged['客户大类'].map(GROUP_ORDER)
    merged = merged[merged['客户大类'].isin(TARGET_GROUPS)].copy()
    merged = merged.sort_values(['group_order', '客户分组']).drop(columns=['group_order'])

    # full comparison table
    out = merged.copy()
    for base in ['总分', '产品服务', '硬件设施', '配套服务', '智慧场馆/服务', '餐饮服务']:
        col1 = f'1-2月-{base}'
        col3 = f'3月-{base}'
        out[f'{base}变化'] = [
            None if pd.isna(a) or pd.isna(b) else float(b) - float(a)
            for a, b in zip(out[col1], out[col3])
        ]

    display_cols = [
        '客户大类', '客户分组', '1-2月样本', '3月样本',
        '1-2月-总分', '3月-总分', '总分变化',
        '1-2月-产品服务', '3月-产品服务', '产品服务变化',
        '1-2月-硬件设施', '3月-硬件设施', '硬件设施变化',
        '1-2月-配套服务', '3月-配套服务', '配套服务变化',
        '1-2月-智慧场馆/服务', '3月-智慧场馆/服务', '智慧场馆/服务变化',
        '1-2月-餐饮服务', '3月-餐饮服务', '餐饮服务变化',
    ]

    printable = out[display_cols].copy()
    for col in printable.columns:
        if col not in ('客户大类', '客户分组', '1-2月样本', '3月样本'):
            printable[col] = printable[col].map(fmt)
    printable['1-2月样本'] = printable['1-2月样本'].fillna(0).astype(int)
    printable['3月样本'] = printable['3月样本'].fillna(0).astype(int)

    print('\n=== 按客户大类+客户分组排列的完整对比表 ===')
    print(printable.to_string(index=False))

    # compact table for ppt
    compact_rows = []
    for _, row in out.iterrows():
        compact_rows.append({
            '客户大类': row['客户大类'],
            '客户分组': row['客户分组'],
            '样本(1-2月→3月)': f"{fmt_int(row['1-2月样本'])}→{fmt_int(row['3月样本'])}",
            '总分': f"{fmt(row['1-2月-总分'])}→{fmt(row['3月-总分'])} ({fmt(row['总分变化'])})",
            '产品服务': f"{fmt(row['1-2月-产品服务'])}→{fmt(row['3月-产品服务'])} ({fmt(row['产品服务变化'])})",
            '硬件设施': f"{fmt(row['1-2月-硬件设施'])}→{fmt(row['3月-硬件设施'])} ({fmt(row['硬件设施变化'])})",
            '配套服务': f"{fmt(row['1-2月-配套服务'])}→{fmt(row['3月-配套服务'])} ({fmt(row['配套服务变化'])})",
            '智慧场馆/服务': f"{fmt(row['1-2月-智慧场馆/服务'])}→{fmt(row['3月-智慧场馆/服务'])} ({fmt(row['智慧场馆/服务变化'])})",
            '餐饮服务': f"{fmt(row['1-2月-餐饮服务'])}→{fmt(row['3月-餐饮服务'])} ({fmt(row['餐饮服务变化'])})",
        })
    compact = pd.DataFrame(compact_rows)
    print('\n=== PPT横向紧凑版 ===')
    print(compact.to_string(index=False))


if __name__ == '__main__':
    main()

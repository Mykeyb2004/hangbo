from __future__ import annotations

from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SUMMARY_FILES = {
    '1-2月': ROOT / '汇总结果' / '1-2月' / '1-2月客户类型满意度汇总表.xlsx',
    '3月': ROOT / '汇总结果' / '3月' / '3月客户类型满意度汇总表.xlsx',
}
SAMPLE_FILE = ROOT / '汇总结果' / 'Q1' / 'Q1客户类型样本统计表.xlsx'
TARGET_GROUPS = ['一、会展客户', '二、餐饮客户', '三、G20峰会体验馆', '五、酒店客户']


def normalize_name(name: str) -> str:
    text = str(name or '').strip()
    mapping = {
        '会议活动主（承）办': '会议活动主（承）办',
        '参会客户': '参会客户',
        '展览活动主（承）办': '展览活动主（承）办',
        '酒店住宿团队': '酒店住宿团队',
        '酒店参会客户': '酒店参会客户',
        '酒店餐饮客户': '酒店餐饮客户',
        '酒店散客': '酒店散客',
    }
    return mapping.get(text, text)


def norm(value):
    if value in ('--', '—', '-', '', None):
        return None
    return value


def read_summary(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    current_group = None
    records: list[dict[str, object]] = []
    for row in rows[2:]:
        first, second, total, product, hardware, support, smart, dining = row[:8]
        if first == '总分':
            break
        if first:
            current_group = str(first).strip()
        if not second:
            continue
        records.append({
            'group': current_group,
            'type': normalize_name(second),
            'score': norm(total),
            'product': norm(product),
            'hardware': norm(hardware),
            'support': norm(support),
            'smart': norm(smart),
            'dining': norm(dining),
        })
    return pd.DataFrame(records)


def read_sample_q1(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    idx_group = header.index('客户大类')
    idx_type = header.index('样本类型')
    idx_12 = header.index('1-2月')
    idx_3 = header.index('3月')
    records = []
    current_group = None
    for row in rows[2:]:
        group = row[idx_group]
        sample_type = row[idx_type]
        if group == '合计':
            break
        if group and group != '小计':
            current_group = str(group).strip()
        if not sample_type or sample_type in ('小计', '四、专项调研', '六、酒店暗访（次）', '会展流失主办客户'):
            continue
        records.append({
            'group': current_group,
            'type': normalize_name(sample_type),
            '1-2月_samples': int(row[idx_12] or 0),
            '3月_samples': int(row[idx_3] or 0),
        })
    return pd.DataFrame(records)


def simple_mean(df: pd.DataFrame, col: str):
    temp = pd.to_numeric(df[col], errors='coerce').dropna()
    return None if temp.empty else float(temp.mean())


def weighted_mean(df: pd.DataFrame, value_col: str, weight_col: str):
    values = pd.to_numeric(df[value_col], errors='coerce')
    weights = pd.to_numeric(df[weight_col], errors='coerce')
    temp = pd.DataFrame({'v': values, 'w': weights}).dropna()
    temp = temp[temp['w'] > 0]
    return None if temp.empty else float((temp['v'] * temp['w']).sum() / temp['w'].sum())


def fmt(v):
    if v is None or pd.isna(v):
        return '-'
    return f'{float(v):.3f}'


def main() -> None:
    summary_frames = []
    for period, path in SUMMARY_FILES.items():
        df = read_summary(path)
        df['period'] = period
        summary_frames.append(df)
    summary_df = pd.concat(summary_frames, ignore_index=True)
    sample_df = read_sample_q1(SAMPLE_FILE)
    merged = summary_df.merge(sample_df, on=['group', 'type'], how='left')
    merged['samples'] = merged.apply(lambda r: r[f"{r['period']}_samples"], axis=1)

    trend_rows = []
    for group in TARGET_GROUPS:
        for period in ['1-2月', '3月']:
            g = merged[(merged['group'].eq(group)) & (merged['period'].eq(period))].copy()
            trend_rows.append({
                'group': group,
                'period': period,
                'type_count': int(pd.to_numeric(g['score'], errors='coerce').notna().sum()),
                'sample_count': int(pd.to_numeric(g.loc[pd.to_numeric(g['score'], errors='coerce').notna(), 'samples'], errors='coerce').fillna(0).sum()),
                'simple_avg_score': simple_mean(g, 'score'),
                'weighted_avg_score': weighted_mean(g, 'score', 'samples'),
                'weighted_product': weighted_mean(g, 'product', 'samples'),
                'weighted_hardware': weighted_mean(g, 'hardware', 'samples'),
                'weighted_support': weighted_mean(g, 'support', 'samples'),
                'weighted_dining': weighted_mean(g, 'dining', 'samples'),
            })

    trend_df = pd.DataFrame(trend_rows)
    print('\n=== 四个客户大类趋势 ===')
    print(trend_df.to_string(index=False))

    print('\n=== 汇报版对比表 ===')
    report_rows = []
    for group in TARGET_GROUPS:
        g = trend_df[trend_df['group'].eq(group)].set_index('period')
        report_rows.append({
            '客户大类': group,
            '1-2月类型数': int(g.at['1-2月','type_count']),
            '1-2月样本': int(g.at['1-2月','sample_count']),
            '1-2月简单均分': fmt(g.at['1-2月','simple_avg_score']),
            '1-2月加权均分': fmt(g.at['1-2月','weighted_avg_score']),
            '3月类型数': int(g.at['3月','type_count']),
            '3月样本': int(g.at['3月','sample_count']),
            '3月简单均分': fmt(g.at['3月','simple_avg_score']),
            '3月加权均分': fmt(g.at['3月','weighted_avg_score']),
            '简单均分变化': fmt(g.at['3月','simple_avg_score'] - g.at['1-2月','simple_avg_score']) if pd.notna(g.at['1-2月','simple_avg_score']) and pd.notna(g.at['3月','simple_avg_score']) else '-',
            '加权均分变化': fmt(g.at['3月','weighted_avg_score'] - g.at['1-2月','weighted_avg_score']) if pd.notna(g.at['1-2月','weighted_avg_score']) and pd.notna(g.at['3月','weighted_avg_score']) else '-',
            '3月主要短板': ', '.join([
                x for x in [
                    f"产品{fmt(g.at['3月','weighted_product'])}" if pd.notna(g.at['3月','weighted_product']) else None,
                    f"硬件{fmt(g.at['3月','weighted_hardware'])}" if pd.notna(g.at['3月','weighted_hardware']) else None,
                    f"配套{fmt(g.at['3月','weighted_support'])}" if pd.notna(g.at['3月','weighted_support']) else None,
                    f"餐饮{fmt(g.at['3月','weighted_dining'])}" if pd.notna(g.at['3月','weighted_dining']) else None,
                ] if x is not None
            ])
        })
    report_df = pd.DataFrame(report_rows)
    print(report_df.to_string(index=False))

    print('\n=== 各大类3月主要短板维度（样本加权）===')
    for group in TARGET_GROUPS:
        g = trend_df[(trend_df['group'].eq(group)) & (trend_df['period'].eq('3月'))].iloc[0]
        print(f"\n[{group}]")
        for col, label in [('weighted_product','产品服务'),('weighted_hardware','硬件设施'),('weighted_support','配套服务'),('weighted_dining','餐饮服务')]:
            if pd.notna(g[col]):
                print(f"{label}: {g[col]:.3f}")


if __name__ == '__main__':
    main()

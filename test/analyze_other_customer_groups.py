from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import sys

import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

SUMMARY_FILES = {
    '1-2月': ROOT / '汇总结果' / '1-2月' / '1-2月客户类型满意度汇总表.xlsx',
    '3月': ROOT / '汇总结果' / '3月' / '3月客户类型满意度汇总表.xlsx',
}
SAMPLE_FILE = ROOT / '汇总结果' / 'Q1' / 'Q1客户类型样本统计表.xlsx'
TARGET_GROUPS = ['二、餐饮客户', '三、G20峰会体验馆', '五、酒店客户']


@dataclass
class GroupSummary:
    period: str
    group: str
    type_count: int
    total_samples: int
    simple_avg_score: float | None
    weighted_avg_score: float | None
    simple_dining: float | None
    simple_hardware: float | None
    simple_support: float | None
    simple_product: float | None


def normalize_name(name: str) -> str:
    text = str(name or '').strip()
    mapping = {
        '会议活动主（承）办': '会议活动主（承）办',
        '参会客户': '参会客户',
        '展览活动主（承）办': '展览活动主（承）办',
        '酒店住宿团队': '酒店住宿团队',
        '酒店参会客户': '酒店参会客户',
        '酒店餐饮客户': '酒店餐饮客户',
        '酒店散客': '酒店散客',
    }
    return mapping.get(text, text)


def read_summary(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    data_rows = rows[2:]
    records: list[dict[str, object]] = []
    current_group = None
    for row in data_rows:
        first, second, total, product, hardware, support, smart, dining = row[:8]
        if first == '总分':
            break
        if first:
            current_group = str(first).strip()
        if not second:
            continue
        def norm(value):
            if value in ('--', '—', '-', ''):
                return None
            return value
        records.append({
            'group': current_group,
            'type': normalize_name(second),
            'score': norm(total),
            'product': norm(product),
            'hardware': norm(hardware),
            'support': norm(support),
            'smart': norm(smart),
            'dining': norm(dining),
        })
    return pd.DataFrame(records)


def read_sample_q1(path: Path) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]
    idx_group = header.index('客户大类')
    idx_type = header.index('样本类型')
    idx_12 = header.index('1-2月')
    idx_3 = header.index('3月')
    records = []
    current_group = None
    for row in rows[2:]:
        group = row[idx_group]
        sample_type = row[idx_type]
        if group == '合计':
            break
        if group and group != '小计':
            current_group = str(group).strip()
        if not sample_type or sample_type in ('小计', '四、专项调研', '六、酒店暗访（次）', '会展流失主办客户'):
            continue
        records.append({
            'group': current_group,
            'type': normalize_name(sample_type),
            '1-2月_samples': int(row[idx_12] or 0),
            '3月_samples': int(row[idx_3] or 0),
        })
    return pd.DataFrame(records)


def weighted_mean(df: pd.DataFrame, value_col: str, weight_col: str) -> float | None:
    temp = df[df[value_col].notna() & df[weight_col].notna() & (df[weight_col] > 0)]
    if temp.empty:
        return None
    return float((temp[value_col] * temp[weight_col]).sum() / temp[weight_col].sum())


def simple_mean(df: pd.DataFrame, value_col: str) -> float | None:
    temp = df[df[value_col].notna()]
    if temp.empty:
        return None
    return float(temp[value_col].mean())


def fmt(v: float | None) -> str:
    if v is None or pd.isna(v):
        return '-'
    return f'{v:.3f}'


def main() -> None:
    summary_frames = []
    for period, path in SUMMARY_FILES.items():
        df = read_summary(path)
        df['period'] = period
        summary_frames.append(df)
    summary_df = pd.concat(summary_frames, ignore_index=True)
    sample_df = read_sample_q1(SAMPLE_FILE)

    merged = summary_df.merge(sample_df, on=['group', 'type'], how='left')
    merged['samples'] = merged.apply(lambda r: r[f"{r['period']}_samples"], axis=1)

    rows: list[GroupSummary] = []
    for period in ['1-2月', '3月']:
        period_df = merged[merged['period'].eq(period)].copy()
        for group in TARGET_GROUPS:
            group_df = period_df[period_df['group'].eq(group)].copy()
            rows.append(GroupSummary(
                period=period,
                group=group,
                type_count=int(group_df['score'].notna().sum()),
                total_samples=int(group_df.loc[group_df['score'].notna(), 'samples'].sum()),
                simple_avg_score=simple_mean(group_df, 'score'),
                weighted_avg_score=weighted_mean(group_df, 'score', 'samples'),
                simple_dining=simple_mean(group_df, 'dining'),
                simple_hardware=simple_mean(group_df, 'hardware'),
                simple_support=simple_mean(group_df, 'support'),
                simple_product=simple_mean(group_df, 'product'),
            ))

    result_df = pd.DataFrame([r.__dict__ for r in rows])
    print('\n=== 其他客户大类总体情况 ===')
    print(result_df.to_string(index=False))

    print('\n=== 客户类型明细（1-2月 vs 3月）===')
    detail = merged[merged['group'].isin(TARGET_GROUPS)][['period','group','type','score','samples','product','hardware','support','dining']]
    print(detail.sort_values(['group','period','samples'], ascending=[True, True, False]).to_string(index=False))

    print('\n=== 变化拆解 ===')
    for group in TARGET_GROUPS:
        g = result_df[result_df['group'].eq(group)].set_index('period')
        if {'1-2月','3月'}.issubset(g.index):
            print(f"\n[{group}]")
            for col in ['simple_avg_score','weighted_avg_score','simple_product','simple_hardware','simple_support','simple_dining']:
                v1 = g.at['1-2月', col]
                v2 = g.at['3月', col]
                if pd.isna(v1) or pd.isna(v2):
                    continue
                print(f"{col}: {v1:.3f} -> {v2:.3f} ({v2-v1:+.3f})")

    print('\n=== 3月低分客户类型（其他大类）===')
    march_low = merged[(merged['period'].eq('3月')) & (merged['group'].isin(TARGET_GROUPS)) & merged['score'].notna()].copy()
    march_low = march_low.sort_values(['score','samples'])
    print(march_low[['group','type','score','samples','product','hardware','support','dining']].to_string(index=False))


if __name__ == '__main__':
    main()

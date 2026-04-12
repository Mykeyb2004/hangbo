from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from survey_stats import OVERALL_FILL, SECTION_FILL, excel_round, mean_ignore_empty, normalize_output_dir

DEFAULT_SUMMARY_TITLE = "杭博客户类型满意度情况表"
DEFAULT_OUTPUT_NAME = "客户类型满意度汇总表.xlsx"
SUMMARY_SHEET_NAME = "汇总表"
SUMMARY_COLUMNS = (
    "总分",
    "产品服务",
    "硬件设施",
    "配套服务",
    "智慧场馆/服务",
    "餐饮服务",
)

SUMMARY_HEADER_FILL = PatternFill(fill_type="solid", start_color="B32046", end_color="B32046")
SUMMARY_SIDE_FILL = PatternFill(fill_type="solid", start_color="B32046", end_color="B32046")
SUMMARY_BODY_FILL = PatternFill(fill_type="solid", start_color="F4E8E8", end_color="F4E8E8")
SUMMARY_NA_FILL = PatternFill(fill_type="solid", start_color="BFBFBF", end_color="BFBFBF")
SUMMARY_NUMBER_FORMAT = "0.00"
SUMMARY_CHINESE_FONT_NAME = "楷体"
SUMMARY_LATIN_FONT_NAME = "Times New Roman"
SUMMARY_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="FFFFFF"),
)
SUMMARY_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUMMARY_TITLE_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=20, bold=True, color="FFFFFF")
SUMMARY_HEADER_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=15, bold=True, color="FFFFFF")
SUMMARY_SIDE_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=14, bold=True, color="FFFFFF")
SUMMARY_BODY_FONT = Font(name=SUMMARY_LATIN_FONT_NAME, size=15, color="000000")

REPORT_HEADER_NAMES = {"指标", "满意度"}


@dataclass(frozen=True)
class ReportSnapshot:
    role_name: str
    source_path: Path
    total: float | None
    sections: dict[str, float | None]
    metrics: dict[str, float | None]


@dataclass(frozen=True)
class ValueSelector:
    source_type: Literal["overall", "section", "metric"]
    source_name: str | None = None


@dataclass(frozen=True)
class SummaryRowDefinition:
    category_label: str
    display_name: str
    source_aliases: tuple[str, ...]
    selectors: dict[str, tuple[ValueSelector, ...]]


@dataclass(frozen=True)
class SummaryRowResult:
    category_label: str
    display_name: str
    values: dict[str, float | None]
    applicable_columns: frozenset[str]


def normalize_text(value: str | None) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .strip()
        .replace(" ", "")
        .replace("（", "(")
        .replace("）", ")")
        .replace("／", "/")
    )


def normalize_color_value(color: str | None) -> str:
    if color is None:
        return ""
    color_text = str(color).upper()
    if len(color_text) == 6:
        return f"00{color_text}"
    return color_text


def normalize_fill_color(cell) -> str:
    return normalize_color_value(cell.fill.start_color.rgb)


OVERALL_FILL_RGB = normalize_color_value(OVERALL_FILL.start_color.rgb)
SECTION_FILL_RGB = normalize_color_value(SECTION_FILL.start_color.rgb)


def coerce_score(value: object) -> float | None:
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(numeric):
        return None
    return excel_round(float(numeric))


def row_kind(worksheet, row_index: int, indicator_column: int) -> Literal["overall", "section", "metric"]:
    if row_index == 2:
        return "overall"

    color = normalize_fill_color(worksheet.cell(row=row_index, column=indicator_column))
    if color == OVERALL_FILL_RGB:
        return "overall"
    if color == SECTION_FILL_RGB:
        return "section"
    return "metric"


def load_report_snapshot(report_path: Path) -> ReportSnapshot | None:
    workbook = load_workbook(report_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_map: dict[str, int] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=column_index).value
        if header_value is None:
            continue
        header_map[str(header_value).strip()] = column_index

    if not REPORT_HEADER_NAMES.issubset(header_map):
        return None

    indicator_column = header_map["指标"]
    satisfaction_column = header_map["满意度"]
    role_name = ""
    total: float | None = None
    sections: dict[str, float | None] = {}
    metrics: dict[str, float | None] = {}

    for row_index in range(2, worksheet.max_row + 1):
        indicator_value = worksheet.cell(row=row_index, column=indicator_column).value
        indicator_name = str(indicator_value).strip() if indicator_value is not None else ""
        if not indicator_name:
            continue

        score = coerce_score(worksheet.cell(row=row_index, column=satisfaction_column).value)
        current_row_kind = row_kind(worksheet, row_index, indicator_column)
        if current_row_kind == "overall":
            role_name = indicator_name
            total = score
        elif current_row_kind == "section":
            sections[indicator_name] = score
        else:
            metrics[indicator_name] = score

    if not role_name:
        return None

    return ReportSnapshot(
        role_name=role_name,
        source_path=report_path.resolve(),
        total=total,
        sections=sections,
        metrics=metrics,
    )


def iter_excel_paths(input_dir: Path, recursive: bool = False) -> list[Path]:
    pattern = "**/*.xlsx" if recursive else "*.xlsx"
    paths = sorted(path for path in input_dir.glob(pattern) if path.is_file())
    return [path for path in paths if not path.name.startswith("~$")]


def load_report_snapshots(input_dir: Path, recursive: bool = False) -> tuple[ReportSnapshot, ...]:
    report_snapshots: list[ReportSnapshot] = []
    for report_path in iter_excel_paths(input_dir, recursive=recursive):
        snapshot = load_report_snapshot(report_path)
        if snapshot is not None:
            report_snapshots.append(snapshot)
    return tuple(report_snapshots)


def overall_selector() -> tuple[ValueSelector, ...]:
    return (ValueSelector("overall"),)


def section_selector(*section_names: str) -> tuple[ValueSelector, ...]:
    return tuple(ValueSelector("section", section_name) for section_name in section_names)


def dining_selector() -> tuple[ValueSelector, ...]:
    return (
        ValueSelector("section", "餐饮服务"),
        ValueSelector("metric", "餐饮服务"),
    )


EVENT_PRODUCT_SELECTORS = section_selector("会展服务", "会场服务")
HOTEL_PRODUCT_SELECTORS = section_selector("入住服务")
G20_PRODUCT_SELECTORS = section_selector("旅游服务")
HARDWARE_SELECTORS = section_selector("硬件设施")
SUPPORT_SELECTORS = section_selector("配套服务")
SMART_SELECTORS = section_selector("智慧场馆", "智慧服务")

SUMMARY_ROW_DEFINITIONS: tuple[SummaryRowDefinition, ...] = (
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="展览活动主（承）办",
        source_aliases=("展览主承办", "展览活动主（承）办"),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="参展商",
        source_aliases=("参展商",),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="专业观众",
        source_aliases=("专业观众",),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="会展服务商",
        source_aliases=("会展服务商",),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="会议活动主（承）办",
        source_aliases=("会议主承办", "会议活动主（承）办"),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="一、会展客户",
        display_name="参会客户",
        source_aliases=("参会人员", "参会客户"),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="二、餐饮客户",
        display_name="商务简餐",
        source_aliases=("商务简餐",),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="二、餐饮客户",
        display_name="特色美食廊",
        source_aliases=("特色美食廊",),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="二、餐饮客户",
        display_name="宴会",
        source_aliases=("宴会",),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="二、餐饮客户",
        display_name="婚宴",
        source_aliases=("婚宴",),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="二、餐饮客户",
        display_name="自助餐",
        source_aliases=("自助餐",),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="三、G20峰会体验馆",
        display_name="旅行社工作人员",
        source_aliases=("旅行社工作人员",),
        selectors={
            "总分": overall_selector(),
            "产品服务": G20_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
        },
    ),
    SummaryRowDefinition(
        category_label="三、G20峰会体验馆",
        display_name="游客",
        source_aliases=("游客",),
        selectors={
            "总分": overall_selector(),
            "产品服务": G20_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
        },
    ),
    SummaryRowDefinition(
        category_label="四、专项调研",
        display_name="",
        source_aliases=(),
        selectors={},
    ),
    SummaryRowDefinition(
        category_label="五、酒店客户",
        display_name="散客",
        source_aliases=("散客",),
        selectors={
            "总分": overall_selector(),
            "产品服务": HOTEL_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="五、酒店客户",
        display_name="住宿团队",
        source_aliases=("住宿团队",),
        selectors={
            "总分": overall_selector(),
            "产品服务": HOTEL_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="五、酒店客户",
        display_name="酒店会议活动主（承）办",
        source_aliases=("酒店会议主承办", "酒店会议活动主（承）办"),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="五、酒店客户",
        display_name="酒店参会客户",
        source_aliases=("酒店参会客户", "酒店参会人员"),
        selectors={
            "总分": overall_selector(),
            "产品服务": EVENT_PRODUCT_SELECTORS,
            "硬件设施": HARDWARE_SELECTORS,
            "配套服务": SUPPORT_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
    SummaryRowDefinition(
        category_label="五、酒店客户",
        display_name="餐饮客户",
        source_aliases=("酒店宴会", "酒店自助餐"),
        selectors={
            "总分": overall_selector(),
            "硬件设施": HARDWARE_SELECTORS,
            "智慧场馆/服务": SMART_SELECTORS,
            "餐饮服务": dining_selector(),
        },
    ),
)


def build_report_index(reports: tuple[ReportSnapshot, ...]) -> dict[str, dict[Path, ReportSnapshot]]:
    report_index: dict[str, dict[Path, ReportSnapshot]] = {}
    for report in reports:
        aliases = {
            normalize_text(report.role_name),
            normalize_text(report.source_path.stem),
        }
        for alias in aliases:
            if not alias:
                continue
            report_index.setdefault(alias, {})[report.source_path] = report
    return report_index


def select_reports_for_row(
    report_index: dict[str, dict[Path, ReportSnapshot]],
    row_definition: SummaryRowDefinition,
) -> tuple[ReportSnapshot, ...]:
    matched_reports: dict[Path, ReportSnapshot] = {}
    for alias in row_definition.source_aliases:
        for source_path, report in report_index.get(normalize_text(alias), {}).items():
            matched_reports[source_path] = report
    return tuple(matched_reports[source_path] for source_path in sorted(matched_reports))


def extract_report_value(report: ReportSnapshot, selectors: tuple[ValueSelector, ...]) -> float | None:
    for selector in selectors:
        if selector.source_type == "overall":
            if report.total is not None:
                return report.total
            continue

        if selector.source_type == "section":
            value = report.sections.get(selector.source_name or "")
            if value is not None:
                return value
            continue

        value = report.metrics.get(selector.source_name or "")
        if value is not None:
            return value
    return None


def build_summary_rows(reports: tuple[ReportSnapshot, ...]) -> tuple[SummaryRowResult, ...]:
    report_index = build_report_index(reports)
    summary_rows: list[SummaryRowResult] = []
    for row_definition in SUMMARY_ROW_DEFINITIONS:
        matched_reports = select_reports_for_row(report_index, row_definition)
        values: dict[str, float | None] = {}
        for column_name in SUMMARY_COLUMNS:
            selectors = row_definition.selectors.get(column_name)
            if selectors is None:
                values[column_name] = None
                continue
            extracted_values = [extract_report_value(report, selectors) for report in matched_reports]
            values[column_name] = mean_ignore_empty(extracted_values)

        summary_rows.append(
            SummaryRowResult(
                category_label=row_definition.category_label,
                display_name=row_definition.display_name,
                values=values,
                applicable_columns=frozenset(row_definition.selectors),
            )
        )

    return tuple(summary_rows)


def build_total_values(rows: tuple[SummaryRowResult, ...]) -> dict[str, float | None]:
    total_values: dict[str, float | None] = {}
    data_rows = [row for row in rows if row.display_name]
    for column_name in SUMMARY_COLUMNS:
        total_values[column_name] = mean_ignore_empty(
            [row.values[column_name] for row in data_rows]
        )
    return total_values


def build_summary_dataframe(rows: tuple[SummaryRowResult, ...]) -> pd.DataFrame:
    result_rows: list[dict[str, object]] = []
    for row in rows:
        row_data: dict[str, object] = {
            "客户大类": row.category_label,
            "样本类型": row.display_name,
        }
        row_data.update(row.values)
        result_rows.append(row_data)

    total_row = {"客户大类": "", "样本类型": "总分"}
    total_row.update(build_total_values(rows))
    result_rows.append(total_row)

    return pd.DataFrame(
        result_rows,
        columns=("客户大类", "样本类型", *SUMMARY_COLUMNS),
    )


def apply_common_style(cell, *, fill: PatternFill | None = None, font: Font | None = None) -> None:
    cell.alignment = SUMMARY_CENTER_ALIGNMENT
    cell.border = SUMMARY_BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def build_inline_font(font: Font, *, font_name: str | None = None) -> InlineFont:
    color = font.color.rgb if font.color is not None else None
    return InlineFont(
        rFont=font_name or font.name,
        b=font.bold,
        i=font.italic,
        color=color,
        sz=font.size,
    )


def is_ascii_alnum_character(character: str) -> bool:
    return character.isascii() and character.isalnum()


def build_styled_text(value: str, *, chinese_font: Font) -> str | CellRichText:
    if not value or not any(is_ascii_alnum_character(character) for character in value):
        return value

    chinese_inline_font = build_inline_font(chinese_font)
    latin_inline_font = build_inline_font(chinese_font, font_name=SUMMARY_LATIN_FONT_NAME)
    segments: list[TextBlock] = []
    current_font = None
    current_text: list[str] = []

    for character in value:
        target_font = latin_inline_font if is_ascii_alnum_character(character) else chinese_inline_font
        if current_font is not None and target_font == current_font:
            current_text.append(character)
            continue
        if current_font is not None:
            segments.append(TextBlock(current_font, "".join(current_text)))
        current_font = target_font
        current_text = [character]

    if current_font is not None:
        segments.append(TextBlock(current_font, "".join(current_text)))

    if len(segments) == 1:
        return value
    return CellRichText(*segments)


def set_text_cell(cell, value: str | None, *, fill: PatternFill, font: Font) -> None:
    if value:
        cell.value = build_styled_text(value, chinese_font=font)
    else:
        cell.value = None
    apply_common_style(cell, fill=fill, font=font)


def style_summary_worksheet(worksheet, rows: tuple[SummaryRowResult, ...]) -> None:
    worksheet.merge_cells("A1:H1")
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 28
    set_text_cell(worksheet["A1"], DEFAULT_SUMMARY_TITLE, fill=SUMMARY_HEADER_FILL, font=SUMMARY_TITLE_FONT)
    for column_index in range(2, 9):
        apply_common_style(
            worksheet.cell(row=1, column=column_index),
            fill=SUMMARY_HEADER_FILL,
            font=SUMMARY_TITLE_FONT,
        )

    worksheet.merge_cells("A2:B2")
    set_text_cell(worksheet["A2"], "样本类型", fill=SUMMARY_HEADER_FILL, font=SUMMARY_HEADER_FONT)
    apply_common_style(worksheet["B2"], fill=SUMMARY_HEADER_FILL, font=SUMMARY_HEADER_FONT)
    for column_index, column_name in enumerate(SUMMARY_COLUMNS, start=3):
        set_text_cell(
            worksheet.cell(row=2, column=column_index),
            column_name,
            fill=SUMMARY_HEADER_FILL,
            font=SUMMARY_HEADER_FONT,
        )

    data_start_row = 3
    for row_offset, row in enumerate(rows):
        excel_row = data_start_row + row_offset
        worksheet.row_dimensions[excel_row].height = 24
        set_text_cell(
            worksheet.cell(row=excel_row, column=1),
            row.category_label,
            fill=SUMMARY_SIDE_FILL,
            font=SUMMARY_SIDE_FONT,
        )
        set_text_cell(
            worksheet.cell(row=excel_row, column=2),
            row.display_name or None,
            fill=SUMMARY_SIDE_FILL,
            font=SUMMARY_SIDE_FONT,
        )

        for column_index, column_name in enumerate(SUMMARY_COLUMNS, start=3):
            value = row.values[column_name]
            if value is not None:
                worksheet.cell(row=excel_row, column=column_index, value=value)
            fill = SUMMARY_BODY_FILL
            if column_name not in row.applicable_columns:
                fill = SUMMARY_NA_FILL
            apply_common_style(
                worksheet.cell(row=excel_row, column=column_index),
                fill=fill,
                font=SUMMARY_BODY_FONT,
            )
            worksheet.cell(row=excel_row, column=column_index).number_format = SUMMARY_NUMBER_FORMAT

    total_row_index = data_start_row + len(rows)
    worksheet.row_dimensions[total_row_index].height = 28
    worksheet.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=2)
    set_text_cell(
        worksheet.cell(row=total_row_index, column=1),
        "总分",
        fill=SUMMARY_HEADER_FILL,
        font=SUMMARY_HEADER_FONT,
    )
    total_values = build_total_values(rows)
    apply_common_style(
        worksheet.cell(row=total_row_index, column=2),
        fill=SUMMARY_HEADER_FILL,
        font=SUMMARY_HEADER_FONT,
    )
    for column_index, column_name in enumerate(SUMMARY_COLUMNS, start=3):
        value = total_values[column_name]
        if value is not None:
            worksheet.cell(row=total_row_index, column=column_index, value=value)
        apply_common_style(
            worksheet.cell(row=total_row_index, column=column_index),
            fill=SUMMARY_BODY_FILL,
            font=SUMMARY_BODY_FONT,
        )
        worksheet.cell(row=total_row_index, column=column_index).number_format = SUMMARY_NUMBER_FORMAT

    merge_category_cells(worksheet, rows, data_start_row)
    worksheet.freeze_panes = "C3"

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 36
    worksheet.column_dimensions["C"].width = 11
    worksheet.column_dimensions["D"].width = 15
    worksheet.column_dimensions["E"].width = 15
    worksheet.column_dimensions["F"].width = 14
    worksheet.column_dimensions["G"].width = 16
    worksheet.column_dimensions["H"].width = 12


def merge_category_cells(worksheet, rows: tuple[SummaryRowResult, ...], data_start_row: int) -> None:
    if not rows:
        return

    current_category = rows[0].category_label
    merge_start = data_start_row
    for index, row in enumerate(rows[1:], start=1):
        excel_row = data_start_row + index
        if row.category_label != current_category:
            if merge_start != excel_row - 1:
                worksheet.merge_cells(
                    start_row=merge_start,
                    start_column=1,
                    end_row=excel_row - 1,
                    end_column=1,
                )
            current_category = row.category_label
            merge_start = excel_row

    final_row = data_start_row + len(rows) - 1
    if merge_start != final_row:
        worksheet.merge_cells(
            start_row=merge_start,
            start_column=1,
            end_row=final_row,
            end_column=1,
        )


def ensure_output_path(output_dir: Path, output_name: str) -> Path:
    final_output_dir = normalize_output_dir(output_dir)
    file_name = output_name if output_name.lower().endswith(".xlsx") else f"{output_name}.xlsx"
    return final_output_dir / file_name


def generate_summary_report(
    input_dir: Path,
    output_dir: Path,
    output_name: str = DEFAULT_OUTPUT_NAME,
    recursive: bool = False,
) -> Path:
    reports = load_report_snapshots(input_dir, recursive=recursive)
    if not reports:
        raise ValueError("输入目录下未找到可识别的统计结果 xlsx 文件。")

    rows = build_summary_rows(reports)
    output_path = ensure_output_path(output_dir, output_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SUMMARY_SHEET_NAME
    style_summary_worksheet(worksheet, rows)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="汇总指定目录下的客户群体统计结果 xlsx，生成客户类型汇总表。")
    parser.add_argument("--input-dir", type=Path, required=True, help="输入目录：survey_stats.py 导出的单群体统计 xlsx 所在目录")
    parser.add_argument("--output-dir", type=Path, required=True, help="输出目录")
    parser.add_argument("--output-name", default=DEFAULT_OUTPUT_NAME, help=f"输出文件名，默认 {DEFAULT_OUTPUT_NAME}")
    parser.add_argument("--recursive", action="store_true", help="递归扫描子目录下的 xlsx 文件")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = generate_summary_report(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        output_name=args.output_name,
        recursive=args.recursive,
    )
    print(f"汇总表已保存到: {output_path}")


if __name__ == "__main__":
    main()

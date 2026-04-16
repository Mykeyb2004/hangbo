from __future__ import annotations

import argparse
import re
import tomllib
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from survey_customer_category_rules import CUSTOMER_CATEGORY_RULE_BY_NAME, CustomerCategoryRule

DEFAULT_SAMPLE_TABLE_TITLE = "杭博客户类型样本统计表"
DEFAULT_SAMPLE_TABLE_OUTPUT_NAME = "客户类型样本统计表.xlsx"
DEFAULT_SAMPLE_TABLE_CONFIG_PATH = Path(__file__).with_name("sample_table.default.toml")
SAMPLE_TABLE_SHEET_NAME = "样本统计"
DEFAULT_SHEET_NAME = "问卷数据"
YEAR_HEADER = "年份"
MONTH_HEADER = "月份"
DEFAULT_GROUP_YEAR = "2026"

COUNT_NUMBER_FORMAT = "0"
PERCENT_NUMBER_FORMAT = "0.00%"
MONTH_RANGE_PATTERN = re.compile(r"^(\d{1,2})\s*[-~至到]\s*(\d{1,2})月?$")
MONTH_SINGLE_PATTERN = re.compile(r"^(\d{1,2})月?$")
SUMMARY_HEADER_FILL = PatternFill(fill_type="solid", start_color="B32046", end_color="B32046")
SUMMARY_SIDE_FILL = PatternFill(fill_type="solid", start_color="B32046", end_color="B32046")
SUMMARY_BODY_FILL = PatternFill(fill_type="solid", start_color="F4E8E8", end_color="F4E8E8")
SUMMARY_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="FFFFFF"),
)
SUMMARY_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUMMARY_CHINESE_FONT_NAME = "楷体"
SUMMARY_LATIN_FONT_NAME = "Times New Roman"
SUMMARY_TITLE_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=20, bold=True, color="FFFFFF")
SUMMARY_HEADER_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=15, bold=True, color="FFFFFF")
SUMMARY_SIDE_FONT = Font(name=SUMMARY_CHINESE_FONT_NAME, size=14, bold=True, color="FFFFFF")
SUMMARY_BODY_FONT = Font(name=SUMMARY_LATIN_FONT_NAME, size=15, color="000000")


@dataclass(frozen=True)
class SampleTableRowConfig:
    category_label: str
    display_name: str
    target_sample_size: int
    rule_name: str | None = None
    actual_count_override: int | None = None


@dataclass(frozen=True)
class SampleTableConfig:
    title: str
    sheet_name: str
    output_name: str
    rows: tuple[SampleTableRowConfig, ...]


@dataclass(frozen=True)
class SampleTableRowResult:
    category_label: str
    display_name: str
    target_sample_size: int
    actual_count: int
    group_counts: dict[str, int]


@dataclass(frozen=True)
class SampleTableGroupSpec:
    label: str
    year: str
    month: str


@dataclass(frozen=True)
class SampleTableBuildResult:
    group_labels: tuple[str, ...]
    rows: tuple[SampleTableRowResult, ...]


@dataclass(frozen=True)
class PreparedSampleTableRow:
    row_config: SampleTableRowConfig
    actual_count: int
    dataframe: pd.DataFrame | None = None
    mask: pd.Series | None = None


@dataclass(frozen=True)
class RowGroup:
    category_label: str
    rows: tuple[SampleTableRowResult, ...]


def apply_common_style(cell, *, fill: PatternFill | None = None, font: Font | None = None) -> None:
    cell.alignment = SUMMARY_CENTER_ALIGNMENT
    cell.border = SUMMARY_BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def set_text_cell(cell, value: str | None, *, fill: PatternFill, font: Font) -> None:
    cell.value = value or None
    apply_common_style(cell, fill=fill, font=font)


def load_sample_table_config(
    config_path: Path = DEFAULT_SAMPLE_TABLE_CONFIG_PATH,
) -> SampleTableConfig:
    data = tomllib.loads(config_path.read_text(encoding="utf-8"))
    rows_data = data.get("rows")
    if not isinstance(rows_data, list) or not rows_data:
        raise ValueError("样本统计配置缺少 rows 列表。")

    rows: list[SampleTableRowConfig] = []
    for row_data in rows_data:
        rows.append(
            SampleTableRowConfig(
                category_label=str(row_data["category_label"]).strip(),
                display_name=str(row_data.get("display_name", "")).strip(),
                target_sample_size=int(row_data["target_sample_size"]),
                rule_name=(
                    str(row_data["rule_name"]).strip()
                    if row_data.get("rule_name") is not None
                    else None
                ),
                actual_count_override=(
                    int(row_data["actual_count_override"])
                    if row_data.get("actual_count_override") is not None
                    else None
                ),
            )
        )

    return SampleTableConfig(
        title=str(data.get("title", DEFAULT_SAMPLE_TABLE_TITLE)).strip() or DEFAULT_SAMPLE_TABLE_TITLE,
        sheet_name=str(data.get("sheet_name", SAMPLE_TABLE_SHEET_NAME)).strip() or SAMPLE_TABLE_SHEET_NAME,
        output_name=str(data.get("output_name", DEFAULT_SAMPLE_TABLE_OUTPUT_NAME)).strip()
        or DEFAULT_SAMPLE_TABLE_OUTPUT_NAME,
        rows=tuple(rows),
    )


def normalize_month_value(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.isdigit():
        return str(int(text))
    if text.endswith(".0"):
        numeric_text = text[:-2]
        if numeric_text.isdigit():
            return str(int(numeric_text))
    range_match = MONTH_RANGE_PATTERN.fullmatch(text)
    if range_match:
        start_month, end_month = range_match.groups()
        return f"{int(start_month)}-{int(end_month)}"
    single_match = MONTH_SINGLE_PATTERN.fullmatch(text)
    if single_match:
        return str(int(single_match.group(1)))
    return text


def normalize_year_value(value: object, *, default_year: str) -> str:
    normalized_value = normalize_month_value(value)
    return normalized_value or str(default_year).strip()


def find_column(df: pd.DataFrame, header_name: str) -> object | None:
    for column_name in df.columns:
        if str(column_name).strip() == header_name:
            return column_name
    return None


def iter_row_year_month_values(
    df: pd.DataFrame,
    mask: pd.Series,
    *,
    default_year: str,
) -> list[tuple[str, str]]:
    month_column = find_column(df, MONTH_HEADER)
    if month_column is None:
        return []

    year_column = find_column(df, YEAR_HEADER)
    year_month_values: list[tuple[str, str]] = []
    for row_index in df.index[mask]:
        month_text = normalize_month_value(df.at[row_index, month_column])
        if not month_text:
            continue

        year_value = df.at[row_index, year_column] if year_column is not None else None
        year_text = normalize_year_value(year_value, default_year=default_year)
        year_month_values.append((year_text, month_text))
    return year_month_values


def build_auto_group_label(year_text: str, month_text: str, *, include_year: bool) -> str:
    month_label = f"{month_text}月" if month_text else ""
    return f"{year_text}年{month_label}" if include_year else month_label


def build_auto_sample_groups(
    year_month_values: set[tuple[str, str]],
    *,
    default_year: str,
) -> tuple[SampleTableGroupSpec, ...]:
    if not year_month_values:
        return ()

    distinct_years = {year_text for year_text, _ in year_month_values}
    include_year = len(distinct_years) > 1 or any(year_text != default_year for year_text in distinct_years)
    groups = [
        SampleTableGroupSpec(
            label=build_auto_group_label(year_text, month_text, include_year=include_year),
            year=year_text,
            month=month_text,
        )
        for year_text, month_text in year_month_values
    ]
    return tuple(sorted(groups, key=lambda item: item.label))


def parse_sample_group_specs(
    raw_group_values: list[str] | None,
    *,
    default_year: str = DEFAULT_GROUP_YEAR,
) -> tuple[SampleTableGroupSpec, ...]:
    if not raw_group_values:
        return ()

    group_specs: list[SampleTableGroupSpec] = []
    seen_labels: set[str] = set()

    for raw_value in raw_group_values:
        text = str(raw_value).strip()
        if not text:
            continue
        if "=" not in text:
            raise ValueError("月份分组格式应为“列名=月份”或“列名=年份:月份”。")

        label_text, selector_text = (part.strip() for part in text.split("=", 1))
        if not label_text or not selector_text:
            raise ValueError("月份分组格式应为“列名=月份”或“列名=年份:月份”。")

        if ":" in selector_text:
            year_text, month_text = (part.strip() for part in selector_text.split(":", 1))
        else:
            year_text, month_text = str(default_year).strip(), selector_text

        if not year_text or not month_text:
            raise ValueError("月份分组格式应为“列名=月份”或“列名=年份:月份”。")
        if label_text in seen_labels:
            raise ValueError(f"月份分组列名重复：{label_text}")

        group_specs.append(
            SampleTableGroupSpec(
                label=label_text,
                year=normalize_year_value(year_text, default_year=default_year),
                month=normalize_month_value(month_text),
            )
        )
        seen_labels.add(label_text)

    return tuple(group_specs)


def load_source_dataframe(
    input_dir: Path,
    source_file_name: str,
    *,
    sheet_name: str,
    dataframe_cache: dict[Path, pd.DataFrame | None],
) -> pd.DataFrame | None:
    input_path = input_dir / source_file_name
    if input_path in dataframe_cache:
        return dataframe_cache[input_path]

    if not input_path.exists() or not input_path.is_file():
        dataframe_cache[input_path] = None
        return None

    dataframe_cache[input_path] = pd.read_excel(input_path, sheet_name=sheet_name)
    return dataframe_cache[input_path]


def excel_column_to_index(column_name: str) -> int:
    index = 0
    for char in column_name.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"非法 Excel 列名: {column_name}")
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def load_text_column(
    df: pd.DataFrame,
    column_name: str,
    *,
    column_label: str,
) -> pd.Series:
    column_index = excel_column_to_index(column_name)
    if column_index >= len(df.columns):
        raise ValueError(
            f"来源数据缺少{column_label}列 {column_name}，当前仅有 {len(df.columns)} 列。"
        )

    return (
        df.iloc[:, column_index]
        .astype("string")
        .fillna("")
        .str.strip()
    )


def normalize_expected_values(expected_value: str | tuple[str, ...]) -> tuple[str, ...]:
    if isinstance(expected_value, tuple):
        return tuple(str(value).strip() for value in expected_value if str(value).strip())
    text = str(expected_value).strip()
    return (text,) if text else ()


def build_customer_category_rule_mask(
    df: pd.DataFrame,
    rule: CustomerCategoryRule,
) -> pd.Series:
    if rule.data_column is None or not rule.data_values:
        return pd.Series([False] * len(df), index=df.index)

    mask = load_text_column(df, rule.data_column, column_label="数据").isin(
        normalize_expected_values(rule.data_values)
    )
    if rule.auxiliary_column and rule.auxiliary_values:
        mask = mask & load_text_column(df, rule.auxiliary_column, column_label="辅助").isin(
            normalize_expected_values(rule.auxiliary_values)
        )
    return mask


def normalize_output_dir(output_dir: Path) -> Path:
    if output_dir.exists() and output_dir.is_dir():
        return output_dir

    if output_dir.exists() and output_dir.is_file():
        return output_dir.parent / f"{output_dir.stem}_outputs"

    if output_dir.suffix:
        return output_dir.with_suffix("")

    return output_dir


def prepare_sample_table_rows(
    input_dir: Path,
    config: SampleTableConfig,
    *,
    source_sheet_name: str,
) -> tuple[PreparedSampleTableRow, ...]:
    dataframe_cache: dict[Path, pd.DataFrame | None] = {}
    prepared_rows: list[PreparedSampleTableRow] = []

    for row_config in config.rows:
        actual_count = row_config.actual_count_override if row_config.actual_count_override is not None else 0
        dataframe: pd.DataFrame | None = None
        mask: pd.Series | None = None

        if row_config.actual_count_override is None and row_config.rule_name is not None:
            rule = CUSTOMER_CATEGORY_RULE_BY_NAME.get(row_config.rule_name)
            if rule is None:
                raise ValueError(f"样本统计配置引用了未知映射规则：{row_config.rule_name}")

            dataframe = load_source_dataframe(
                input_dir,
                rule.source_file_name,
                sheet_name=source_sheet_name,
                dataframe_cache=dataframe_cache,
            )
            if dataframe is not None:
                mask = build_customer_category_rule_mask(dataframe, rule)
                actual_count = int(mask.sum())

        prepared_rows.append(
            PreparedSampleTableRow(
                row_config=row_config,
                actual_count=int(actual_count),
                dataframe=dataframe,
                mask=mask,
            )
        )

    return tuple(prepared_rows)


def resolve_sample_groups(
    prepared_rows: tuple[PreparedSampleTableRow, ...],
    *,
    sample_groups: tuple[SampleTableGroupSpec, ...] | None,
    default_year: str,
) -> tuple[SampleTableGroupSpec, ...]:
    if sample_groups:
        return sample_groups

    discovered_year_months: set[tuple[str, str]] = set()
    for row in prepared_rows:
        if row.dataframe is None or row.mask is None:
            continue
        discovered_year_months.update(
            iter_row_year_month_values(
                row.dataframe,
                row.mask,
                default_year=default_year,
            )
        )
    return build_auto_sample_groups(discovered_year_months, default_year=default_year)


def count_samples_by_group(
    df: pd.DataFrame | None,
    mask: pd.Series | None,
    *,
    sample_groups: tuple[SampleTableGroupSpec, ...],
    default_year: str,
) -> dict[str, int]:
    group_counts = {group.label: 0 for group in sample_groups}
    if df is None or mask is None or not sample_groups:
        return group_counts

    group_lookup = {(group.year, group.month): group.label for group in sample_groups}
    counter: Counter[str] = Counter()
    for year_text, month_text in iter_row_year_month_values(df, mask, default_year=default_year):
        matched_label = group_lookup.get((year_text, month_text))
        if matched_label is not None:
            counter[matched_label] += 1

    for label, count in counter.items():
        group_counts[label] = int(count)
    return group_counts


def build_sample_table_rows(
    input_dir: Path,
    config: SampleTableConfig | None = None,
    *,
    source_sheet_name: str = DEFAULT_SHEET_NAME,
    sample_groups: tuple[SampleTableGroupSpec, ...] | None = None,
    default_year: str = DEFAULT_GROUP_YEAR,
) -> SampleTableBuildResult:
    resolved_config = config or load_sample_table_config()
    prepared_rows = prepare_sample_table_rows(
        input_dir,
        resolved_config,
        source_sheet_name=source_sheet_name,
    )
    sample_groups = resolve_sample_groups(
        prepared_rows,
        sample_groups=sample_groups,
        default_year=default_year,
    )
    group_labels = tuple(group.label for group in sample_groups)
    row_results: list[SampleTableRowResult] = []

    for prepared_row in prepared_rows:
        row_config = prepared_row.row_config
        row_results.append(
            SampleTableRowResult(
                category_label=row_config.category_label,
                display_name=row_config.display_name,
                target_sample_size=row_config.target_sample_size,
                actual_count=prepared_row.actual_count,
                group_counts=count_samples_by_group(
                    prepared_row.dataframe,
                    prepared_row.mask,
                    sample_groups=sample_groups,
                    default_year=default_year,
                ),
            )
        )

    return SampleTableBuildResult(
        group_labels=group_labels,
        rows=tuple(row_results),
    )


def build_row_groups(rows: tuple[SampleTableRowResult, ...]) -> tuple[RowGroup, ...]:
    if not rows:
        return ()

    groups: list[RowGroup] = []
    current_category = rows[0].category_label
    current_rows: list[SampleTableRowResult] = []

    for row in rows:
        if row.category_label != current_category:
            groups.append(RowGroup(category_label=current_category, rows=tuple(current_rows)))
            current_category = row.category_label
            current_rows = []
        current_rows.append(row)

    groups.append(RowGroup(category_label=current_category, rows=tuple(current_rows)))
    return tuple(groups)


def set_count_cell(cell, value: int | str | None) -> None:
    cell.value = value
    apply_common_style(cell, fill=SUMMARY_BODY_FILL, font=SUMMARY_BODY_FONT)
    cell.number_format = COUNT_NUMBER_FORMAT


def set_percent_cell(cell, value: str) -> None:
    cell.value = value
    apply_common_style(cell, fill=SUMMARY_BODY_FILL, font=SUMMARY_BODY_FONT)
    cell.number_format = PERCENT_NUMBER_FORMAT


def merge_category_column(worksheet, start_row: int, end_row: int) -> None:
    if start_row >= end_row:
        return
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)


def ensure_output_path(output_dir: Path, output_name: str) -> Path:
    final_output_dir = normalize_output_dir(output_dir)
    file_name = output_name if output_name.lower().endswith(".xlsx") else f"{output_name}.xlsx"
    return final_output_dir / file_name


def build_sum_formula(column_index: int, row_ranges: list[tuple[int, int]]) -> str:
    column_letter = get_column_letter(column_index)
    joined_ranges = ",".join(
        f"{column_letter}{start_row}:{column_letter}{end_row}"
        for start_row, end_row in row_ranges
    )
    return f"=SUM({joined_ranges})"


def style_sample_table_worksheet(
    worksheet,
    result: SampleTableBuildResult,
    *,
    title: str,
) -> None:
    header_names = ("客户大类", "样本类型", "样本量", "样本进度百分比", "总执行样本量", *result.group_labels)
    last_column_index = len(header_names)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 36
    worksheet.row_dimensions[2].height = 28
    set_text_cell(worksheet["A1"], title, fill=SUMMARY_HEADER_FILL, font=SUMMARY_TITLE_FONT)
    for column_index in range(2, last_column_index + 1):
        apply_common_style(
            worksheet.cell(row=1, column=column_index),
            fill=SUMMARY_HEADER_FILL,
            font=SUMMARY_TITLE_FONT,
        )

    for column_index, column_name in enumerate(header_names, start=1):
        set_text_cell(
            worksheet.cell(row=2, column=column_index),
            column_name,
            fill=SUMMARY_HEADER_FILL,
            font=SUMMARY_HEADER_FONT,
        )

    data_row_ranges: list[tuple[int, int]] = []
    current_excel_row = 3

    for group in build_row_groups(result.rows):
        group_start_row = current_excel_row

        for row in group.rows:
            worksheet.row_dimensions[current_excel_row].height = 24
            set_text_cell(
                worksheet.cell(row=current_excel_row, column=1),
                row.category_label,
                fill=SUMMARY_SIDE_FILL,
                font=SUMMARY_SIDE_FONT,
            )
            set_text_cell(
                worksheet.cell(row=current_excel_row, column=2),
                row.display_name or None,
                fill=SUMMARY_SIDE_FILL,
                font=SUMMARY_SIDE_FONT,
            )
            set_count_cell(worksheet.cell(row=current_excel_row, column=3), row.target_sample_size)
            set_percent_cell(
                worksheet.cell(row=current_excel_row, column=4),
                f"=IFERROR(E{current_excel_row}/C{current_excel_row},0)",
            )
            set_count_cell(worksheet.cell(row=current_excel_row, column=5), row.actual_count)
            for column_offset, group_label in enumerate(result.group_labels, start=6):
                set_count_cell(
                    worksheet.cell(row=current_excel_row, column=column_offset),
                    row.group_counts.get(group_label, 0),
                )
            current_excel_row += 1

        group_end_row = current_excel_row - 1
        data_row_ranges.append((group_start_row, group_end_row))
        merge_category_column(worksheet, group_start_row, group_end_row)

        if len(group.rows) > 1:
            worksheet.row_dimensions[current_excel_row].height = 24
            set_text_cell(
                worksheet.cell(row=current_excel_row, column=1),
                None,
                fill=SUMMARY_SIDE_FILL,
                font=SUMMARY_SIDE_FONT,
            )
            set_text_cell(
                worksheet.cell(row=current_excel_row, column=2),
                "小计",
                fill=SUMMARY_SIDE_FILL,
                font=SUMMARY_SIDE_FONT,
            )
            set_count_cell(
                worksheet.cell(row=current_excel_row, column=3),
                f"=SUM(C{group_start_row}:C{group_end_row})",
            )
            set_percent_cell(
                worksheet.cell(row=current_excel_row, column=4),
                f"=IFERROR(E{current_excel_row}/C{current_excel_row},0)",
            )
            set_count_cell(
                worksheet.cell(row=current_excel_row, column=5),
                f"=SUM(E{group_start_row}:E{group_end_row})",
            )
            for column_offset in range(6, last_column_index + 1):
                set_count_cell(
                    worksheet.cell(row=current_excel_row, column=column_offset),
                    f"=SUM({get_column_letter(column_offset)}{group_start_row}:{get_column_letter(column_offset)}{group_end_row})",
                )
            current_excel_row += 1

    total_row_index = current_excel_row
    worksheet.row_dimensions[total_row_index].height = 28
    worksheet.merge_cells(start_row=total_row_index, start_column=1, end_row=total_row_index, end_column=2)
    set_text_cell(
        worksheet.cell(row=total_row_index, column=1),
        "合计",
        fill=SUMMARY_HEADER_FILL,
        font=SUMMARY_HEADER_FONT,
    )
    apply_common_style(
        worksheet.cell(row=total_row_index, column=2),
        fill=SUMMARY_HEADER_FILL,
        font=SUMMARY_HEADER_FONT,
    )

    set_count_cell(worksheet.cell(row=total_row_index, column=3), build_sum_formula(3, data_row_ranges))
    set_percent_cell(
        worksheet.cell(row=total_row_index, column=4),
        f"=IFERROR(E{total_row_index}/C{total_row_index},0)",
    )
    set_count_cell(worksheet.cell(row=total_row_index, column=5), build_sum_formula(5, data_row_ranges))
    for column_offset in range(6, last_column_index + 1):
        set_count_cell(
            worksheet.cell(row=total_row_index, column=column_offset),
            build_sum_formula(column_offset, data_row_ranges),
        )

    worksheet.freeze_panes = "C3"
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 30
    worksheet.column_dimensions["C"].width = 12
    worksheet.column_dimensions["D"].width = 16
    worksheet.column_dimensions["E"].width = 14
    for column_offset in range(6, last_column_index + 1):
        worksheet.column_dimensions[get_column_letter(column_offset)].width = 14


def generate_sample_table_report(
    input_dir: Path,
    output_dir: Path,
    *,
    output_name: str | None = None,
    config_path: Path = DEFAULT_SAMPLE_TABLE_CONFIG_PATH,
    source_sheet_name: str = DEFAULT_SHEET_NAME,
    sample_groups: tuple[SampleTableGroupSpec, ...] | None = None,
    default_year: str = DEFAULT_GROUP_YEAR,
) -> Path:
    config = load_sample_table_config(config_path)
    result = build_sample_table_rows(
        input_dir=input_dir,
        config=config,
        source_sheet_name=source_sheet_name,
        sample_groups=sample_groups,
        default_year=default_year,
    )
    output_path = ensure_output_path(output_dir, output_name or config.output_name)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = config.sheet_name
    style_sample_table_worksheet(worksheet, result, title=config.title)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按客户类型映射规则统计原始问卷样本量，生成样本统计表。")
    parser.add_argument("--input-dir", type=Path, required=True, help="输入目录：原始问卷 xlsx 所在目录")
    parser.add_argument("--output-dir", type=Path, required=True, help="输出目录")
    parser.add_argument("--output-name", help="输出文件名，默认读取配置中的 output_name")
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_SAMPLE_TABLE_CONFIG_PATH,
        help=f"样本统计配置文件，默认 {DEFAULT_SAMPLE_TABLE_CONFIG_PATH.name}",
    )
    parser.add_argument(
        "--source-sheet-name",
        default=DEFAULT_SHEET_NAME,
        help=f"原始数据 sheet 名称，默认 {DEFAULT_SHEET_NAME}",
    )
    parser.add_argument(
        "--default-year",
        default=DEFAULT_GROUP_YEAR,
        help=f"当原始数据缺少“{YEAR_HEADER}”列或为空时，用于月份分组匹配的默认年份，默认 {DEFAULT_GROUP_YEAR}",
    )
    parser.add_argument(
        "--month-group",
        action="append",
        help="月份分组，格式“列名=月份”或“列名=年份:月份”；可重复传入多次。",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    sample_groups = parse_sample_group_specs(args.month_group, default_year=str(args.default_year))
    output_path = generate_sample_table_report(
        input_dir=args.input_dir,
        output_dir=args.output_dir,
        output_name=args.output_name,
        config_path=args.config,
        source_sheet_name=args.source_sheet_name,
        sample_groups=sample_groups,
        default_year=str(args.default_year),
    )
    print(f"样本统计表已保存到: {output_path}")


if __name__ == "__main__":
    main()
